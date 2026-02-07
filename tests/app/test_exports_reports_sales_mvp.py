from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.order import Order, OrderItem, OrderSource, OrderStatus
from app.models.product import Product

pytestmark = pytest.mark.asyncio


def _make_order(*, company_id: int, ext_id: str, created_at: datetime) -> Order:
    return Order(
        company_id=company_id,
        order_number=f"SALE-{company_id}-{ext_id}",
        external_id=ext_id,
        source=OrderSource.MANUAL,
        status=OrderStatus.PAID,
        customer_name="Bob",
        customer_phone="+70000000000",
        total_amount=Decimal("1500.00"),
        currency="KZT",
        created_at=created_at.replace(tzinfo=None),
        updated_at=created_at.replace(tzinfo=None),
    )


async def _add_item(session, order_id: int, sku: str, qty: int = 1) -> None:
    session.add(
        OrderItem(
            order_id=order_id,
            sku=sku,
            name="Item",
            quantity=qty,
            unit_price=Decimal("500.00"),
            total_price=Decimal("500.00") * qty,
            cost_price=Decimal("100.00"),
        )
    )
    await session.commit()


def _make_product(*, company_id: int, sku: str, name: str) -> Product:
    return Product(
        company_id=company_id,
        sku=sku,
        name=name,
        price=Decimal("2500.00"),
    )


async def test_sales_export_xlsx(async_client, async_db_session, company_a_admin_headers):
    now = datetime.now(UTC)

    order_a1 = _make_order(company_id=1001, ext_id="s1", created_at=now)
    order_a2 = _make_order(company_id=1001, ext_id="s2", created_at=now)
    order_b1 = _make_order(company_id=2001, ext_id="s3", created_at=now)

    async_db_session.add_all([order_a1, order_a2, order_b1])
    await async_db_session.flush()

    await _add_item(async_db_session, order_a1.id, "SKU-A1", qty=2)
    await _add_item(async_db_session, order_a2.id, "SKU-A2", qty=1)
    await _add_item(async_db_session, order_b1.id, "SKU-B1", qty=1)

    resp = await async_client.get("/api/v1/exports/sales.xlsx", headers=company_a_admin_headers)
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers.get("content-type", "")

    wb = load_workbook(filename=BytesIO(resp.content))
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    assert headers == ["order_id", "created_at", "total_amount", "items_count"]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    order_ids = {int(row[0]) for row in rows}
    assert order_a1.id in order_ids
    assert order_a2.id in order_ids
    assert order_b1.id not in order_ids


async def test_products_export_xlsx(async_client, async_db_session, company_a_admin_headers):
    product_a1 = _make_product(company_id=1001, sku="P-A1", name="Product A1")
    product_a2 = _make_product(company_id=1001, sku="P-A2", name="Product A2")
    product_b1 = _make_product(company_id=2001, sku="P-B1", name="Product B1")

    async_db_session.add_all([product_a1, product_a2, product_b1])
    await async_db_session.commit()

    resp = await async_client.get("/api/v1/exports/products.xlsx", headers=company_a_admin_headers)
    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers.get("content-type", "")

    wb = load_workbook(filename=BytesIO(resp.content))
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    assert headers == ["product_id", "sku", "name", "price", "created_at"]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    product_ids = {int(row[0]) for row in rows}
    assert product_a1.id in product_ids
    assert product_a2.id in product_ids
    assert product_b1.id not in product_ids


async def test_sales_report_pdf(async_client, async_db_session, company_a_admin_headers):
    now = datetime.now(UTC)
    date_str = now.date().isoformat()

    order_a1 = _make_order(company_id=1001, ext_id="p1", created_at=now)
    order_a2 = _make_order(company_id=1001, ext_id="p2", created_at=now)
    order_b1 = _make_order(company_id=2001, ext_id="p3", created_at=now)

    async_db_session.add_all([order_a1, order_a2, order_b1])
    await async_db_session.flush()

    await _add_item(async_db_session, order_a1.id, "SKU-A1", qty=2)
    await _add_item(async_db_session, order_a2.id, "SKU-A2", qty=1)
    await _add_item(async_db_session, order_b1.id, "SKU-B1", qty=5)

    resp = await async_client.get(
        f"/api/v1/reports/sales.pdf?date_from={date_str}&date_to={date_str}",
        headers=company_a_admin_headers,
    )
    assert resp.status_code == 200
    assert resp.content.startswith(b"%PDF")
    assert "application/pdf" in resp.headers.get("content-type", "")

    text = resp.content.decode("latin1", errors="ignore")
    assert "Total orders:" in text
    assert "SKU-A1" in text
    assert "SKU-B1" not in text
