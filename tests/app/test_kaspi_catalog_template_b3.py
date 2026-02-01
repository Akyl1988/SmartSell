from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from app.api.v1.kaspi import KASPI_CATALOG_TEMPLATE_HEADERS

pytestmark = pytest.mark.asyncio


async def test_kaspi_catalog_template_defaults_to_xlsx(async_client, company_a_admin_headers):
    resp = await async_client.get(
        "/api/v1/kaspi/catalog/template",
        headers=company_a_admin_headers,
    )

    assert resp.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in resp.headers.get("content-type", "")
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == KASPI_CATALOG_TEMPLATE_HEADERS


async def test_kaspi_catalog_template_csv(async_client, company_a_admin_headers):
    resp = await async_client.get(
        "/api/v1/kaspi/catalog/template?format=csv",
        headers=company_a_admin_headers,
    )

    assert resp.status_code == 200
    assert "text/csv" in resp.headers.get("content-type", "")
    first_line = resp.text.splitlines()[0]
    assert first_line == ",".join(KASPI_CATALOG_TEMPLATE_HEADERS)


async def test_kaspi_catalog_template_xlsx(async_client, company_a_admin_headers):
    resp = await async_client.get(
        "/api/v1/kaspi/catalog/template?format=xlsx",
        headers=company_a_admin_headers,
    )

    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == KASPI_CATALOG_TEMPLATE_HEADERS
