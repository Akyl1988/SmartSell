from __future__ import annotations

"""
app/api/v1/kaspi.py — Полный, боевой роутер интеграции с Kaspi.

Что реализовано (по ТЗ и договорённостям):
- POST   /api/v1/kaspi/connect                — единая точка «подключить магазин» (verify/save).
- POST   /api/v1/kaspi/tokens                 — создать/обновить токен магазина (upsert).
- GET    /api/v1/kaspi/tokens                 — список подключённых магазинов (алиасы).
- GET    /api/v1/kaspi/tokens/{store_name}    — карточка токена (маска + метаданные, без расшифровки).
- DELETE /api/v1/kaspi/tokens/{store_name}    — удалить токен (безвозвратно).
- GET    /api/v1/kaspi/health/{store}         — health-проверка адаптера Kaspi для конкретного магазина.
- POST   /api/v1/kaspi/orders                 — получить заказы (через адаптер).
- POST   /api/v1/kaspi/import                 — запустить импорт офферов (фид) в Kaspi.
- POST   /api/v1/kaspi/import/status          — проверить статус импорта офферов.
- POST   /api/v1/kaspi/orders/sync            — синхронизировать свежие заказы Kaspi в локальную БД.
- GET    /api/v1/kaspi/feed      — сгенерировать XML-фид активных товаров компании.
- POST   /api/v1/kaspi/availability/sync      — синхронизировать доступность одного товара.
- POST   /api/v1/kaspi/availability/bulk      — массовая синхронизация доступности по компании.
- GET    /api/v1/kaspi/_debug/ping            — диагностический ping.

Принципы:
- Предсказуемые ответы: 4xx/5xx с внятным detail, без «просто 500».
- Pydantic v2, SQLAlchemy 2.x (AsyncSession).
- Безопасность: наружу не отдаём сырые токены, только маска + метаданные.
- Расширяемость: аккуратные модели ввода/вывода; готово к будущим эндпоинтам.
"""

import asyncio
import csv
import hashlib
import inspect
import io
import json
import os
import secrets
import time
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from email.utils import formatdate, parsedate_to_datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Literal
from uuid import UUID, uuid4
from xml.etree import ElementTree as ET

import httpx
import sqlalchemy as sa
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Request, Response, UploadFile, status
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sqlalchemy import bindparam, literal_column, select, text
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.kaspi_autosync_routes import (
    KaspiAutoSyncStatusOut,
    register_kaspi_autosync_routes,
)
from app.api.v1.kaspi_catalog_routes import register_kaspi_catalog_routes
from app.api.v1.kaspi_core_routes import register_kaspi_core_routes
from app.api.v1.kaspi_debug_routes import register_kaspi_debug_routes
from app.api.v1.kaspi_feed_routes import (
    register_kaspi_feed_routes_phase_one,
    register_kaspi_feed_routes_phase_two,
)
from app.api.v1.kaspi_goods_routes import register_kaspi_goods_routes
from app.api.v1.kaspi_mc_routes import register_kaspi_mc_routes
from app.api.v1.kaspi_orders_routes import register_kaspi_orders_routes
from app.api.v1.kaspi_public_routes import register_kaspi_public_routes
from app.api.v1.kaspi_status_routes import register_kaspi_status_routes
from app.api.v1.kaspi_sync_routes import register_kaspi_sync_routes
from app.api.v1.kaspi_tooling_routes import register_kaspi_tooling_routes
from app.api.v1.kaspi_utils import (
    build_kaspi_orders_params as _build_kaspi_orders_params,
)
from app.api.v1.kaspi_utils import (
    ci_diag_enabled as _ci_diag_enabled,
)
from app.api.v1.kaspi_utils import (
    classify_httpx_error as _classify_httpx_error,
)
from app.api.v1.kaspi_utils import (
    extract_httpx_root_cause as _extract_httpx_root_cause,
)
from app.api.v1.kaspi_utils import (
    extract_import_code as _extract_import_code,
)
from app.api.v1.kaspi_utils import (
    extract_schema_required_fields as _extract_schema_required_fields,
)
from app.api.v1.kaspi_utils import (
    extract_schema_types as _extract_schema_types,
)
from app.api.v1.kaspi_utils import (
    is_dev_environment as _is_dev_environment,
)
from app.api.v1.kaspi_utils import (
    kaspi_stub_enabled as _kaspi_stub_enabled,
)
from app.api.v1.kaspi_utils import (
    mask_secret as _mask_secret,
)
from app.api.v1.kaspi_utils import (
    probe_response_snippet as _probe_response_snippet,
)
from app.core.config import settings
from app.core.db import get_async_db  # noqa — для совместимости импорт-алиас
from app.core.dependencies import enforce_rate_limit, require_store_admin_company
from app.core.errors import safe_error_message
from app.core.exceptions import AuthorizationError
from app.core.logging import get_logger
from app.core.rbac import is_store_admin, require_platform_admin
from app.core.security import get_current_user, resolve_tenant_company_id
from app.core.subscriptions import (
    FEATURE_KASPI_AUTOSYNC,
    FEATURE_KASPI_FEED_UPLOADS,
    FEATURE_KASPI_GOODS_IMPORTS,
    FEATURE_KASPI_ORDERS_LIST,
    FEATURE_KASPI_SYNC_NOW,
    require_feature,
)
from app.integrations.kaspi_adapter import KaspiAdapter, KaspiAdapterError
from app.models import Product
from app.models.catalog_import import CatalogImportBatch, CatalogImportRow
from app.models.company import Company
from app.models.kaspi_catalog_item import KaspiCatalogItem
from app.models.kaspi_catalog_product import KaspiCatalogProduct
from app.models.kaspi_feed_export import KaspiFeedExport
from app.models.kaspi_feed_public_token import KaspiFeedPublicToken
from app.models.kaspi_feed_upload import KaspiFeedUpload
from app.models.kaspi_goods_import import KaspiGoodsImport
from app.models.kaspi_import_run import KaspiImportRun
from app.models.kaspi_mc_session import KaspiMcSession
from app.models.kaspi_offer import KaspiOffer
from app.models.kaspi_order_sync_state import KaspiOrderSyncState
from app.models.marketplace import KaspiStoreToken
from app.models.order import Order, OrderSource, OrderStatus, OrderStatusHistory
from app.models.user import User
from app.schemas.kaspi import (
    ImportRequest,
    ImportStatusQuery,
    KaspiConnectIn,
    KaspiConnectOut,
    KaspiTokenIn,
    KaspiTokenOut,
    OrdersQuery,
)
from app.schemas.kaspi_local_legacy import (
    AvailabilityBulkIn,
    AvailabilitySyncIn,
    KaspiMcSessionIn,
    KaspiMcSessionListOut,
    KaspiMcSessionOut,
    KaspiMcSyncOut,
    KaspiTokenMaskedOut,
)
from app.services.integration_events import record_integration_event
from app.services.kaspi_feed_upload_service import (
    compute_feed_payload_hash,
    compute_next_attempt_at,
    create_feed_upload_job,
    find_recent_successful_upload_by_hash,
    get_feed_upload_by_request_id,
    get_or_create_feed_export,
    is_unsupported_content_type_error,
    normalize_kaspi_payload,
    should_block_feed_upload_url,
    update_feed_upload_job,
)
from app.services.kaspi_goods_client import KaspiGoodsClient, KaspiNotAuthenticated
from app.services.kaspi_goods_import_client import (
    KaspiGoodsImportClient,
    KaspiImportNotAuthenticated,
    KaspiImportUpstreamError,
    KaspiImportUpstreamUnavailable,
)
from app.services.kaspi_goods_import_service import (
    build_payload_json,
    compute_payload_hash,
    load_offers_payload,
)
from app.services.kaspi_mc_sync import mark_mc_session_error, sync_kaspi_mc_offers
from app.services.kaspi_service import KaspiService, KaspiSyncAlreadyRunning
from app.services.otp_providers import is_otp_active, require_otp_provider_or_admin_bypass

logger = get_logger(__name__)
router = APIRouter(prefix="/api/v1/kaspi", tags=["kaspi"])
public_router = APIRouter(tags=["kaspi-public"])

__all__ = ["KaspiAutoSyncStatusOut", "public_router", "router"]


# ----------------------------- Константы/утилиты -----------------------------

MASK_HEX_LEN = 10
MASK_CHAR = "..."
STATUS_LAST_ERROR_MAX_LEN = 500
FAST_PROBE_TIMEOUT = 5.0
SYNC_NOW_TIMEOUT_SEC = 25.0


try:  # pragma: no cover - py<3.11 compatibility
    _ = ExceptionGroup  # type: ignore[name-defined]
    _HAS_EXCEPTION_GROUP = True
except NameError:  # pragma: no cover
    _HAS_EXCEPTION_GROUP = False


def _safe_log_info(event: str, **extra: Any) -> None:
    try:
        logger.info(event, extra=extra)
    except Exception:
        pass


def _safe_log_warning(event: str, **extra: Any) -> None:
    try:
        logger.warning(event, extra=extra)
    except Exception:
        pass


def _task_cancelled_flag() -> bool | None:
    try:
        task = asyncio.current_task()
        return bool(task.cancelled()) if task else None
    except Exception:
        return None


def _exception_group_info(exc: BaseException) -> dict[str, Any]:
    info: dict[str, Any] = {}
    if _HAS_EXCEPTION_GROUP and isinstance(exc, ExceptionGroup):  # type: ignore[name-defined]
        sub_exc = list(exc.exceptions)[:5]
        info["sub_exception_types"] = [type(e).__name__ for e in sub_exc]
        if _ci_diag_enabled():
            info["sub_exception_reprs"] = [repr(e) for e in sub_exc]
    return info


def _kaspi_user_agent() -> str:
    return (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) " "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
    )


def _build_kaspi_httpx_client() -> httpx.AsyncClient:
    total_timeout = max(60.0, float(getattr(settings, "KASPI_HTTP_TIMEOUT_SEC", 60) or 60))
    connect_timeout = max(20.0, float(getattr(settings, "KASPI_ORDERS_CONNECT_TIMEOUT_SEC", 20) or 20))
    timeout = httpx.Timeout(total_timeout, connect=connect_timeout)
    limits = httpx.Limits(max_connections=1, max_keepalive_connections=0)
    headers = {
        "Connection": "close",
        "User-Agent": _kaspi_user_agent(),
    }
    return httpx.AsyncClient(timeout=timeout, limits=limits, headers=headers, http2=False)


def _log_kaspi_probe_response(
    *,
    request_id: str | None,
    company_id: int | None,
    store_name: str,
    method: str,
    url: str,
    status_code: int | None,
    response_text: str | None,
    elapsed_ms: int | None,
) -> None:
    logger.info(
        "kaspi_probe_response",
        extra={
            "request_id": request_id,
            "company_id": company_id,
            "store_name": store_name,
            "method": method,
            "url": url,
            "status_code": status_code,
            "response_snippet": _probe_response_snippet(response_text),
            "elapsed_ms": elapsed_ms,
        },
    )


def _log_kaspi_probe_error(
    *,
    request_id: str | None,
    company_id: int | None,
    store_name: str,
    method: str,
    url: str,
    exc: Exception,
) -> None:
    root_type, root_message = _extract_httpx_root_cause(exc)
    logger.warning(
        "kaspi_probe_error",
        extra={
            "request_id": request_id,
            "company_id": company_id,
            "store_name": store_name,
            "method": method,
            "url": url,
            "error_class": type(exc).__name__,
            "error_kind": _classify_httpx_error(exc, root_type, root_message),
            "root_cause_type": root_type,
            "root_cause_message": root_message,
        },
    )


async def _maybe_await(value):
    if inspect.isawaitable(value):
        return await value
    return value


async def _auth_user(current_user: User = Depends(get_current_user)) -> User:
    return current_user


def _resolve_company_id(current_user: User) -> int:
    return resolve_tenant_company_id(current_user, not_found_detail="Company not set")


async def _require_store_admin_company_scoped(current_user: User) -> None:
    await require_store_admin_company(current_user)
    if not is_store_admin(current_user):
        raise AuthorizationError("Admin role required", "ADMIN_REQUIRED")


def require_store_admin_then_feature(feature: str) -> Any:
    async def _dep(
        request: Request,
        current_user: User = Depends(require_store_admin_company),  # noqa: B008
        db: AsyncSession = Depends(get_async_db),  # noqa: B008
    ) -> User:
        if not is_store_admin(current_user):
            raise AuthorizationError("Admin role required", "ADMIN_REQUIRED")
        resolve_tenant_company_id(current_user, not_found_detail="Company not set")
        await require_feature(feature)(request=request, current_user=current_user, db=db)
        return current_user

    return _dep


def _load_company_settings(company: Company | None) -> dict[str, Any]:
    if not company or not company.settings:
        return {}
    try:
        return json.loads(company.settings) or {}
    except (TypeError, json.JSONDecodeError):
        return {}


def _get_goods_import_flags(company: Company | None, merchant_uid: str) -> dict[str, bool]:
    settings_obj = _load_company_settings(company)
    defaults = {
        "include_price": bool(settings_obj.get("kaspi.goods_import.include_price", False)),
        "include_stock": bool(settings_obj.get("kaspi.goods_import.include_stock", False)),
    }
    merchant_map = settings_obj.get("kaspi.goods_import.merchants") or {}
    merchant_cfg = merchant_map.get(merchant_uid) if isinstance(merchant_map, dict) else None
    if isinstance(merchant_cfg, dict):
        defaults["include_price"] = bool(merchant_cfg.get("include_price", defaults["include_price"]))
        defaults["include_stock"] = bool(merchant_cfg.get("include_stock", defaults["include_stock"]))
    return defaults


def _sync_now_lock_key(company_id: int, merchant_uid: str) -> int:
    raw = f"kaspi-sync-now-{company_id}-{merchant_uid}".encode()
    h = int.from_bytes(hashlib.sha1(raw).digest()[:8], "big", signed=False)
    return h % (2**63 - 1)


async def _try_sync_now_lock(session: AsyncSession, *, company_id: int, merchant_uid: str) -> bool:
    lock_key = _sync_now_lock_key(company_id, merchant_uid)
    res = await session.execute(text("SELECT pg_try_advisory_lock(:lock_key)").bindparams(lock_key=lock_key))
    return bool(res.scalar_one_or_none())


async def _release_sync_now_lock(session: AsyncSession, *, company_id: int, merchant_uid: str) -> None:
    lock_key = _sync_now_lock_key(company_id, merchant_uid)
    await session.execute(text("SELECT pg_advisory_unlock(:lock_key)").bindparams(lock_key=lock_key))


async def _resolve_kaspi_token(session: AsyncSession, company_id: int) -> tuple[str, str]:
    company = (await session.execute(sa.select(Company).where(Company.id == company_id))).scalars().first()
    if not company:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")
    store_name = (company.kaspi_store_id or "").strip()
    if not store_name:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="kaspi_store_not_configured")
    token = await KaspiStoreToken.get_token(session, store_name)
    if not token:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="kaspi_token_not_found")
    return store_name, token


def _resolve_merchant_uid(
    *,
    company: Company | None,
    raw_merchant_uid: str | None,
    missing_status: int,
    missing_detail: str,
) -> tuple[str, str]:
    query_value = (raw_merchant_uid or "").strip()
    if query_value:
        return query_value, "query"
    fallback = (company.kaspi_store_id or "").strip() if company else ""
    if fallback:
        return fallback, "company"
    raise HTTPException(status_code=missing_status, detail=missing_detail)


def _resolve_company_store_uid(
    *,
    company: Company | None,
    raw_merchant_uid: str | None,
) -> tuple[str | None, str | None, str | None]:
    provided = (raw_merchant_uid or "").strip()
    store_uid = (company.kaspi_store_id or "").strip() if company else ""
    if provided:
        if not store_uid:
            return None, None, "missing_merchant_uid"
        if provided != store_uid:
            return None, None, "merchant_not_found"
        return provided, "query", None
    if store_uid:
        return store_uid, "company", None
    return None, None, "missing_merchant_uid"


def _resolve_kaspi_offers_merchant_uid(
    *,
    company: Company | None,
    raw_merchant_uid: str | None,
) -> str:
    store_uid, _source, error = _resolve_company_store_uid(company=company, raw_merchant_uid=raw_merchant_uid)
    if error == "missing_merchant_uid":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="kaspi_store_not_configured")
    if error == "merchant_not_found":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="merchant_not_found")
    if not store_uid:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="kaspi_store_not_configured")
    return store_uid


async def _build_offers_from_products(
    session: AsyncSession,
    *,
    company_id: int,
    merchant_uid: str,
    include_inactive: bool = False,
) -> list[dict[str, Any]]:
    conditions = [Product.company_id == company_id, Product.deleted_at.is_(None)]
    if not include_inactive:
        conditions.append(Product.is_active.is_(True))

    result = await session.execute(sa.select(Product).where(*conditions).order_by(Product.id.asc()))
    products = result.scalars().all()
    offers: list[dict[str, Any]] = []
    for product in products:
        sku = (product.sku or "").strip()
        if not sku:
            continue
        title = (product.name or "").strip() or sku
        price = product.sale_price or product.price or product.min_price or product.max_price
        stock_raw = int(product.stock_quantity or 0) - int(product.reserved_quantity or 0)
        stock_count = max(0, stock_raw)
        offers.append(
            {
                "company_id": company_id,
                "merchant_uid": merchant_uid,
                "sku": sku,
                "master_sku": None,
                "title": title,
                "price": price,
                "old_price": None,
                "stock_count": stock_count,
                "pre_order": bool(getattr(product, "is_preorder_enabled", False)),
                "stock_specified": True,
                "raw": {
                    "source": "products",
                    "product_id": product.id,
                },
            }
        )
    return offers


async def _upsert_kaspi_offers(
    session: AsyncSession,
    *,
    offers: list[dict[str, Any]],
) -> dict[str, int]:
    if not offers:
        return {"fetched": 0, "inserted": 0, "updated": 0}

    now = datetime.utcnow()
    stmt = insert(KaspiOffer).values(offers)
    update_values = {
        "master_sku": stmt.excluded.master_sku,
        "title": stmt.excluded.title,
        "price": stmt.excluded.price,
        "old_price": stmt.excluded.old_price,
        "stock_count": stmt.excluded.stock_count,
        "pre_order": stmt.excluded.pre_order,
        "stock_specified": stmt.excluded.stock_specified,
        "raw": stmt.excluded.raw,
        "updated_at": now,
    }
    stmt = stmt.on_conflict_do_update(
        index_elements=["company_id", "merchant_uid", "sku"],
        set_=update_values,
    ).returning(KaspiOffer.id, literal_column("xmax = 0").label("inserted"))
    result = await session.execute(stmt)
    rows = result.fetchall()
    inserted = sum(1 for row in rows if row.inserted)
    updated = len(rows) - inserted
    return {"fetched": len(offers), "inserted": inserted, "updated": updated}


async def _create_import_run(
    session: AsyncSession,
    *,
    company_id: int,
    merchant_uid: str,
    request_id: str | None,
) -> KaspiImportRun:
    run_code = uuid4().hex[:16]
    run = KaspiImportRun(
        company_id=company_id,
        merchant_uid=merchant_uid,
        import_code=run_code,
        status="created",
        request_id=request_id,
    )
    session.add(run)
    await session.commit()
    await session.refresh(run)
    return run


async def _get_import_run(
    session: AsyncSession,
    *,
    company_id: int,
    import_code: str,
) -> KaspiImportRun | None:
    result = await session.execute(
        sa.select(KaspiImportRun).where(
            KaspiImportRun.company_id == company_id,
            KaspiImportRun.import_code == import_code,
        )
    )
    return result.scalars().first()


async def _find_recent_import_run_by_hash(
    session: AsyncSession,
    *,
    company_id: int,
    merchant_uid: str,
    payload_hash: str,
    window_hours: int = 24,
) -> KaspiImportRun | None:
    cutoff = datetime.utcnow() - timedelta(hours=window_hours)
    result = await session.execute(
        sa.select(KaspiImportRun)
        .where(
            KaspiImportRun.company_id == company_id,
            KaspiImportRun.merchant_uid == merchant_uid,
            KaspiImportRun.payload_hash == payload_hash,
            KaspiImportRun.created_at >= cutoff,
        )
        .order_by(KaspiImportRun.created_at.desc())
        .limit(1)
    )
    return result.scalars().first()


def _validate_payload_against_schema(
    payload: list[dict[str, Any]],
    schema: dict[str, Any],
    *,
    max_errors: int = 10,
) -> list[dict[str, Any]]:
    required = _extract_schema_required_fields(schema)
    types = _extract_schema_types(schema)
    additional_allowed = schema.get("additionalProperties") is not False
    allowed_fields = set(types.keys()) | set(required)
    errors: list[dict[str, Any]] = []

    def _sku(item: dict[str, Any]) -> str:
        return str(item.get("sku") or item.get("offerId") or item.get("offer_id") or "unknown")

    for item in payload:
        if len(errors) >= max_errors:
            break
        if not additional_allowed:
            for key in item.keys():
                if key not in allowed_fields:
                    errors.append(
                        {
                            "code": "schema_additional_property",
                            "detail": "additional_property_not_allowed",
                            "field": key,
                            "sku": _sku(item),
                        }
                    )
                    if len(errors) >= max_errors:
                        break
        for field in required:
            value = item.get(field)
            if value is None or (isinstance(value, str) and not value.strip()):
                errors.append(
                    {
                        "code": "schema_missing_required",
                        "detail": "missing_required_field",
                        "field": field,
                        "sku": _sku(item),
                    }
                )
                if len(errors) >= max_errors:
                    break
        if len(errors) >= max_errors:
            break
        for field, expected in types.items():
            if len(errors) >= max_errors:
                break
            if field not in item:
                continue
            value = item.get(field)
            if value is None:
                continue
            if expected == "string" and not isinstance(value, str):
                errors.append(
                    {
                        "code": "schema_type_mismatch",
                        "detail": "expected_string",
                        "field": field,
                        "sku": _sku(item),
                    }
                )
            elif expected in {"number", "integer"} and not isinstance(value, int | float | Decimal):
                errors.append(
                    {
                        "code": "schema_type_mismatch",
                        "detail": "expected_number",
                        "field": field,
                        "sku": _sku(item),
                    }
                )
            elif expected == "boolean" and not isinstance(value, bool):
                errors.append(
                    {
                        "code": "schema_type_mismatch",
                        "detail": "expected_boolean",
                        "field": field,
                        "sku": _sku(item),
                    }
                )
    return errors


def _product_to_goods_payload(product: Product) -> dict[str, Any]:
    sku = product.sku or f"PID-{product.id}"
    return {
        "sku": sku,
        "name": product.name or sku,
        "price": float(product.price) if product.price is not None else None,
        "quantity": int(product.stock_quantity or 0),
        "isActive": bool(product.is_active),
    }


def _norm_header(name: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in (name or "").strip().lower()).strip("_")


_CATALOG_HEADER_ALIASES: dict[str, list[str]] = {
    "sku": ["sku", "offer_id", "offerid", "offersku", "merchant_sku"],
    "masterSku": ["master_sku", "mastersku", "master sku", "parent_sku"],
    "title": ["title", "name", "product_name"],
    "price": ["price", "price_kzt", "price_kz", "minprice", "maxprice"],
    "oldprice": ["oldprice", "old_price", "oldprice_kzt", "oldprice_kz", "price_old"],
    "stockCount": ["stock", "stock_count", "stockcount", "qty", "quantity"],
    "preOrder": ["preorder", "pre_order", "pre_order_flag", "preorder_flag"],
    "images": ["images", "image", "image_urls", "pictures"],
    "attributes": ["attributes", "attrs"],
    "stock_specified": ["stock_specified", "stockspecified", "stock_flag", "stockspecified_flag"],
    "updated_at": ["updated_at", "updated", "last_update", "updatedat"],
}


def _parse_int(value: str | None) -> int | None:
    if value is None:
        return None
    v = str(value).strip().replace(" ", "")
    if not v:
        return None
    v = v.replace(",", ".")
    try:
        return int(float(v))
    except ValueError:
        return None


def _parse_bool(value: str | None) -> bool | None:
    if value is None:
        return None
    v = str(value).strip().lower()
    if v in {"1", "true", "yes", "y", "on"}:
        return True
    if v in {"0", "false", "no", "n", "off"}:
        return False
    return None


def _get_json_value(normalized_raw: dict[str, Any], keys: list[str]) -> Any | None:
    for key in keys:
        value = normalized_raw.get(_norm_header(key))
        if value is None:
            continue
        if isinstance(value, str):
            stripped = value.strip()
            if stripped == "":
                continue
            return stripped
        return value
    return None


def _select_alias_value(normalized_raw: dict[str, Any], aliases: list[str]) -> Any | None:
    for alias in aliases:
        value = _get_json_value(normalized_raw, [alias])
        if value is not None:
            return value
    return None


def _normalize_catalog_row(raw: dict[str, Any]) -> dict[str, Any]:
    normalized_raw = {_norm_header(str(k)): v for k, v in raw.items()}

    def clean_str(value: Any | None) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def pick(key: str) -> Any | None:
        return _select_alias_value(normalized_raw, _CATALOG_HEADER_ALIASES.get(key, []))

    sku = pick("sku")
    master_sku = pick("masterSku")
    title = pick("title")
    price = _parse_int(pick("price"))
    old_price = _parse_int(pick("oldprice"))
    stock_count = _parse_int(pick("stockCount"))
    pre_order = _parse_bool(pick("preOrder"))
    stock_specified = _parse_bool(pick("stock_specified"))
    images = pick("images")
    attributes = pick("attributes")
    updated_at_raw = pick("updated_at")

    updated_at = None
    if updated_at_raw:
        try:
            updated_at = datetime.fromisoformat(str(updated_at_raw).replace("Z", ""))
        except Exception:
            updated_at = None

    return {
        "raw": raw,
        "sku": clean_str(sku) if isinstance(sku, str) or sku is not None else None,
        "master_sku": clean_str(master_sku) if isinstance(master_sku, str) or master_sku is not None else None,
        "title": clean_str(title) if isinstance(title, str) or title is not None else None,
        "price": price,
        "old_price": old_price,
        "stock_count": stock_count,
        "pre_order": pre_order,
        "stock_specified": stock_specified,
        "images": images,
        "attributes": attributes,
        "updated_at": updated_at,
    }


def _parse_csv_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig", errors="replace")
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_encoding")

    try:
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ","

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_headers")

    rows: list[dict[str, Any]] = []
    for raw in reader:
        if raw is None:
            continue
        if all((str(v).strip() == "" if v is not None else True) for v in raw.values()):
            continue
        rows.append(raw)
    return rows


def _parse_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_xlsx")

    sheet = wb.worksheets[0]
    header_row: list[str] | None = None
    rows: list[dict[str, Any]] = []

    for row in sheet.iter_rows(values_only=True):
        values = [str(c).strip() if c is not None else "" for c in row]
        if header_row is None:
            if any(values):
                header_row = values
            continue

        if not any(values):
            continue

        raw: dict[str, Any] = {}
        for idx, header in enumerate(header_row):
            if not header:
                continue
            raw[header] = row[idx] if idx < len(row) else None
        rows.append(raw)

    if header_row is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_headers")

    return rows


def _parse_json_rows(content: bytes, is_jsonl: bool) -> list[dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig", errors="replace")
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_encoding")

    rows: list[dict[str, Any]] = []
    if is_jsonl:
        for line in text.splitlines():
            if not line.strip():
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_json")
            if not isinstance(payload, dict):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_json_row")
            rows.append(payload)
        return rows

    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_json")

    if isinstance(payload, dict):
        payload = payload.get("data") or payload.get("items") or payload.get("rows") or payload

    if not isinstance(payload, list):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_json_payload")

    for item in payload:
        if not isinstance(item, dict):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_json_row")
        rows.append(item)
    return rows


def parse_catalog_file(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    name = (filename or "").lower()

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        rows = _parse_xlsx_rows(file_bytes)
    elif name.endswith(".csv"):
        rows = _parse_csv_rows(file_bytes)
    elif name.endswith(".jsonl"):
        rows = _parse_json_rows(file_bytes, is_jsonl=True)
    elif name.endswith(".json"):
        rows = _parse_json_rows(file_bytes, is_jsonl=False)
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="unsupported_file_type")

    return [_normalize_catalog_row(raw) for raw in rows]


def _truncate_raw(raw: Any, max_len: int = 2000) -> str | None:
    if raw is None:
        return None
    try:
        text_value = json.dumps(raw, ensure_ascii=False, default=str)
    except Exception:
        text_value = str(raw)
    if len(text_value) <= max_len:
        return text_value
    return f"{text_value[:max_len]}..."


def _normalize_kaspi_response(payload: Any) -> dict[str, Any]:
    if isinstance(payload, dict):
        return payload
    if isinstance(payload, list):
        return {"data": payload}
    if payload is None:
        return {}
    return {"raw": payload}


def _extract_error_info(payload: Any) -> tuple[str | None, str | None]:
    if not isinstance(payload, dict):
        return None, None
    code = payload.get("errorCode") or payload.get("error_code") or payload.get("code") or payload.get("error")
    message = payload.get("errorMessage") or payload.get("error_message") or payload.get("message")
    return (str(code) if code else None), (str(message) if message else None)


def _serialize_raw_response(payload: Any) -> str | None:
    if payload is None:
        return None
    if isinstance(payload, dict | list):
        try:
            return json.dumps(payload, ensure_ascii=False, default=str)
        except Exception:
            return str(payload)
    return str(payload)


_KASPI_NS = "kaspiShopping"
ET.register_namespace("", _KASPI_NS)


def _to_unsigned_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        num = int(float(value))
    except Exception:
        return None
    return max(num, 0)


def _format_price(value: Any) -> str:
    try:
        dec = Decimal(str(value if value is not None else 0))
    except Exception:
        dec = Decimal("0")
    try:
        dec = dec.quantize(Decimal("0.01"))
    except Exception:
        dec = Decimal("0.00")
    return str(dec)


def _extract_city_prices(raw: Any) -> list[dict[str, str]]:
    if not isinstance(raw, dict):
        return []
    city_prices = raw.get("cityPrices") or raw.get("cityprices")
    items: list[dict[str, str]] = []
    if isinstance(city_prices, list):
        for entry in city_prices:
            if not isinstance(entry, dict):
                continue
            city_id = entry.get("cityId") or entry.get("city_id")
            price = _to_unsigned_int(entry.get("value") or entry.get("price"))
            if city_id is None or price is None:
                continue
            oldprice = _to_unsigned_int(entry.get("oldprice") or entry.get("oldPrice"))
            item = {"cityId": str(city_id), "value": str(price)}
            if oldprice is not None:
                item["oldprice"] = str(oldprice)
            items.append(item)
    elif isinstance(city_prices, dict):
        if "cityId" in city_prices:
            city_id = city_prices.get("cityId")
            price = _to_unsigned_int(city_prices.get("value") or city_prices.get("price"))
            if city_id is not None and price is not None:
                oldprice = _to_unsigned_int(city_prices.get("oldprice") or city_prices.get("oldPrice"))
                item = {"cityId": str(city_id), "value": str(price)}
                if oldprice is not None:
                    item["oldprice"] = str(oldprice)
                items.append(item)
        else:
            for city_id, price_val in city_prices.items():
                price = _to_unsigned_int(price_val)
                if city_id is None or price is None:
                    continue
                items.append({"cityId": str(city_id), "value": str(price)})
    return items


def _build_kaspi_offers_xml(offers: list[KaspiOffer], *, company: str, merchant_id: str) -> str:
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    root = ET.Element(f"{{{_KASPI_NS}}}kaspi_catalog", {"date": date_str})
    company_el = ET.SubElement(root, f"{{{_KASPI_NS}}}company")
    company_el.text = str(company)
    merchant_el = ET.SubElement(root, f"{{{_KASPI_NS}}}merchantid")
    merchant_el.text = str(merchant_id)
    offers_el = ET.SubElement(root, f"{{{_KASPI_NS}}}offers")

    for offer in offers:
        sku = (offer.sku or "").strip()
        if not sku:
            continue
        model_text = (offer.title or offer.master_sku or sku).strip() or sku
        offer_el = ET.SubElement(offers_el, f"{{{_KASPI_NS}}}offer", {"sku": sku})
        model_el = ET.SubElement(offer_el, f"{{{_KASPI_NS}}}model")
        model_el.text = model_text

        raw = offer.raw or {}
        city_prices = _extract_city_prices(raw)
        if city_prices:
            cityprices_el = ET.SubElement(offer_el, f"{{{_KASPI_NS}}}cityprices")
            for entry in city_prices:
                attrs = {"cityId": entry["cityId"]}
                if "oldprice" in entry:
                    attrs["oldprice"] = entry["oldprice"]
                city_el = ET.SubElement(cityprices_el, f"{{{_KASPI_NS}}}cityprice", attrs)
                city_el.text = entry["value"]
        else:
            price_value = _format_price(offer.price)
            attrs: dict[str, str] = {}
            if offer.old_price is not None:
                attrs["oldprice"] = _format_price(offer.old_price)
            price_el = ET.SubElement(offer_el, f"{{{_KASPI_NS}}}price", attrs)
            price_el.text = price_value

    xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return xml_bytes.decode("utf-8")


def _goods_import_to_out(record: KaspiGoodsImport) -> KaspiGoodsImportRecordOut:
    return KaspiGoodsImportRecordOut(
        id=str(record.id),
        merchant_uid=record.merchant_uid,
        import_code=record.import_code,
        status=record.status,
        source=record.source,
        payload_hash=record.payload_hash,
        attempts=record.attempts,
        request_json=record.request_json,
        status_json=record.status_json,
        raw_status_json=record.raw_status_json,
        result_json=record.result_json,
        error_code=record.error_code,
        error_message=record.error_message,
        created_at=record.created_at,
        updated_at=record.updated_at,
        last_checked_at=record.last_checked_at,
        revoked_at=record.revoked_at,
    )


def _feed_upload_to_out(record: KaspiFeedUpload) -> KaspiFeedUploadRecordOut:
    return KaspiFeedUploadRecordOut(
        id=str(record.id),
        merchant_uid=record.merchant_uid,
        import_code=record.import_code,
        status=record.status,
        source=record.source,
        comment=record.comment,
        attempts=record.attempts or 0,
        last_error_code=record.last_error_code,
        last_error_message=record.last_error_message,
        request_id=record.request_id,
        payload_hash=record.payload_hash,
        response_json=record.response_json,
        next_attempt_at=record.next_attempt_at,
        created_at=record.created_at,
        updated_at=record.updated_at,
    )


def _feed_export_to_out(export: KaspiFeedExport) -> KaspiFeedExportOut:
    return KaspiFeedExportOut(
        id=export.id,
        kind=export.kind,
        format=export.format,
        status=export.status,
        checksum=export.checksum,
        stats_json=export.stats_json,
        last_error=export.last_error,
        attempts=export.attempts or 0,
        last_attempt_at=export.last_attempt_at.isoformat() if export.last_attempt_at else None,
        uploaded_at=export.uploaded_at.isoformat() if export.uploaded_at else None,
        duration_ms=export.duration_ms,
        created_at=export.created_at.isoformat() if export.created_at else None,
        updated_at=export.updated_at.isoformat() if export.updated_at else None,
    )


# ------------------------------- Локальные схемы (deprecated - use app/schemas/kaspi.py) ------


async def _record_kaspi_event(
    session: AsyncSession,
    *,
    company_id: int,
    kind: str,
    status: str,
    request_id: str | None,
    merchant_uid: str | None = None,
    error_code: str | None = None,
    error_message: str | None = None,
    meta_json: dict[str, Any] | None = None,
    commit: bool = True,
) -> None:
    try:
        await record_integration_event(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            kind=kind,
            status=status,
            error_code=error_code,
            error_message=error_message,
            request_id=request_id,
            meta_json=meta_json,
            commit=commit,
        )
    except Exception as exc:
        logger.warning("Kaspi integration event write failed: kind=%s err=%s", kind, exc)


async def connect_store(
    request: Request,
    body: KaspiConnectIn,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    """
    Main Kaspi onboarding endpoint:
    1) Resolve tenant company from current user.
    2) Validate and optionally verify token with Kaspi HTTP API (NOT PowerShell adapter).
    3) Update Company.name with provided company_name.
    4) Store encrypted token via KaspiStoreToken.upsert_token().
    5) Store optional private metadata in Company.settings JSON.
    6) Return only safe fields (no token, no metadata exposed).

    Requires: company_name (min_length=2).
    Optional: meta (private marketplace metadata, not exposed).

    Token verification (verify=true):
    - Uses KaspiService.verify_token() which makes minimal HTTP call to Kaspi API
    - 401/403 -> returns 400 with detail="kaspi_invalid_token"
    - Network/timeout errors -> returns 502 with detail="kaspi_upstream_error"
    - Never calls KaspiAdapter (no PowerShell dependency)
    """
    # Resolve company from current user (tenant isolation)
    company_id = _resolve_company_id(current_user)

    # Load Company
    result = await session.execute(sa.select(Company).where(Company.id == company_id))
    company = result.scalars().first()
    if not company:
        logger.error("Kaspi connect: company not found id=%s", company_id)
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found",
        )

    if not is_otp_active():
        require_otp_provider_or_admin_bypass(
            current_user,
            action="kaspi_connect",
            company_id=company_id,
            owner_id=company.owner_id,
        )

    # Verify token if requested (before persisting) - HTTP only, no PowerShell
    if body.verify:
        try:
            logger.info("Kaspi connect: verifying token via HTTP for store=%s", body.store_name)
            kaspi_service = KaspiService()
            await kaspi_service.verify_token(store_name=body.store_name, token=body.token)
            logger.info("Kaspi connect: HTTP verification succeeded for store=%s", body.store_name)
        except Exception as e:
            # Handle different error types
            import httpx

            if isinstance(e, httpx.HTTPStatusError):
                if e.response.status_code in (401, 403):
                    # Invalid token
                    logger.warning(
                        "Kaspi connect: invalid token store=%s status=%s", body.store_name, e.response.status_code
                    )
                    await _record_kaspi_event(
                        session,
                        company_id=company_id,
                        kind="kaspi_connect",
                        status="failed",
                        request_id=getattr(getattr(request, "state", None), "request_id", None),
                        error_code="kaspi_invalid_token",
                        error_message="NOT_AUTHENTICATED",
                        meta_json={"store_name": body.store_name, "verify": body.verify},
                    )
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="kaspi_invalid_token",
                    )
                else:
                    # Other HTTP errors are upstream problems
                    logger.error(
                        "Kaspi connect: upstream HTTP error store=%s status=%s", body.store_name, e.response.status_code
                    )
                    await _record_kaspi_event(
                        session,
                        company_id=company_id,
                        kind="kaspi_connect",
                        status="failed",
                        request_id=getattr(getattr(request, "state", None), "request_id", None),
                        error_code="kaspi_upstream_error",
                        error_message="upstream_error",
                        meta_json={"store_name": body.store_name, "verify": body.verify},
                    )
                    raise HTTPException(
                        status_code=status.HTTP_502_BAD_GATEWAY,
                        detail="kaspi_upstream_error",
                    )
            elif isinstance(e, httpx.TimeoutException | httpx.NetworkError):
                # Network/timeout errors
                logger.error("Kaspi connect: network error store=%s error=%s", body.store_name, type(e).__name__)
                await _record_kaspi_event(
                    session,
                    company_id=company_id,
                    kind="kaspi_connect",
                    status="failed",
                    request_id=getattr(getattr(request, "state", None), "request_id", None),
                    error_code="kaspi_upstream_unavailable",
                    error_message="upstream_unavailable",
                    meta_json={"store_name": body.store_name, "verify": body.verify},
                )
                raise HTTPException(
                    status_code=status.HTTP_502_BAD_GATEWAY,
                    detail="kaspi_upstream_error",
                )
            else:
                # Unexpected errors
                logger.error("Kaspi connect: verification error store=%s error=%s", body.store_name, str(e))
                await _record_kaspi_event(
                    session,
                    company_id=company_id,
                    kind="kaspi_connect",
                    status="failed",
                    request_id=getattr(getattr(request, "state", None), "request_id", None),
                    error_code="kaspi_upstream_error",
                    error_message=str(e),
                    meta_json={"store_name": body.store_name, "verify": body.verify},
                )
                raise HTTPException(
                    status_code=status.HTTP_502_BAD_GATEWAY,
                    detail="kaspi_upstream_error",
                )

    # Update Company with provided company_name
    company.name = body.company_name.strip()

    requested_store = body.store_name.strip()
    if requested_store:
        conflict = await session.execute(
            sa.select(Company.id).where(Company.kaspi_store_id == requested_store, Company.id != company_id)
        )
        conflict_id = conflict.scalar_one_or_none()
        if conflict_id is not None:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="kaspi_store_id_conflict")

        if not company.kaspi_store_id:
            company.kaspi_store_id = requested_store
        elif company.kaspi_store_id != requested_store:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="kaspi_store_id_conflict")

    # Store private metadata in Company.settings if provided
    if body.meta:
        try:
            settings_dict = {}
            if company.settings:
                try:
                    settings_dict = json.loads(company.settings)
                except (json.JSONDecodeError, TypeError):
                    settings_dict = {}
            settings_dict["kaspi_meta"] = body.meta
            company.settings = json.dumps(settings_dict)
            logger.debug("Kaspi connect: stored private metadata for company_id=%s", company_id)
        except Exception as e:
            logger.warning("Kaspi connect: failed to store metadata company_id=%s error=%s", company_id, e)
            # Don't fail the entire request if metadata storage fails

    # Upsert encrypted token (never expose plaintext)
    try:
        logger.info("Kaspi connect: upserting token for store=%s", body.store_name)
        await KaspiStoreToken.upsert_token(session, body.store_name, body.token)
        logger.info("Kaspi connect: token upserted for store=%s company_id=%s", body.store_name, company_id)
    except Exception as e:
        logger.error("Kaspi connect: token upsert failed store=%s error=%s", body.store_name, e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save token: {str(e)}",
        )

    # Commit all changes in single transaction
    try:
        await session.commit()
        logger.info("Kaspi connect: transaction committed company_id=%s store=%s", company_id, body.store_name)
    except Exception as e:
        await session.rollback()
        logger.error("Kaspi connect: commit failed company_id=%s error=%s", company_id, e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save configuration: {str(e)}",
        )

    await _record_kaspi_event(
        session,
        company_id=company_id,
        kind="kaspi_connect",
        status="success",
        request_id=getattr(getattr(request, "state", None), "request_id", None),
        meta_json={"store_name": body.store_name, "verify": body.verify},
    )

    return KaspiConnectOut(
        store_name=body.store_name,
        company_id=company_id,
        connected=True,
        message="Successfully connected to Kaspi store",
    )


# ================================= TOKENS ====================================


async def upsert_token(
    payload: KaspiTokenIn,
    session: AsyncSession = Depends(get_async_db),
):
    try:
        await KaspiStoreToken.upsert_token(session, payload.store_name, payload.token)
    except Exception as e:
        logger.error("Kaspi upsert_token failed: store=%s err=%s", payload.store_name, e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))
    return KaspiTokenOut(store_name=payload.store_name)


async def list_tokens(session: AsyncSession = Depends(get_async_db)):
    try:
        stores = await KaspiStoreToken.list_stores(session)
    except Exception as e:
        logger.error("Kaspi list_tokens failed: %s", e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))
    return [KaspiTokenOut(store_name=s) for s in stores]


async def get_token_by_store_name(
    store_name: str,
    session: AsyncSession = Depends(get_async_db),
):
    """
    Возвращает запись токена по имени магазина.
    Токен НЕ раскрываем — только маска (первые MASK_HEX_LEN hex-символов) и метаданные.
    """
    q = text(
        """
        SELECT
            id,
            store_name,
            left(encode(token_ciphertext,'hex'), :mask_len) || :mask_char AS token_hex_masked,
            created_at,
            updated_at,
            last_selftest_at,
            last_selftest_status,
            last_selftest_error_code,
            last_selftest_error_message
        FROM kaspi_store_tokens
        WHERE lower(trim(store_name)) = lower(trim(:name))
        LIMIT 1
        """
    ).bindparams(
        bindparam("mask_len", type_=sa.Integer),
        bindparam("mask_char", type_=sa.String),
        bindparam("name", type_=sa.String),
    )

    try:
        res = await session.execute(q, {"name": store_name, "mask_len": MASK_HEX_LEN, "mask_char": MASK_CHAR})
        row = res.mappings().first()
    except Exception as e:
        logger.error("Kaspi get_token_by_store_name failed: store=%s err=%s", store_name, e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="db_error")

    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Not Found")

    return KaspiTokenMaskedOut(
        id=str(row["id"]),
        store_name=row["store_name"],
        token_hex_masked=row["token_hex_masked"],
        created_at=row["created_at"],
        updated_at=row["updated_at"],
        last_selftest_at=row.get("last_selftest_at"),
        last_selftest_status=row.get("last_selftest_status"),
        last_selftest_error_code=row.get("last_selftest_error_code"),
        last_selftest_error_message=row.get("last_selftest_error_message"),
    )


async def delete_token(store_name: str, session: AsyncSession = Depends(get_async_db)):
    try:
        deleted = await KaspiStoreToken.delete_by_store(session, store_name)
        if not deleted:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Not Found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Kaspi delete_token failed: store=%s err=%s", store_name, e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ============================ Операции через адаптер ==========================


async def kaspi_health(store: str):
    try:
        return KaspiAdapter().health(store)
    except KaspiAdapterError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))
    except Exception as e:
        logger.error("Kaspi health unexpected error: store=%s err=%s", store, e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


register_kaspi_core_routes(
    router,
    connect_store=connect_store,
    upsert_token=upsert_token,
    list_tokens=list_tokens,
    get_token_by_store_name=get_token_by_store_name,
    delete_token=delete_token,
    kaspi_health=kaspi_health,
    kaspi_connect_out_model=KaspiConnectOut,
    kaspi_token_out_model=KaspiTokenOut,
    kaspi_token_masked_out_model=KaspiTokenMaskedOut,
)


class KaspiOrderEntryOut(BaseModel):
    id: int
    sku: str
    name: str
    quantity: int
    unit_price: Decimal
    total_price: Decimal


class KaspiOrderListItemOut(BaseModel):
    id: int
    external_id: str | None = None
    order_number: str
    status: str
    created_at: datetime
    updated_at: datetime
    total_amount: Decimal
    currency: str
    customer_name: str | None = None
    customer_phone: str | None = None
    delivery_date: str | None = None
    kaspi_preorder: bool | None = None
    kaspi_planned_delivery_date: str | None = None
    kaspi_reservation_date: str | None = None


class KaspiOrderDetailOut(KaspiOrderListItemOut):
    items: list[KaspiOrderEntryOut]


class KaspiOrdersListOut(BaseModel):
    items: list[KaspiOrderListItemOut]
    page: int
    limit: int
    total: int
    offset: int = 0


class KaspiOrderActionOut(BaseModel):
    ok: bool
    status: str
    code: str | None = None
    message: str | None = None
    request_id: str | None = None
    upstream_status_code: int | None = None
    retry_after: int | None = None


def _parse_iso_dt(value: str) -> datetime:
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_datetime") from exc


def _to_ms(value: datetime) -> int:
    if value.tzinfo is None:
        return int(value.timestamp() * 1000)
    return int(value.astimezone(tz=UTC).timestamp() * 1000)


def _parse_order_created_at(raw: Any) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, int | float):
        return datetime.fromtimestamp(int(raw) / 1000.0, tz=UTC)
    if isinstance(raw, str):
        try:
            return _parse_iso_dt(raw)
        except Exception:
            return None
    return None


def _status_to_str(value: OrderStatus | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, OrderStatus):
        return value.value
    return str(value)


def _parse_order_status(raw: str) -> OrderStatus:
    value = str(raw or "").strip()
    if not value:
        raise ValueError("empty_status")
    if "." in value:
        value = value.split(".")[-1].strip()
    try:
        return OrderStatus[value.upper()]
    except KeyError as exc:
        raise ValueError("invalid_status") from exc


def _compute_sync_now_budgets(
    timeout_sec: float,
    *,
    safety_margin: float = 1.5,
    default_orders: float = 12.0,
    default_goods: float = 10.0,
    default_feed: float = 3.0,
) -> dict[str, float]:
    budget_total = max(0.1, float(timeout_sec) - safety_margin)
    default_total = default_orders + default_goods + default_feed
    scale = (budget_total / default_total) if default_total > 0 else 1.0

    orders_timeout_sec = max(0.1, default_orders * scale)
    goods_timeout_sec = max(0.0, default_goods * scale)
    feed_timeout_sec = max(0.0, default_feed * scale)

    min_orders_budget = min(10.0, budget_total)
    non_orders_total = goods_timeout_sec + feed_timeout_sec
    max_non_orders = max(0.0, budget_total - min_orders_budget)
    if non_orders_total > max_non_orders and non_orders_total > 0:
        shrink = max_non_orders / non_orders_total
        goods_timeout_sec = max(0.0, goods_timeout_sec * shrink)
        feed_timeout_sec = max(0.0, feed_timeout_sec * shrink)

    orders_budget = max(0.1, budget_total - goods_timeout_sec - feed_timeout_sec)
    final_orders_timeout = min(max(orders_budget, 0.1), 60.0)

    return {
        "budget_total": budget_total,
        "orders_timeout_sec": orders_timeout_sec,
        "goods_timeout_sec": goods_timeout_sec,
        "feed_timeout_sec": feed_timeout_sec,
        "orders_budget": orders_budget,
        "final_orders_timeout": final_orders_timeout,
    }


def _normalize_db_dt(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(tz=UTC).replace(tzinfo=None)


def _extract_kaspi_notes(order: Order) -> dict[str, Any]:
    raw = order.internal_notes
    if raw is None:
        return {}
    if isinstance(raw, dict):
        data = raw
    elif isinstance(raw, str):
        try:
            data = json.loads(raw) if raw.strip() else {}
        except json.JSONDecodeError:
            return {}
    else:
        return {}
    if not isinstance(data, dict):
        return {}
    kaspi = data.get("kaspi")
    return kaspi if isinstance(kaspi, dict) else {}


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return str(value)


def _order_to_list_item(order: Order) -> KaspiOrderListItemOut:
    kaspi_notes = _extract_kaspi_notes(order)
    kaspi_preorder = kaspi_notes.get("preOrder")
    if not isinstance(kaspi_preorder, bool):
        kaspi_preorder = None
    return KaspiOrderListItemOut(
        id=order.id,
        external_id=order.external_id,
        order_number=order.order_number,
        status=_status_to_str(order.status),
        created_at=order.created_at,
        updated_at=order.updated_at,
        total_amount=Decimal(order.total_amount or 0),
        currency=order.currency,
        customer_name=order.customer_name,
        customer_phone=order.customer_phone,
        delivery_date=order.delivery_date,
        kaspi_preorder=kaspi_preorder,
        kaspi_planned_delivery_date=_optional_str(kaspi_notes.get("plannedDeliveryDate")),
        kaspi_reservation_date=_optional_str(kaspi_notes.get("reservationDate")),
    )


def _order_to_detail(order: Order) -> KaspiOrderDetailOut:
    items = [
        KaspiOrderEntryOut(
            id=item.id,
            sku=item.sku,
            name=item.name,
            quantity=int(item.quantity or 0),
            unit_price=Decimal(item.unit_price or 0),
            total_price=Decimal(item.total_price or 0),
        )
        for item in (order.items or [])
    ]
    base = _order_to_list_item(order)
    return KaspiOrderDetailOut(**base.model_dump(), items=items)


async def kaspi_orders_list(
    request: Request,
    merchant_uid: str | None = Query(None, min_length=1, alias="merchantUid"),
    status_filter: str | None = Query(None, alias="status"),
    state: str | None = Query(None),
    q: str | None = Query(None),
    created_from: str | None = Query(None),
    created_to: str | None = Query(None),
    page: int = Query(1, ge=1),
    offset: int | None = Query(None, ge=0),
    limit: int = Query(50, ge=1, le=200),
    sort: Literal["created_at_desc", "created_at_asc"] = Query("created_at_desc"),
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_ORDERS_LIST)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    request_id = getattr(getattr(request, "state", None), "request_id", None)
    rid = request_id or request.headers.get("X-Request-ID") or str(uuid4())
    company = await session.get(Company, company_id)
    resolved_merchant_uid, merchant_source, merchant_error = _resolve_company_store_uid(
        company=company,
        raw_merchant_uid=merchant_uid,
    )
    if merchant_error == "missing_merchant_uid":
        payload = {"detail": "missing_merchant_uid", "code": "missing_merchant_uid", "request_id": rid}
        return JSONResponse(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            content=payload,
            headers={"X-Request-ID": rid},
        )
    if merchant_error == "merchant_not_found":
        payload = {"detail": "merchant_not_found", "code": "merchant_not_found", "request_id": rid}
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content=payload,
            headers={"X-Request-ID": rid},
        )

    dt_from = None
    dt_to = None
    if created_from:
        try:
            dt_from = _normalize_db_dt(_parse_iso_dt(created_from))
        except HTTPException:
            payload = {"detail": "invalid_datetime", "code": "invalid_datetime", "request_id": rid}
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content=payload,
                headers={"X-Request-ID": rid},
            )
    if created_to:
        try:
            dt_to = _normalize_db_dt(_parse_iso_dt(created_to))
        except HTTPException:
            payload = {"detail": "invalid_datetime", "code": "invalid_datetime", "request_id": rid}
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content=payload,
                headers={"X-Request-ID": rid},
            )

    stmt = select(Order).where(
        Order.company_id == company_id,
        Order.source == OrderSource.KASPI,
    )
    # status has priority; state is legacy alias.
    effective_status = status_filter or state
    if effective_status:
        try:
            parsed_status = _parse_order_status(effective_status)
        except ValueError:
            payload = {"detail": "invalid_status", "code": "invalid_status", "request_id": rid}
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content=payload,
                headers={"X-Request-ID": rid},
            )
        stmt = stmt.where(Order.status == parsed_status)
    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            sa.or_(
                Order.order_number.ilike(like),
                Order.external_id.ilike(like),
                Order.customer_name.ilike(like),
                Order.customer_phone.ilike(like),
            )
        )
    if dt_from:
        stmt = stmt.where(Order.created_at >= dt_from)
    if dt_to:
        stmt = stmt.where(Order.created_at <= dt_to)
    _ = resolved_merchant_uid
    _ = merchant_source

    total = (await session.scalar(select(sa.func.count()).select_from(stmt.subquery()))) or 0

    effective_offset = offset if offset is not None else (page - 1) * limit
    order_by = [Order.created_at.desc(), Order.id.desc()]
    if sort == "created_at_asc":
        order_by = [Order.created_at.asc(), Order.id.asc()]
    rows = (await session.execute(stmt.order_by(*order_by).limit(limit).offset(effective_offset))).scalars().all()

    return KaspiOrdersListOut(
        items=[_order_to_list_item(order) for order in rows],
        page=(effective_offset // limit) + 1,
        limit=limit,
        total=int(total),
        offset=effective_offset,
    )


async def kaspi_order_detail(
    order_id: int,
    request: Request,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_ORDERS_LIST)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    request_id = getattr(getattr(request, "state", None), "request_id", None)
    rid = request_id or request.headers.get("X-Request-ID") or str(uuid4())

    order = await session.get(Order, order_id)
    if not order or order.company_id != company_id or order.source != OrderSource.KASPI:
        payload = {"detail": "order_not_found", "code": "order_not_found", "request_id": rid}
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content=payload,
            headers={"X-Request-ID": rid},
        )

    await session.refresh(order, attribute_names=["items"])
    return _order_to_detail(order)


def _action_response_payload(result: dict[str, Any], *, request_id: str) -> dict[str, Any]:
    payload = {
        "ok": bool(result.get("ok")),
        "status": result.get("status") or "error",
        "code": result.get("code"),
        "message": result.get("message"),
        "request_id": request_id,
    }
    if result.get("upstream_status_code") is not None:
        payload["upstream_status_code"] = result.get("upstream_status_code")
    if result.get("retry_after") is not None:
        payload["retry_after"] = result.get("retry_after")
    return payload


async def _handle_order_action(
    *,
    action: Literal["accept", "cancel"],
    request: Request,
    external_id: str,
    merchant_uid: str | None,
    current_user: User,
    session: AsyncSession,
):
    company_id = _resolve_company_id(current_user)
    request_id = getattr(getattr(request, "state", None), "request_id", None)
    rid = request_id or request.headers.get("X-Request-ID") or str(uuid4())

    company = await session.get(Company, company_id)
    resolved_merchant_uid, _merchant_source, merchant_error = _resolve_company_store_uid(
        company=company,
        raw_merchant_uid=merchant_uid,
    )
    if merchant_error == "missing_merchant_uid":
        payload = {"detail": "missing_merchant_uid", "code": "missing_merchant_uid", "request_id": rid}
        return JSONResponse(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            content=payload,
            headers={"X-Request-ID": rid},
        )
    if merchant_error == "merchant_not_found":
        payload = {"detail": "merchant_not_found", "code": "merchant_not_found", "request_id": rid}
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content=payload,
            headers={"X-Request-ID": rid},
        )

    order = (
        (
            await session.execute(
                select(Order).where(
                    Order.company_id == company_id,
                    Order.external_id == external_id,
                    Order.source == OrderSource.KASPI,
                )
            )
        )
        .scalars()
        .first()
    )
    if order is None:
        payload = {"detail": "order_not_found", "code": "order_not_found", "request_id": rid}
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content=payload,
            headers={"X-Request-ID": rid},
        )

    svc = KaspiService()
    if action == "accept":
        result = await svc.accept_order(
            external_id=external_id,
            merchant_uid=resolved_merchant_uid,
            request_id=rid,
        )
    else:
        result = await svc.cancel_order(
            external_id=external_id,
            merchant_uid=resolved_merchant_uid,
            request_id=rid,
        )

    http_status = int(result.get("http_status") or 500)
    if not result.get("ok"):
        return JSONResponse(
            status_code=http_status,
            content=_action_response_payload(result, request_id=rid),
            headers={"X-Request-ID": rid},
        )

    mapped_status = result.get("mapped_status")
    now = datetime.utcnow()
    if mapped_status:
        try:
            new_status = OrderStatus(mapped_status)
        except Exception:
            new_status = None
        if new_status and order.status != new_status:
            old_status = order.status
            order.status = new_status
            session.add(
                OrderStatusHistory(
                    order_id=order.id,
                    old_status=old_status,
                    new_status=new_status,
                    changed_by=getattr(current_user, "id", None),
                    changed_at=now,
                    note=f"kaspi_order_action:{action}",
                )
            )
    order.updated_at = now

    await _record_kaspi_event(
        session,
        company_id=company_id,
        merchant_uid=resolved_merchant_uid,
        kind="kaspi_order_action",
        status="success",
        request_id=rid,
        meta_json={
            "action": action,
            "external_id": external_id,
            "merchant_uid": resolved_merchant_uid,
        },
        commit=False,
    )

    await session.commit()

    return JSONResponse(
        status_code=200,
        content=_action_response_payload(result, request_id=rid),
        headers={"X-Request-ID": rid},
    )


async def kaspi_order_accept(
    request: Request,
    external_id: str,
    merchant_uid: str | None = Query(None, min_length=1, alias="merchantUid"),
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_ORDERS_LIST)),
    session: AsyncSession = Depends(get_async_db),
):
    return await _handle_order_action(
        action="accept",
        request=request,
        external_id=external_id,
        merchant_uid=merchant_uid,
        current_user=current_user,
        session=session,
    )


async def kaspi_order_cancel(
    request: Request,
    external_id: str,
    merchant_uid: str | None = Query(None, min_length=1, alias="merchantUid"),
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_ORDERS_LIST)),
    session: AsyncSession = Depends(get_async_db),
):
    return await _handle_order_action(
        action="cancel",
        request=request,
        external_id=external_id,
        merchant_uid=merchant_uid,
        current_user=current_user,
        session=session,
    )


async def kaspi_orders(query: OrdersQuery):
    try:
        return KaspiAdapter().orders(query.store, state=query.state)
    except KaspiAdapterError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))
    except Exception as e:
        logger.error("Kaspi orders unexpected error: payload=%s err=%s", query.model_dump(), e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


async def kaspi_import(req: ImportRequest):
    try:
        return KaspiAdapter().publish_feed(req.store, req.offers_json_path)
    except KaspiAdapterError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))
    except Exception as e:
        logger.error("Kaspi import unexpected error: payload=%s err=%s", req.model_dump(), e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


async def kaspi_import_status(req: ImportStatusQuery):
    try:
        return KaspiAdapter().import_status(req.store, import_id=req.import_id)
    except KaspiAdapterError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))
    except Exception as e:
        logger.error("Kaspi import_status unexpected error: payload=%s err=%s", req.model_dump(), e)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


register_kaspi_orders_routes(
    router,
    kaspi_orders_list=kaspi_orders_list,
    kaspi_order_detail=kaspi_order_detail,
    kaspi_order_accept=kaspi_order_accept,
    kaspi_order_cancel=kaspi_order_cancel,
    kaspi_orders=kaspi_orders,
    kaspi_import=kaspi_import,
    kaspi_import_status=kaspi_import_status,
    kaspi_orders_list_out_model=KaspiOrdersListOut,
    kaspi_order_detail_out_model=KaspiOrderDetailOut,
    kaspi_order_action_out_model=KaspiOrderActionOut,
)


# ================================== Service ==================================


async def kaspi_orders_sync(
    request: Request,
    response: Response,
    timeout_sec: int = Query(
        30,
        ge=5,
        le=300,
        description="Overall timeout budget for this sync call in seconds",
    ),
    max_pages: int = Query(2, ge=1, le=50),
    max_window_minutes: int = Query(1440, ge=5, le=10080),
    backfill_days: int = Query(0, ge=0, le=365),
    merchant_uid: str | None = Query(None, min_length=1, alias="merchantUid"),
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    resolved_company_id: int | None = None
    svc: KaspiService | None = None
    try:
        resolved_company_id = _resolve_company_id(current_user)
        company = await session.get(Company, resolved_company_id)
        resolved_merchant_uid, merchant_source = _resolve_merchant_uid(
            company=company,
            raw_merchant_uid=merchant_uid,
            missing_status=status.HTTP_422_UNPROCESSABLE_CONTENT,
            missing_detail="missing_merchant_uid",
        )

        logger.info(
            "kaspi_orders_sync merchant_uid resolved",
            extra={
                "company_id": resolved_company_id,
                "request_id": getattr(getattr(request, "state", None), "request_id", None),
                "merchant_uid": resolved_merchant_uid,
                "merchant_uid_source": merchant_source,
                "store_selected": resolved_merchant_uid,
            },
        )

        stub_mode = _kaspi_stub_enabled()

        env_value = (settings.ENVIRONMENT or "").lower()
        if backfill_days > 0 and env_value not in {"dev", "development", "test", "testing", "local"}:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="backfill_days_allowed_only_in_development",
            )

        if stub_mode:
            payload = {
                "ok": False,
                "status": "not_implemented",
                "code": "KASPI_STUB_NOT_IMPLEMENTED",
                "detail": "kaspi_stub_not_implemented",
                "company_id": resolved_company_id,
                "merchant_uid": resolved_merchant_uid,
            }
            return JSONResponse(status_code=status.HTTP_501_NOT_IMPLEMENTED, content=payload)

        svc = KaspiService()
        request_id = getattr(getattr(request, "state", None), "request_id", None) if request else None
        result = await svc.sync_orders(
            db=session,
            company_id=resolved_company_id,
            merchant_uid=resolved_merchant_uid,
            request_id=request_id,
            timeout_seconds=timeout_sec,
            max_pages=max_pages,
            max_window_minutes=max_window_minutes,
            backfill_days=backfill_days,
        )
        if "ok" not in result and "status" not in result and "code" not in result:
            response.status_code = status.HTTP_200_OK
            return result

        if result.get("ok") is True:
            await _record_kaspi_event(
                session,
                company_id=resolved_company_id,
                kind="kaspi_orders_sync",
                status="success",
                request_id=request_id,
                meta_json={
                    "fetched": result.get("fetched"),
                    "inserted": result.get("inserted"),
                    "updated": result.get("updated"),
                    "watermark": result.get("watermark"),
                },
            )
            return result

        code = str(result.get("code") or "internal_error")
        status_value = str(result.get("status") or "failed")
        retry_after = result.get("retry_after")

        if code in {"locked", "sync_locked"} or status_value == "locked":
            response.status_code = status.HTTP_423_LOCKED
            detail = "kaspi sync already running"
        elif code in {"timeout", "connect_timeout", "read_timeout"} or status_value == "timeout":
            response.status_code = status.HTTP_504_GATEWAY_TIMEOUT
            detail = "kaspi timeout"
        elif code == "rate_limited" or status_value == "rate_limited":
            response.status_code = status.HTTP_429_TOO_MANY_REQUESTS
            detail = "kaspi rate limited"
            if retry_after is not None:
                response.headers["Retry-After"] = str(retry_after)
        elif code.upper() == "KASPI_BAD_REQUEST":
            response.status_code = status.HTTP_422_UNPROCESSABLE_CONTENT
            detail = str(result.get("message") or result.get("detail") or "kaspi bad request")
        elif code == "upstream_unavailable":
            response.status_code = status.HTTP_502_BAD_GATEWAY
            detail = "kaspi upstream error"
        else:
            response.status_code = status.HTTP_500_INTERNAL_SERVER_ERROR
            detail = "kaspi sync failed"

        await _record_kaspi_event(
            session,
            company_id=resolved_company_id,
            kind="kaspi_orders_sync",
            status="failed" if response.status_code >= 500 else "skipped",
            request_id=request_id,
            error_code=code,
            error_message=detail,
        )

        return {
            **result,
            "detail": detail,
            "code": code,
            "request_id": request_id,
        }
    except KaspiSyncAlreadyRunning:
        request_id = getattr(getattr(request, "state", None), "request_id", None) if request else None
        if resolved_company_id is not None:
            await _record_kaspi_event(
                session,
                company_id=resolved_company_id,
                kind="kaspi_orders_sync",
                status="skipped",
                request_id=request_id,
                error_code="sync_locked",
                error_message="already_running",
            )
        response.status_code = status.HTTP_423_LOCKED
        return {
            "ok": False,
            "status": "locked",
            "code": "locked",
            "detail": "kaspi sync already running",
            "request_id": request_id,
        }
    except HTTPException:
        raise
    except Exception as e:
        svc = svc or KaspiService()
        error_code = svc.classify_sync_error(e)
        retry_after = svc.get_retry_after_seconds(e)
        logger.error("Kaspi orders sync failed: company_id=%s code=%s err=%s", resolved_company_id, error_code, e)
        request_id = getattr(getattr(request, "state", None), "request_id", None) if request else None
        if resolved_company_id is not None:
            await _record_kaspi_event(
                session,
                company_id=resolved_company_id,
                kind="kaspi_orders_sync",
                status="failed",
                request_id=request_id,
                error_code=error_code,
                error_message=safe_error_message(e),
            )

        if error_code == "kaspi_http_429":
            response.status_code = status.HTTP_429_TOO_MANY_REQUESTS
            if retry_after is not None:
                response.headers["Retry-After"] = str(retry_after)
            detail = "kaspi rate limited"
        elif error_code in {"kaspi_timeout", "timeout"}:
            response.status_code = status.HTTP_504_GATEWAY_TIMEOUT
            detail = "kaspi timeout"
        elif error_code.startswith("kaspi_http_") or error_code == "kaspi_adapter_error":
            response.status_code = status.HTTP_502_BAD_GATEWAY
            detail = "kaspi upstream error"
        else:
            response.status_code = status.HTTP_500_INTERNAL_SERVER_ERROR
            detail = "kaspi sync failed"

        return {
            "ok": False,
            "status": "failed",
            "code": error_code,
            "detail": detail,
            "request_id": request_id,
        }


async def kaspi_generate_feed(
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    resolved_company_id: int | None = None
    try:
        resolved_company_id = _resolve_company_id(current_user)
        svc = KaspiService()
        xml_body = await svc.generate_product_feed(company_id=resolved_company_id, db=session)
        return Response(content=xml_body, media_type="application/xml")
    except HTTPException:
        raise
    except RuntimeError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))
    except Exception as e:
        logger.exception("Kaspi generate feed unexpected error: company_id=%s", resolved_company_id)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


async def kaspi_availability_sync_one(
    payload: AvailabilitySyncIn,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    try:
        resolved_company_id = _resolve_company_id(current_user)
        res = await session.execute(
            sa.select(Product).where(Product.id == payload.product_id, Product.company_id == resolved_company_id)
        )
        product: Product | None = res.scalar_one_or_none()
        if not product:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")

        svc = KaspiService()
        ok = await svc.sync_product_availability(product, db=session)
        return {"ok": bool(ok)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Kaspi availability sync one failed: payload=%s err=%s", payload.model_dump(), e)
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))


async def kaspi_availability_bulk(
    payload: AvailabilityBulkIn,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    try:
        resolved_company_id = _resolve_company_id(current_user)
        svc = KaspiService()
        stats = await svc.bulk_sync_availability(company_id=resolved_company_id, db=session, limit=payload.limit)
        return stats
    except Exception as e:
        logger.error("Kaspi availability bulk failed: payload=%s err=%s", payload.model_dump(), e)
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))


class KaspiSyncStateOut(BaseModel):
    watermark: Any | None = None
    last_success_at: Any | None = None
    last_attempt_at: Any | None = None
    last_duration_ms: int | None = None
    last_result: str | None = None
    last_fetched: int | None = None
    last_inserted: int | None = None
    last_updated: int | None = None
    last_error_at: Any | None = None
    last_error_code: str | None = None
    last_error_message: str | None = None


class KaspiSyncOpsOut(KaspiSyncStateOut):
    lock_available: bool


class KaspiCatalogItemOut(BaseModel):
    sku: str
    merchant_uid: str
    offer_code: str | None = None
    product_code: str | None = None
    last_seen_name: str | None = None
    last_seen_price: Decimal | None = None
    last_seen_qty: int | None = None
    last_seen_at: datetime | None = None


class KaspiCatalogItemsOut(BaseModel):
    items: list[KaspiCatalogItemOut]
    total: int
    limit: int
    offset: int


async def kaspi_orders_sync_state(
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    res = await session.execute(sa.select(KaspiOrderSyncState).where(KaspiOrderSyncState.company_id == company_id))
    state = res.scalar_one_or_none()
    watermark = getattr(state, "last_synced_at", None) if state else None
    last_success_at = getattr(state, "last_synced_at", None) if state else None
    last_attempt_at = getattr(state, "last_attempt_at", None) if state else None
    last_duration_ms = getattr(state, "last_duration_ms", None) if state else None
    last_result = getattr(state, "last_result", None) if state else None
    last_fetched = getattr(state, "last_fetched", None) if state else None
    last_inserted = getattr(state, "last_inserted", None) if state else None
    last_updated = getattr(state, "last_updated", None) if state else None
    last_error_at = getattr(state, "last_error_at", None) if state else None
    last_error_code = getattr(state, "last_error_code", None) if state else None
    last_error_message = getattr(state, "last_error_message", None) if state else None
    return KaspiSyncStateOut(
        watermark=watermark,
        last_success_at=last_success_at,
        last_attempt_at=last_attempt_at,
        last_duration_ms=last_duration_ms,
        last_result=last_result,
        last_fetched=last_fetched,
        last_inserted=last_inserted,
        last_updated=last_updated,
        last_error_at=last_error_at,
        last_error_code=last_error_code,
        last_error_message=last_error_message,
    )


async def kaspi_orders_sync_ops(
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    res = await session.execute(sa.select(KaspiOrderSyncState).where(KaspiOrderSyncState.company_id == company_id))
    state = res.scalar_one_or_none()

    watermark = getattr(state, "last_synced_at", None) if state else None
    last_success_at = getattr(state, "last_synced_at", None) if state else None
    last_attempt_at = getattr(state, "last_attempt_at", None) if state else None
    last_duration_ms = getattr(state, "last_duration_ms", None) if state else None
    last_result = getattr(state, "last_result", None) if state else None
    last_fetched = getattr(state, "last_fetched", None) if state else None
    last_inserted = getattr(state, "last_inserted", None) if state else None
    last_updated = getattr(state, "last_updated", None) if state else None
    last_error_at = getattr(state, "last_error_at", None) if state else None
    last_error_code = getattr(state, "last_error_code", None) if state else None
    last_error_message = getattr(state, "last_error_message", None) if state else None

    svc = KaspiService()
    lock_available = False
    try:
        lock_available = await svc.check_lock_available(session, company_id)
    except Exception:
        lock_available = False

    return KaspiSyncOpsOut(
        watermark=watermark,
        last_success_at=last_success_at,
        last_attempt_at=last_attempt_at,
        last_duration_ms=last_duration_ms,
        last_result=last_result,
        last_fetched=last_fetched,
        last_inserted=last_inserted,
        last_updated=last_updated,
        last_error_at=last_error_at,
        last_error_code=last_error_code,
        last_error_message=last_error_message,
        lock_available=bool(lock_available),
    )


async def kaspi_catalog_items(
    request: Request,
    merchant_uid: str | None = Query(None, min_length=1),
    q: str | None = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_ORDERS_LIST)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    request_id = getattr(getattr(request, "state", None), "request_id", None)
    rid = request_id or request.headers.get("X-Request-ID") or str(uuid4())
    company = await session.get(Company, company_id)
    resolved_merchant_uid, _merchant_source, merchant_error = _resolve_company_store_uid(
        company=company,
        raw_merchant_uid=merchant_uid,
    )
    if merchant_error == "missing_merchant_uid":
        payload = {"detail": "missing_merchant_uid", "code": "missing_merchant_uid", "request_id": rid}
        return JSONResponse(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            content=payload,
            headers={"X-Request-ID": rid},
        )
    if merchant_error == "merchant_not_found":
        payload = {"detail": "merchant_not_found", "code": "merchant_not_found", "request_id": rid}
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content=payload,
            headers={"X-Request-ID": rid},
        )

    stmt = select(KaspiCatalogItem).where(KaspiCatalogItem.company_id == company_id)
    if resolved_merchant_uid:
        stmt = stmt.where(KaspiCatalogItem.merchant_uid == resolved_merchant_uid)
    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            sa.or_(
                KaspiCatalogItem.sku.ilike(like),
                KaspiCatalogItem.last_seen_name.ilike(like),
            )
        )

    total = (await session.scalar(select(sa.func.count()).select_from(stmt.subquery()))) or 0
    rows = (
        (
            await session.execute(
                stmt.order_by(sa.nullslast(KaspiCatalogItem.last_seen_at.desc()), KaspiCatalogItem.id.desc())
                .limit(limit)
                .offset(offset)
            )
        )
        .scalars()
        .all()
    )

    return KaspiCatalogItemsOut(
        items=[
            KaspiCatalogItemOut(
                sku=item.sku,
                merchant_uid=item.merchant_uid,
                offer_code=item.offer_code,
                product_code=item.product_code,
                last_seen_name=item.last_seen_name,
                last_seen_price=item.last_seen_price,
                last_seen_qty=item.last_seen_qty,
                last_seen_at=item.last_seen_at,
            )
            for item in rows
        ],
        total=int(total),
        limit=limit,
        offset=offset,
    )


register_kaspi_sync_routes(
    router,
    kaspi_orders_sync=kaspi_orders_sync,
    kaspi_generate_feed=kaspi_generate_feed,
    kaspi_availability_sync_one=kaspi_availability_sync_one,
    kaspi_availability_bulk=kaspi_availability_bulk,
    kaspi_orders_sync_state=kaspi_orders_sync_state,
    kaspi_orders_sync_ops=kaspi_orders_sync_ops,
    kaspi_catalog_items=kaspi_catalog_items,
    kaspi_sync_state_out_model=KaspiSyncStateOut,
    kaspi_sync_ops_out_model=KaspiSyncOpsOut,
    kaspi_catalog_items_out_model=KaspiCatalogItemsOut,
)


register_kaspi_debug_routes(
    router,
    auth_dependency=_auth_user,
    get_async_db_dependency=get_async_db,
    require_platform_admin_fn=require_platform_admin,
    require_store_admin_company_scoped_fn=_require_store_admin_company_scoped,
    kaspi_store_token_model=KaspiStoreToken,
    build_kaspi_httpx_client_fn=_build_kaspi_httpx_client,
    build_kaspi_orders_params_fn=_build_kaspi_orders_params,
    is_dev_environment_fn=lambda: _is_dev_environment(),
    kaspi_user_agent_fn=_kaspi_user_agent,
    probe_response_snippet_fn=_probe_response_snippet,
    log_kaspi_probe_error_fn=_log_kaspi_probe_error,
    log_kaspi_probe_response_fn=_log_kaspi_probe_response,
    safe_log_info_fn=_safe_log_info,
    kaspi_ns=_KASPI_NS,
    fast_probe_timeout=FAST_PROBE_TIMEOUT,
)

register_kaspi_autosync_routes(
    router,
    require_store_admin_then_feature_fn=require_store_admin_then_feature,
    feature_kaspi_autosync=FEATURE_KASPI_AUTOSYNC,
    logger=logger,
)


# ============================= CATALOG PRODUCTS ==============================


class KaspiProductSyncOut(BaseModel):
    """Response model for catalog sync operation."""

    ok: bool
    company_id: int
    fetched: int
    inserted: int
    updated: int


class KaspiCatalogPullUnsupportedOut(BaseModel):
    detail: str
    code: str
    request_id: str | None = None
    errors: list[dict[str, Any]] = Field(default_factory=list)


class KaspiImportRunOut(BaseModel):
    ok: bool
    import_code: str
    status: str
    kaspi_import_code: str | None = None
    request_id: str | None = None


class KaspiImportRunPollOut(KaspiImportRunOut):
    status_payload: dict[str, Any] | None = None
    result_payload: dict[str, Any] | None = None


class KaspiProductOut(BaseModel):
    """Response model for single product in catalog list."""

    offer_id: str
    name: str | None = None
    sku: str | None = None
    price: str | None = None
    qty: int | None = None
    is_active: bool


class KaspiProductListOut(BaseModel):
    """Response model for catalog products list."""

    items: list[KaspiProductOut]
    total: int
    limit: int
    offset: int


class KaspiGoodsImportIn(BaseModel):
    payload: list[dict[str, Any]] | None = None
    product_ids: list[int] | None = None
    content_type: str | None = None


class KaspiGoodsImportCreateIn(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    merchant_uid: str = Field(..., min_length=1, alias="merchantUid")
    source: str = Field("db")
    comment: str | None = None


class KaspiGoodsImportOut(BaseModel):
    ok: bool
    import_code: str
    status: str


class KaspiGoodsUploadOut(BaseModel):
    import_code: str
    status: str
    payload: dict[str, Any] | None = None


class KaspiGoodsStatusOut(BaseModel):
    import_code: str
    status: str
    payload: dict[str, Any] | None = None


class KaspiGoodsResultOut(BaseModel):
    import_code: str
    status: str
    payload: dict[str, Any] | None = None


class KaspiOffersPreviewOut(BaseModel):
    items: list[dict[str, Any]]
    total: int
    payload_hash: str


class KaspiGoodsImportRecordOut(BaseModel):
    id: str
    merchant_uid: str | None = None
    import_code: str
    status: str
    source: str | None = None
    payload_hash: str | None = None
    attempts: int | None = None
    request_json: list[dict[str, Any]] | dict[str, Any]
    status_json: dict[str, Any] | None = None
    raw_status_json: dict[str, Any] | None = None
    result_json: dict[str, Any] | None = None
    error_code: str | None = None
    error_message: str | None = None
    created_at: datetime
    updated_at: datetime
    last_checked_at: datetime | None = None
    revoked_at: datetime | None = None


class KaspiImportValidationOut(BaseModel):
    ok: bool
    errors: list[dict[str, Any]] = Field(default_factory=list)
    total: int = 0
    payload_hash: str | None = None


class KaspiSyncNowIn(BaseModel):
    merchant_uid: str = Field(..., min_length=1)
    refresh_once: bool = True


class KaspiSyncNowOut(BaseModel):
    ok: bool
    status: str | None = None
    result: str | None = None
    phase: str | None = None
    errors: list[dict[str, Any]] = Field(default_factory=list)
    company_id: int
    merchant_uid: str
    orders_sync: dict[str, Any] | None = None
    goods_import_id: str | None = None
    goods_import_code: str | None = None
    goods_import_status: str | None = None
    goods_import_result: dict[str, Any] | None = None
    feed_last_generated_at: datetime | None = None
    offers_feed_result: dict[str, Any] | None = None
    offers_feed_upload_result: dict[str, Any] | None = None


class KaspiFeedUploadIn(BaseModel):
    merchant_uid: str = Field(..., min_length=3, max_length=128)
    source: str = Field(..., description="public_token | export_id | local_file_path")
    comment: str | None = Field(None, max_length=500)
    export_id: int | None = None
    local_file_path: str | None = None

    @field_validator("export_id", mode="before")
    @classmethod
    def _coerce_export_id(cls, value: object) -> object:
        if value is None or isinstance(value, int):
            return value
        if isinstance(value, str):
            raw = value.strip()
            if raw.isdigit():
                return int(raw)
        raise ValueError("export_id must be an integer")


class KaspiOffersFeedUploadIn(BaseModel):
    merchant_uid: str = Field(..., min_length=3, max_length=128)
    refresh: bool = False
    comment: str | None = Field(None, max_length=500)


class KaspiOffersFeedUploadOut(BaseModel):
    url: str
    token_id: int
    payload_hash: str
    generated_at: datetime
    merchant_uid: str


class KaspiFeedUploadRecordOut(BaseModel):
    id: str
    merchant_uid: str
    import_code: str | None = None
    status: str
    source: str | None = None
    comment: str | None = None
    attempts: int = 0
    last_error_code: str | None = None
    last_error_message: str | None = None
    request_id: str | None = None
    payload_hash: str | None = None
    response_json: dict[str, Any] | None = None
    next_attempt_at: datetime | None = None
    created_at: datetime
    updated_at: datetime


class KaspiTokenHealthOut(BaseModel):
    ok: bool
    orders_http: int
    goods_http: int
    cause: str | None = None


class KaspiTokenSelftestOut(BaseModel):
    orders_http: int
    goods_schema_http: int
    goods_categories_http: int
    goods_access: str | None = None
    orders_error: str | None = None


class KaspiCatalogImportOut(BaseModel):
    batch_id: str | None = None
    status: str
    rows_total: int
    rows_ok: int
    rows_skipped: int
    top_errors: list[dict[str, Any]]
    dry_run: bool = False


class KaspiCatalogImportBatchOut(BaseModel):
    id: str
    merchant_uid: str | None = None
    filename: str
    status: str
    rows_total: int
    rows_ok: int
    rows_failed: int
    started_at: datetime | None = None
    finished_at: datetime | None = None
    created_at: datetime
    error_summary: str | None = None


class KaspiCatalogImportBatchDetailOut(KaspiCatalogImportBatchOut):
    duration_seconds: int | None = None


class KaspiCatalogImportErrorOut(BaseModel):
    row_num: int
    error: str | None = None
    sku: str | None = None
    master_sku: str | None = None
    title: str | None = None
    raw: str | None = None


class KaspiOfferOut(BaseModel):
    id: int
    merchant_uid: str
    sku: str
    master_sku: str | None = None
    title: str | None = None
    price: float | None = None
    old_price: float | None = None
    stock_count: int | None = None
    pre_order: bool | None = None
    stock_specified: bool | None = None
    updated_at: datetime


class KaspiOffersRebuildOut(BaseModel):
    ok: bool
    company_id: int
    merchant_uid: str
    fetched: int
    inserted: int
    updated: int


class KaspiOffersImportOut(BaseModel):
    ok: bool
    company_id: int
    merchant_uid: str
    fetched: int
    inserted: int
    updated: int
    skipped: int


class KaspiOfferListOut(BaseModel):
    items: list[KaspiOfferOut]
    total: int
    limit: int
    offset: int


async def kaspi_offers_rebuild(
    request: Request,
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    include_inactive: bool = Query(False, alias="includeInactive"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)
    company = await session.get(Company, company_id)
    resolved_uid = _resolve_kaspi_offers_merchant_uid(company=company, raw_merchant_uid=merchant_uid)

    offers = await _build_offers_from_products(
        session,
        company_id=company_id,
        merchant_uid=resolved_uid,
        include_inactive=include_inactive,
    )
    summary = await _upsert_kaspi_offers(session, offers=offers)
    await session.commit()

    return KaspiOffersRebuildOut(
        ok=True,
        company_id=company_id,
        merchant_uid=resolved_uid,
        fetched=summary["fetched"],
        inserted=summary["inserted"],
        updated=summary["updated"],
    )


async def kaspi_offers_import(
    request: Request,
    file: UploadFile = File(...),
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)
    company = await session.get(Company, company_id)
    resolved_uid = _resolve_kaspi_offers_merchant_uid(company=company, raw_merchant_uid=merchant_uid)

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="empty_file")

    filename = file.filename or "offers.csv"
    rows = parse_catalog_file(content, filename)
    offers: list[dict[str, Any]] = []
    skipped = 0
    for row in rows:
        sku = (row.get("sku") or "").strip()
        if not sku:
            skipped += 1
            continue
        offers.append(
            {
                "company_id": company_id,
                "merchant_uid": resolved_uid,
                "sku": sku,
                "master_sku": row.get("master_sku"),
                "title": row.get("title") or sku,
                "price": row.get("price"),
                "old_price": row.get("old_price"),
                "stock_count": row.get("stock_count"),
                "pre_order": row.get("pre_order"),
                "stock_specified": row.get("stock_specified"),
                "raw": row,
            }
        )

    if not offers:
        request_id = getattr(getattr(request, "state", None), "request_id", None)
        payload = {
            "detail": "offers_missing",
            "code": "offers_missing",
            "request_id": request_id,
            "errors": [{"code": "offers_missing", "detail": "offers_missing", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_409_CONFLICT, content=payload)

    summary = await _upsert_kaspi_offers(session, offers=offers)
    await session.commit()

    return KaspiOffersImportOut(
        ok=True,
        company_id=company_id,
        merchant_uid=resolved_uid,
        fetched=summary["fetched"],
        inserted=summary["inserted"],
        updated=summary["updated"],
        skipped=skipped,
    )


async def kaspi_offers_preview(
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    limit: int = Query(5, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)
    company = await session.get(Company, company_id)
    resolved_uid = _resolve_kaspi_offers_merchant_uid(company=company, raw_merchant_uid=merchant_uid)
    flags = _get_goods_import_flags(company, resolved_uid)
    payload = await load_offers_payload(
        session,
        company_id=company_id,
        merchant_uid=resolved_uid,
        include_price=flags["include_price"],
        include_stock=flags["include_stock"],
    )
    payload_json = build_payload_json(payload)
    payload_hash = compute_payload_hash(payload_json)
    return KaspiOffersPreviewOut(items=payload[:limit], total=len(payload), payload_hash=payload_hash)


async def kaspi_products_sync(
    request: Request,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    """
    Kaspi Shop API не поддерживает получение полного каталога через X-Auth-Token.
    Используйте /api/v1/kaspi/products/import и /api/v1/kaspi/products/import/result.
    """
    _ = _resolve_company_id(current_user)
    _ = session
    request_id = getattr(getattr(request, "state", None), "request_id", None)
    payload = {
        "detail": "catalog_pull_not_supported",
        "code": "catalog_pull_not_supported",
        "request_id": request_id,
        "errors": [
            {
                "code": "catalog_pull_not_supported",
                "detail": "catalog_pull_not_supported",
                "request_id": request_id,
            }
        ],
    }
    return JSONResponse(status_code=status.HTTP_409_CONFLICT, content=payload)


async def kaspi_products_import_start(
    request: Request,
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    company = await session.get(Company, company_id)
    resolved_uid = _resolve_kaspi_offers_merchant_uid(company=company, raw_merchant_uid=merchant_uid)
    request_id = getattr(getattr(request, "state", None), "request_id", None)

    run = await _create_import_run(
        session,
        company_id=company_id,
        merchant_uid=resolved_uid,
        request_id=request_id,
    )

    return KaspiImportRunOut(
        ok=True,
        import_code=run.import_code,
        status=run.status,
        kaspi_import_code=run.kaspi_import_code,
        request_id=request_id,
    )


async def kaspi_products_import_upload(
    request: Request,
    import_code: str = Query(..., alias="i"),
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    force: bool = Query(False, description="Force re-upload even if already uploaded"),
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    request_id = getattr(getattr(request, "state", None), "request_id", None)

    run = await _get_import_run(session, company_id=company_id, import_code=import_code)
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import_not_found")

    if run.kaspi_import_code and not force:
        return KaspiImportRunOut(
            ok=True,
            import_code=run.import_code,
            status=run.status,
            kaspi_import_code=run.kaspi_import_code,
            request_id=request_id,
        )

    if merchant_uid:
        resolved_uid = _resolve_kaspi_offers_merchant_uid(
            company=await session.get(Company, company_id),
            raw_merchant_uid=merchant_uid,
        )
        if resolved_uid != run.merchant_uid:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="merchant_not_found")

    company = await session.get(Company, company_id)
    flags = _get_goods_import_flags(company, run.merchant_uid)
    payload = await load_offers_payload(
        session,
        company_id=company_id,
        merchant_uid=run.merchant_uid,
        include_price=flags["include_price"],
        include_stock=flags["include_stock"],
    )
    if not payload:
        run.error_code = "offers_missing"
        run.error_message = "offers_missing"
        await session.commit()
        payload_error = {
            "detail": "offers_missing",
            "code": "offers_missing",
            "request_id": request_id,
            "errors": [{"code": "offers_missing", "detail": "offers_missing", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_409_CONFLICT, content=payload_error)

    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsImportClient(token=token, base_url="https://kaspi.kz")
    try:
        schema = await client.get_schema()
    except KaspiImportNotAuthenticated:
        run.error_code = "NOT_AUTHENTICATED"
        run.error_message = "NOT_AUTHENTICATED"
        await session.commit()
        payload_error = {
            "detail": "NOT_AUTHENTICATED",
            "code": "NOT_AUTHENTICATED",
            "request_id": request_id,
            "errors": [{"code": "NOT_AUTHENTICATED", "detail": "NOT_AUTHENTICATED", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content=payload_error)
    except KaspiImportUpstreamUnavailable:
        run.error_code = "upstream_unavailable"
        run.error_message = "kaspi_upstream_unavailable"
        await session.commit()
        payload_error = {
            "detail": "kaspi_upstream_unavailable",
            "code": "upstream_unavailable",
            "request_id": request_id,
            "errors": [
                {"code": "upstream_unavailable", "detail": "kaspi_upstream_unavailable", "request_id": request_id}
            ],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload_error)
    except KaspiImportUpstreamError:
        run.error_code = "upstream_error"
        run.error_message = "kaspi_upstream_error"
        await session.commit()
        payload_error = {
            "detail": "kaspi_upstream_error",
            "code": "upstream_error",
            "request_id": request_id,
            "errors": [{"code": "upstream_error", "detail": "kaspi_upstream_error", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload_error)

    schema_errors = _validate_payload_against_schema(payload, schema)
    if schema_errors:
        for err in schema_errors:
            err["request_id"] = request_id
        run.error_code = "schema_validation_failed"
        run.error_message = "schema_validation_failed"
        await session.commit()
        payload_error = {
            "detail": "schema_validation_failed",
            "code": "schema_validation_failed",
            "request_id": request_id,
            "errors": schema_errors,
        }
        return JSONResponse(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, content=payload_error)

    payload_json = build_payload_json(payload)
    payload_hash = compute_payload_hash(payload_json)

    existing = None
    if not force:
        existing = await _find_recent_import_run_by_hash(
            session,
            company_id=company_id,
            merchant_uid=run.merchant_uid,
            payload_hash=payload_hash,
        )
    if existing and existing.id != run.id:
        run.status = "DUPLICATE"
        run.kaspi_import_code = existing.kaspi_import_code
        run.payload_hash = payload_hash
        run.request_payload = payload
        run.status_json = existing.status_json
        run.result_json = existing.result_json
        run.last_checked_at = existing.last_checked_at
        run.next_poll_at = existing.next_poll_at
        run.error_code = "duplicate_payload"
        run.error_message = "duplicate_payload"
        await session.commit()
        return KaspiImportRunOut(
            ok=True,
            import_code=existing.import_code,
            status=existing.status,
            kaspi_import_code=existing.kaspi_import_code,
            request_id=request_id,
        )
    try:
        response = await client.submit_import(payload_json)
    except KaspiImportNotAuthenticated:
        run.error_code = "NOT_AUTHENTICATED"
        run.error_message = "NOT_AUTHENTICATED"
        await session.commit()
        payload_error = {
            "detail": "NOT_AUTHENTICATED",
            "code": "NOT_AUTHENTICATED",
            "request_id": request_id,
            "errors": [{"code": "NOT_AUTHENTICATED", "detail": "NOT_AUTHENTICATED", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content=payload_error)
    except KaspiImportUpstreamUnavailable:
        run.error_code = "upstream_unavailable"
        run.error_message = "kaspi_upstream_unavailable"
        await session.commit()
        payload_error = {
            "detail": "kaspi_upstream_unavailable",
            "code": "upstream_unavailable",
            "request_id": request_id,
            "errors": [
                {"code": "upstream_unavailable", "detail": "kaspi_upstream_unavailable", "request_id": request_id}
            ],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload_error)
    except KaspiImportUpstreamError:
        run.error_code = "upstream_error"
        run.error_message = "kaspi_upstream_error"
        await session.commit()
        payload_error = {
            "detail": "kaspi_upstream_error",
            "code": "upstream_error",
            "request_id": request_id,
            "errors": [{"code": "upstream_error", "detail": "kaspi_upstream_error", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload_error)

    kaspi_code = _extract_import_code(response or {})
    if not kaspi_code:
        run.error_code = "kaspi_import_code_missing"
        run.error_message = "kaspi_import_code_missing"
        await session.commit()
        payload_error = {
            "detail": "kaspi_import_code_missing",
            "code": "kaspi_import_code_missing",
            "request_id": request_id,
            "errors": [
                {"code": "kaspi_import_code_missing", "detail": "kaspi_import_code_missing", "request_id": request_id}
            ],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload_error)

    run.kaspi_import_code = str(kaspi_code)
    run.status = str(response.get("status") or "UPLOADED")
    run.request_payload = payload
    run.status_json = response
    run.payload_hash = payload_hash
    run.error_code = None
    run.error_message = None
    run.attempts = int(getattr(run, "attempts", 0) or 0) + 1
    from app.core.config import settings
    from app.services.kaspi_import_run_utils import compute_next_poll_at

    now = datetime.utcnow()
    run.last_checked_at = now
    run.next_poll_at = compute_next_poll_at(
        now=now,
        status=run.status,
        attempts=run.attempts,
        base_delay_seconds=settings.KASPI_IMPORT_POLL_BACKOFF_BASE_SECONDS,
        max_delay_seconds=settings.KASPI_IMPORT_POLL_BACKOFF_MAX_SECONDS,
        result_payload=run.result_json,
    )
    await session.commit()
    await session.refresh(run)

    return KaspiImportRunOut(
        ok=True,
        import_code=run.import_code,
        status=run.status,
        kaspi_import_code=run.kaspi_import_code,
        request_id=request_id,
    )


async def kaspi_products_import_poll(
    request: Request,
    import_code: str = Query(..., alias="i"),
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    request_id = getattr(getattr(request, "state", None), "request_id", None)

    run = await _get_import_run(session, company_id=company_id, import_code=import_code)
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import_not_found")
    if not run.kaspi_import_code:
        payload_error = {
            "detail": "import_not_uploaded",
            "code": "import_not_uploaded",
            "request_id": request_id,
            "errors": [{"code": "import_not_uploaded", "detail": "import_not_uploaded", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_409_CONFLICT, content=payload_error)

    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsImportClient(token=token, base_url="https://kaspi.kz")
    try:
        status_response = await client.get_status(import_code=run.kaspi_import_code)
        result_response = await client.get_result(import_code=run.kaspi_import_code)
    except KaspiImportNotAuthenticated:
        run.error_code = "NOT_AUTHENTICATED"
        run.error_message = "NOT_AUTHENTICATED"
        await session.commit()
        payload_error = {
            "detail": "NOT_AUTHENTICATED",
            "code": "NOT_AUTHENTICATED",
            "request_id": request_id,
            "errors": [{"code": "NOT_AUTHENTICATED", "detail": "NOT_AUTHENTICATED", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content=payload_error)
    except KaspiImportUpstreamUnavailable:
        run.error_code = "upstream_unavailable"
        run.error_message = "kaspi_upstream_unavailable"
        await session.commit()
        payload_error = {
            "detail": "kaspi_upstream_unavailable",
            "code": "upstream_unavailable",
            "request_id": request_id,
            "errors": [
                {"code": "upstream_unavailable", "detail": "kaspi_upstream_unavailable", "request_id": request_id}
            ],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload_error)
    except KaspiImportUpstreamError:
        run.error_code = "upstream_error"
        run.error_message = "kaspi_upstream_error"
        await session.commit()
        payload_error = {
            "detail": "kaspi_upstream_error",
            "code": "upstream_error",
            "request_id": request_id,
            "errors": [{"code": "upstream_error", "detail": "kaspi_upstream_error", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload_error)

    status_value = status_response.get("status") or run.status
    run.status = str(status_value)
    run.status_json = status_response
    run.result_json = result_response
    run.last_checked_at = datetime.utcnow()
    from app.services.kaspi_import_run_utils import classify_import_result

    result_class, _summary = classify_import_result(run.status, run.result_json)
    if result_class == "failed":
        run.error_code = "import_failed"
        run.error_message = "kaspi_import_failed"
    else:
        run.error_code = None
        run.error_message = None
    run.attempts = int(getattr(run, "attempts", 0) or 0) + 1
    from app.core.config import settings
    from app.services.kaspi_import_run_utils import compute_next_poll_at

    run.next_poll_at = compute_next_poll_at(
        now=run.last_checked_at,
        status=run.status,
        attempts=run.attempts,
        base_delay_seconds=settings.KASPI_IMPORT_POLL_BACKOFF_BASE_SECONDS,
        max_delay_seconds=settings.KASPI_IMPORT_POLL_BACKOFF_MAX_SECONDS,
        result_payload=run.result_json,
    )
    await session.commit()
    await session.refresh(run)

    return KaspiImportRunPollOut(
        ok=True,
        import_code=run.import_code,
        status=run.status,
        kaspi_import_code=run.kaspi_import_code,
        request_id=request_id,
        status_payload=status_response,
        result_payload=result_response,
    )


async def kaspi_products_import_status(
    request: Request,
    import_code: str = Query(..., alias="i"),
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    from app.services.kaspi_goods_import_client import (
        KaspiGoodsImportClient,
        KaspiImportNotAuthenticated,
        KaspiImportUpstreamError,
        KaspiImportUpstreamUnavailable,
    )

    client = KaspiGoodsImportClient(token=token, base_url="https://kaspi.kz")
    request_id = getattr(getattr(request, "state", None), "request_id", None)
    try:
        response = await client.get_status(import_code=import_code)
    except KaspiImportNotAuthenticated:
        payload = {
            "detail": "NOT_AUTHENTICATED",
            "code": "NOT_AUTHENTICATED",
            "request_id": request_id,
            "errors": [{"code": "NOT_AUTHENTICATED", "detail": "NOT_AUTHENTICATED", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content=payload)
    except KaspiImportUpstreamUnavailable:
        payload = {
            "detail": "kaspi_upstream_unavailable",
            "code": "upstream_unavailable",
            "request_id": request_id,
            "errors": [
                {
                    "code": "upstream_unavailable",
                    "detail": "kaspi_upstream_unavailable",
                    "request_id": request_id,
                }
            ],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload)
    except KaspiImportUpstreamError:
        payload = {
            "detail": "kaspi_upstream_error",
            "code": "upstream_error",
            "request_id": request_id,
            "errors": [{"code": "upstream_error", "detail": "kaspi_upstream_error", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload)

    status_value = response.get("status") or "unknown"
    return KaspiGoodsStatusOut(import_code=import_code, status=str(status_value), payload=response)


async def kaspi_products_import_schema(
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsImportClient(token=token, base_url="https://kaspi.kz")
    try:
        return await client.get_schema()
    except KaspiImportNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc
    except KaspiImportUpstreamUnavailable:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_unavailable")
    except KaspiImportUpstreamError:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_error")


async def kaspi_products_import_result(
    request: Request,
    import_code: str = Query(..., alias="i"),
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    from app.services.kaspi_goods_import_client import (
        KaspiGoodsImportClient,
        KaspiImportNotAuthenticated,
        KaspiImportUpstreamError,
        KaspiImportUpstreamUnavailable,
    )

    client = KaspiGoodsImportClient(token=token, base_url="https://kaspi.kz")
    request_id = getattr(getattr(request, "state", None), "request_id", None)
    try:
        response = await client.get_result(import_code=import_code)
    except KaspiImportNotAuthenticated:
        payload = {
            "detail": "NOT_AUTHENTICATED",
            "code": "NOT_AUTHENTICATED",
            "request_id": request_id,
            "errors": [{"code": "NOT_AUTHENTICATED", "detail": "NOT_AUTHENTICATED", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content=payload)
    except KaspiImportUpstreamUnavailable:
        payload = {
            "detail": "kaspi_upstream_unavailable",
            "code": "upstream_unavailable",
            "request_id": request_id,
            "errors": [
                {
                    "code": "upstream_unavailable",
                    "detail": "kaspi_upstream_unavailable",
                    "request_id": request_id,
                }
            ],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload)
    except KaspiImportUpstreamError:
        payload = {
            "detail": "kaspi_upstream_error",
            "code": "upstream_error",
            "request_id": request_id,
            "errors": [{"code": "upstream_error", "detail": "kaspi_upstream_error", "request_id": request_id}],
        }
        return JSONResponse(status_code=status.HTTP_502_BAD_GATEWAY, content=payload)

    status_value = response.get("status") or "unknown"
    return KaspiGoodsResultOut(import_code=import_code, status=str(status_value), payload=response)


register_kaspi_catalog_routes(
    router,
    kaspi_offers_rebuild=kaspi_offers_rebuild,
    kaspi_offers_import=kaspi_offers_import,
    kaspi_offers_preview=kaspi_offers_preview,
    kaspi_products_sync=kaspi_products_sync,
    kaspi_products_import_start=kaspi_products_import_start,
    kaspi_products_import_upload=kaspi_products_import_upload,
    kaspi_products_import_poll=kaspi_products_import_poll,
    kaspi_products_import_status=kaspi_products_import_status,
    kaspi_products_import_schema=kaspi_products_import_schema,
    kaspi_products_import_result=kaspi_products_import_result,
    kaspi_offers_rebuild_out_model=KaspiOffersRebuildOut,
    kaspi_offers_import_out_model=KaspiOffersImportOut,
    kaspi_offers_preview_out_model=KaspiOffersPreviewOut,
    kaspi_catalog_pull_unsupported_out_model=KaspiCatalogPullUnsupportedOut,
    kaspi_import_run_out_model=KaspiImportRunOut,
    kaspi_import_run_poll_out_model=KaspiImportRunPollOut,
    kaspi_goods_status_out_model=KaspiGoodsStatusOut,
    kaspi_goods_result_out_model=KaspiGoodsResultOut,
)


# ============================= GOODS API ==============================


async def kaspi_goods_schema(
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        return await client.get_schema()
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc


async def kaspi_goods_categories(
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        return await client.get_categories()
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc


async def kaspi_goods_attributes(
    category: str,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        return await client.get_attributes(category_code=category)
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc


async def kaspi_goods_attribute_values(
    category: str,
    attribute: str,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        return await client.get_attribute_values(category_code=category, attribute_code=attribute)
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc


async def kaspi_goods_import_upload(
    file: UploadFile = File(...),
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    current_user: User = Depends(require_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="file_is_empty")

    filename = file.filename or "kaspi_goods_import"
    content_type = file.content_type or "application/octet-stream"

    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        response = await client.post_import_upload(
            filename=filename,
            file_bytes=file_bytes,
            content_type=content_type,
        )
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc

    import_code = _extract_import_code(response)
    if not import_code:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_import_code_missing")

    status_value = response.get("status") or "submitted"

    record = KaspiGoodsImport(
        company_id=company_id,
        created_by_user_id=current_user.id,
        merchant_uid=merchant_uid,
        filename=filename,
        import_code=str(import_code),
        status=str(status_value),
        source="upload",
        raw_response=json.dumps(response, ensure_ascii=False),
    )
    session.add(record)
    await session.commit()
    await session.refresh(record)

    return KaspiGoodsUploadOut(import_code=str(import_code), status=str(status_value), payload=response)


async def kaspi_goods_import_status_by_code(
    import_code: str = Query(..., alias="importCode"),
    current_user: User = Depends(require_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        response = await client.get_import_status_by_code(import_code=import_code)
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc

    status_value = response.get("status") or "unknown"

    record = (
        (
            await session.execute(
                sa.select(KaspiGoodsImport).where(
                    sa.and_(KaspiGoodsImport.company_id == company_id, KaspiGoodsImport.import_code == import_code)
                )
            )
        )
        .scalars()
        .first()
    )
    if record:
        record.status = str(status_value)
        record.status_json = response
        record.last_checked_at = datetime.utcnow()
        await session.commit()

    return KaspiGoodsStatusOut(import_code=import_code, status=str(status_value), payload=response)


async def kaspi_goods_import(
    body: KaspiGoodsImportIn,
    current_user: User = Depends(require_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)

    payload: list[dict[str, Any]]
    if body.payload:
        payload = body.payload
    elif body.product_ids:
        res = await session.execute(
            sa.select(Product).where(sa.and_(Product.company_id == company_id, Product.id.in_(body.product_ids)))
        )
        products = res.scalars().all()
        if not products:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="products_not_found")
        payload = [_product_to_goods_payload(p) for p in products]
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="payload_or_product_ids_required")

    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        response = await client.post_import(payload, content_type=body.content_type)
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc

    import_code = _extract_import_code(response)
    if not import_code:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_import_code_missing")

    status_value = response.get("status") or "submitted"

    record = KaspiGoodsImport(
        company_id=company_id,
        created_by_user_id=current_user.id,
        import_code=str(import_code),
        status=str(status_value),
        request_payload=payload,
        result_payload=None,
        last_error=None,
    )
    session.add(record)
    await session.commit()

    return KaspiGoodsImportOut(ok=True, import_code=str(import_code), status=str(status_value))


async def kaspi_goods_import_status(
    code: str,
    current_user: User = Depends(require_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        response = await client.get_import_status(import_code=code)
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc

    status_value = response.get("status") or "unknown"

    res = await session.execute(
        sa.select(KaspiGoodsImport).where(
            sa.and_(KaspiGoodsImport.company_id == company_id, KaspiGoodsImport.import_code == code)
        )
    )
    record = res.scalars().first()
    if record:
        record.status = str(status_value)
        record.result_payload = response
        await session.commit()

    return KaspiGoodsStatusOut(import_code=code, status=str(status_value), payload=response)


async def kaspi_goods_import_result(
    code: str,
    current_user: User = Depends(require_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    _, token = await _resolve_kaspi_token(session, company_id)
    client = KaspiGoodsClient(token=token, base_url="https://kaspi.kz")
    try:
        response = await client.get_import_result(import_code=code)
    except KaspiNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc

    status_value = response.get("status") or "unknown"

    res = await session.execute(
        sa.select(KaspiGoodsImport).where(
            sa.and_(KaspiGoodsImport.company_id == company_id, KaspiGoodsImport.import_code == code)
        )
    )
    record = res.scalars().first()
    if record:
        record.status = str(status_value)
        record.result_payload = response
        await session.commit()

    return KaspiGoodsResultOut(import_code=code, status=str(status_value), payload=response)


async def kaspi_goods_import_create(
    body: KaspiGoodsImportCreateIn,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    merchant_uid = (body.merchant_uid or "").strip()
    if not merchant_uid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_merchant_uid")
    source = (body.source or "db").strip() or "db"
    if source != "db":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_source")

    company_id = _resolve_company_id(current_user)
    company = await session.get(Company, company_id)
    _, token = await _resolve_kaspi_token(session, company_id)

    from app.services.kaspi_goods_import_client import (
        KaspiGoodsImportClient,
        KaspiImportNotAuthenticated,
        KaspiImportUpstreamError,
        KaspiImportUpstreamUnavailable,
    )
    from app.services.kaspi_goods_import_service import (
        build_payload_json,
        compute_payload_hash,
        load_offers_payload,
    )

    flags = _get_goods_import_flags(company, merchant_uid)
    payload = await load_offers_payload(
        session,
        company_id=company_id,
        merchant_uid=merchant_uid,
        include_price=flags["include_price"],
        include_stock=flags["include_stock"],
    )
    if not payload:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="offers_not_found")

    payload_json = build_payload_json(payload)
    payload_hash = compute_payload_hash(payload_json)

    client = KaspiGoodsImportClient(token=token, base_url="https://kaspi.kz")
    try:
        response = await client.submit_import(payload_json)
    except KaspiImportNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc
    except KaspiImportUpstreamUnavailable:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_unavailable")
    except KaspiImportUpstreamError:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_error")

    import_code = _extract_import_code(response)
    if not import_code:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_import_code_missing")

    status_value = response.get("status") or "UPLOADED"
    now = datetime.utcnow()

    record = KaspiGoodsImport(
        company_id=company_id,
        created_by_user_id=current_user.id,
        merchant_uid=merchant_uid,
        import_code=str(import_code),
        status=str(status_value),
        source=source,
        comment=body.comment,
        payload_hash=payload_hash,
        attempts=1,
        request_json=payload,
        raw_status_json=response,
        status_json=response,
        result_json=None,
        error_code=None,
        error_message=None,
        last_checked_at=now,
        request_payload=payload,
        result_payload=None,
        last_error=None,
    )
    session.add(record)
    await session.commit()
    await session.refresh(record)

    return _goods_import_to_out(record)


async def kaspi_goods_import_list(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)

    result = await session.execute(
        sa.select(KaspiGoodsImport)
        .where(KaspiGoodsImport.company_id == company_id)
        .order_by(KaspiGoodsImport.created_at.desc())
        .limit(limit)
        .offset(offset)
    )
    records = result.scalars().all()
    return [_goods_import_to_out(record) for record in records]


async def kaspi_goods_import_get(
    import_id: str,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)

    record = (
        (
            await session.execute(
                sa.select(KaspiGoodsImport).where(
                    KaspiGoodsImport.company_id == company_id,
                    KaspiGoodsImport.id == import_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import_not_found")
    return _goods_import_to_out(record)


async def kaspi_goods_import_refresh(
    import_id: str,
    request: Request,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_GOODS_IMPORTS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)

    record = (
        (
            await session.execute(
                sa.select(KaspiGoodsImport).where(
                    KaspiGoodsImport.company_id == company_id,
                    KaspiGoodsImport.id == import_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import_not_found")

    if not record.import_code:
        rid = getattr(getattr(request, "state", None), "request_id", None) or request.headers.get("X-Request-ID")
        if not rid:
            rid = str(uuid4())
        payload = {
            "detail": "kaspi_import_missing_code",
            "code": "kaspi_import_missing_code",
            "request_id": rid or "",
        }
        return JSONResponse(status_code=status.HTTP_409_CONFLICT, content=payload, headers={"X-Request-ID": rid})

    _, token = await _resolve_kaspi_token(session, company_id)
    from app.services.kaspi_goods_import_client import (
        KaspiGoodsImportClient,
        KaspiImportNotAuthenticated,
        KaspiImportUpstreamError,
        KaspiImportUpstreamUnavailable,
    )

    client = KaspiGoodsImportClient(token=token, base_url="https://kaspi.kz")
    now = datetime.utcnow()
    try:
        status_response = await client.get_status(import_code=record.import_code)
    except KaspiImportNotAuthenticated as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc
    except KaspiImportUpstreamUnavailable:
        record.error_code = "upstream_unavailable"
        record.error_message = "kaspi_upstream_unavailable"
        record.last_checked_at = now
        await session.commit()
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_unavailable")
    except KaspiImportUpstreamError:
        record.error_code = "upstream_error"
        record.error_message = "kaspi_upstream_error"
        record.last_checked_at = now
        await session.commit()
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_error")

    status_value = status_response.get("status") or record.status
    record.status = str(status_value)
    record.raw_status_json = status_response
    record.status_json = status_response
    record.error_code = None
    record.error_message = None
    record.last_checked_at = now
    await session.commit()
    await session.refresh(record)

    return _goods_import_to_out(record)


async def kaspi_sync_now(
    request: Request,
    body: KaspiSyncNowIn | None = Body(None),
    merchant_uid: str | None = Query(None, min_length=1, alias="merchantUid"),
    timeout_sec: float = Query(SYNC_NOW_TIMEOUT_SEC, ge=0.1, le=60.0),
    hard: int = Query(0, ge=0, le=1),
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_SYNC_NOW)),
    session: AsyncSession = Depends(get_async_db),
):
    refresh_once = True
    if body is not None:
        refresh_once = bool(body.refresh_once)
    query_merchant_uid = ((body.merchant_uid if body else merchant_uid) or "").strip()
    resolved_company_id = resolve_tenant_company_id(current_user, not_found_detail="Company not set")
    company = await session.get(Company, resolved_company_id)
    merchant_uid, merchant_source = _resolve_merchant_uid(
        company=company,
        raw_merchant_uid=query_merchant_uid,
        missing_status=status.HTTP_400_BAD_REQUEST,
        missing_detail="missing_merchant_uid",
    )

    company_id = _resolve_company_id(current_user)
    request_id = getattr(getattr(request, "state", None), "request_id", None)
    rid = request_id or request.headers.get("X-Request-ID") or str(uuid4())

    phase = "init"

    def _err(code: str, detail: str, phase_for_error: str, status: str | None = None) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "code": code,
            "detail": detail,
            "phase": phase_for_error,
            "request_id": rid,
        }
        if status is not None:
            payload["status"] = status
        return payload

    def _with_phase(payload: dict[str, Any]) -> dict[str, Any]:
        if "phase" not in payload:
            payload["phase"] = phase
        return payload

    started_mono = time.perf_counter()
    outcome: str | None = None
    http_status: int | None = None
    exc_type: str | None = None
    exc_repr: str | None = None
    _safe_log_info(
        "kaspi_sync_now_enter",
        request_id=rid,
        company_id=company_id,
        merchant_uid=merchant_uid,
        merchant_uid_source=merchant_source,
        hard=hard,
        timeout_sec=timeout_sec,
        phase=phase,
        monotonic_start=started_mono,
    )

    async def _run_sync_now_bounded() -> KaspiSyncNowOut | JSONResponse:
        lock_acquired = False
        token: str | None = None
        try:
            try:
                _store_name, token = await _resolve_kaspi_token(session, company_id)
            except HTTPException as exc:
                if exc.status_code in {status.HTTP_404_NOT_FOUND, status.HTTP_409_CONFLICT}:
                    payload = _with_phase(
                        {
                            "detail": "kaspi_not_configured",
                            "code": "kaspi_not_configured",
                            "request_id": rid,
                        }
                    )
                    return JSONResponse(
                        status_code=status.HTTP_409_CONFLICT,
                        content=payload,
                        headers={"X-Request-ID": rid},
                    )
                raise

            lock_acquired = await _try_sync_now_lock(session, company_id=company_id, merchant_uid=merchant_uid)
            if not lock_acquired:
                payload = _with_phase(
                    {
                        "detail": "kaspi_sync_in_progress",
                        "code": "kaspi_sync_in_progress",
                        "request_id": rid,
                    }
                )
                return JSONResponse(
                    status_code=status.HTTP_409_CONFLICT,
                    content=payload,
                    headers={"X-Request-ID": rid},
                )

            async def _run_sync_now() -> KaspiSyncNowOut:
                nonlocal phase
                started_at = time.perf_counter()
                logger.info(
                    "kaspi_sync_now start",
                    extra={"company_id": company_id, "merchant_uid": merchant_uid, "request_id": rid},
                )

                svc = KaspiService()
                try:
                    svc_timeout = float(getattr(svc, "_sync_timeout_seconds", timeout_sec) or timeout_sec)
                except Exception:
                    svc_timeout = float(timeout_sec)
                svc._sync_timeout_seconds = max(0.1, min(svc_timeout, float(timeout_sec)))

                safety_margin = 1.5
                budgets = _compute_sync_now_budgets(
                    timeout_sec,
                    safety_margin=safety_margin,
                    default_orders=12.0,
                    default_goods=10.0,
                    default_feed=3.0,
                )
                budget_total = budgets["budget_total"]
                orders_timeout_sec = budgets["orders_timeout_sec"]
                goods_timeout_sec = budgets["goods_timeout_sec"]
                feed_timeout_sec = budgets["feed_timeout_sec"]
                orders_budget = budgets["orders_budget"]
                final_orders_timeout = budgets["final_orders_timeout"]
                orders_http_timeout = svc._orders_timeout(final_orders_timeout)

                logger.info(
                    "kaspi_sync_now budgets",
                    extra={
                        "company_id": company_id,
                        "merchant_uid": merchant_uid,
                        "request_id": rid,
                        "orders_timeout_sec": orders_timeout_sec,
                        "goods_timeout_sec": goods_timeout_sec,
                        "feed_timeout_sec": feed_timeout_sec,
                        "safety_margin_sec": safety_margin,
                        "budget_total_sec": budget_total,
                        "orders_budget": orders_budget,
                        "orders_final_timeout_sec": final_orders_timeout,
                        "orders_http_connect": getattr(orders_http_timeout, "connect", None),
                        "orders_http_read": getattr(orders_http_timeout, "read", None),
                        "orders_http_write": getattr(orders_http_timeout, "write", None),
                        "orders_http_pool": getattr(orders_http_timeout, "pool", None),
                    },
                )

                errors: list[dict[str, Any]] = []
                orders_timed_out = False
                try:
                    phase = "orders_sync"
                    orders_result = await asyncio.wait_for(
                        svc.sync_orders(
                            db=session,
                            company_id=company_id,
                            merchant_uid=merchant_uid,
                            request_id=rid,
                            timeout_seconds=final_orders_timeout,
                            orders_max_attempts=1,
                            client_retries=0,
                        ),
                        timeout=final_orders_timeout,
                    )
                except asyncio.TimeoutError:
                    orders_timed_out = True
                    orders_result = {
                        "ok": False,
                        "status": "timeout",
                        "code": "upstream_timeout",
                        "detail": "Kaspi orders sync timed out",
                        "request_id": rid,
                    }
                    errors.append(
                        _err(
                            code="upstream_timeout",
                            detail="Kaspi orders sync timed out",
                            phase_for_error="orders_sync",
                            status="timeout",
                        )
                    )
                    elapsed_ms = int((time.perf_counter() - started_at) * 1000)
                    logger.warning(
                        "kaspi_sync_now orders timeout",
                        extra={
                            "company_id": company_id,
                            "merchant_uid": merchant_uid,
                            "request_id": rid,
                            "elapsed_ms": elapsed_ms,
                        },
                    )
                    phase = "goods_import"
                    await _record_kaspi_event(
                        session,
                        company_id=company_id,
                        kind="kaspi_orders_sync",
                        status="timeout",
                        request_id=rid,
                        merchant_uid=merchant_uid,
                        error_code="upstream_timeout",
                        error_message="Kaspi orders sync timed out",
                        meta_json={"source": "sync_now"},
                    )

                if not isinstance(orders_result, dict) or "ok" not in orders_result:
                    orders_result = {
                        "ok": True,
                        "status": "success",
                        "result": orders_result,
                    }
                else:
                    status_value = str(orders_result.get("status") or "")
                    code_value = str(orders_result.get("code") or "")
                    if (not orders_timed_out) and (
                        status_value == "timeout" or code_value in {"timeout", "read_timeout", "connect_timeout"}
                    ):
                        orders_timed_out = True
                        duration_ms = orders_result.get("duration_ms")
                        orders_result = {
                            "ok": False,
                            "status": "timeout",
                            "code": "upstream_timeout",
                            "detail": "Kaspi orders sync timed out",
                            "request_id": rid,
                        }
                        if duration_ms is not None:
                            orders_result["duration_ms"] = duration_ms
                        errors.append(
                            _err(
                                code="upstream_timeout",
                                detail="Kaspi orders sync timed out",
                                phase_for_error="orders_sync",
                            )
                        )
                        elapsed_ms = int((time.perf_counter() - started_at) * 1000)
                        logger.warning(
                            "kaspi_sync_now orders timeout",
                            extra={
                                "company_id": company_id,
                                "merchant_uid": merchant_uid,
                                "request_id": rid,
                                "elapsed_ms": elapsed_ms,
                            },
                        )
                        phase = "goods_import"
                        await _record_kaspi_event(
                            session,
                            company_id=company_id,
                            kind="kaspi_orders_sync",
                            status="timeout",
                            request_id=rid,
                            merchant_uid=merchant_uid,
                            error_code="upstream_timeout",
                            error_message="Kaspi orders sync timed out",
                            meta_json={"source": "sync_now"},
                        )

                orders_failed = orders_timed_out or (
                    isinstance(orders_result, dict) and orders_result.get("ok") is False
                )
                remaining = None
                if orders_failed:
                    elapsed_sec = time.perf_counter() - started_at
                    remaining = max(0.0, float(timeout_sec) - elapsed_sec - safety_margin)
                    if remaining <= 0:
                        phase = "goods_import"
                        errors.insert(
                            0,
                            _err(
                                code="kaspi_sync_timeout",
                                detail="Kaspi sync now timed out",
                                phase_for_error="goods_import",
                                status="timeout",
                            ),
                        )
                        if hard:
                            hard_payload = _with_phase(
                                {
                                    "detail": "kaspi_sync_timeout",
                                    "code": "kaspi_sync_timeout",
                                    "request_id": rid,
                                    "ok": False,
                                }
                            )
                            return JSONResponse(
                                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                                content=hard_payload,
                                headers={"X-Request-ID": rid},
                            )
                        payload = _with_phase(
                            {
                                "ok": True,
                                "status": "partial",
                                "errors": errors,
                                "company_id": company_id,
                                "merchant_uid": merchant_uid,
                                "orders_sync": orders_result,
                                "goods_import_result": {
                                    "status": "skipped",
                                    "code": "kaspi_sync_timeout",
                                    "detail": "Kaspi sync now timed out",
                                    "request_id": rid,
                                },
                                "offers_feed_result": {
                                    "status": "skipped",
                                    "code": "kaspi_sync_timeout",
                                    "detail": "Kaspi sync now timed out",
                                    "phase": "goods_import",
                                    "request_id": rid,
                                },
                            }
                        )
                        return JSONResponse(
                            status_code=status.HTTP_200_OK, content=payload, headers={"X-Request-ID": rid}
                        )

                phase = "goods_import"
                company = await session.get(Company, company_id)
                flags = _get_goods_import_flags(company, merchant_uid)
                payload = await load_offers_payload(
                    session,
                    company_id=company_id,
                    merchant_uid=merchant_uid,
                    include_price=flags["include_price"],
                    include_stock=flags["include_stock"],
                )
                if not payload:
                    rebuild_offers = await _build_offers_from_products(
                        session,
                        company_id=company_id,
                        merchant_uid=merchant_uid,
                        include_inactive=False,
                    )
                    rebuild_summary = await _upsert_kaspi_offers(session, offers=rebuild_offers)
                    if rebuild_summary["fetched"] > 0:
                        await session.commit()
                        payload = await load_offers_payload(
                            session,
                            company_id=company_id,
                            merchant_uid=merchant_uid,
                            include_price=flags["include_price"],
                            include_stock=flags["include_stock"],
                        )
                if not payload:
                    errors.append(
                        _err(
                            code="offers_missing",
                            detail="offers_missing",
                            phase_for_error="goods_import",
                        )
                    )
                    payload = _with_phase(
                        {
                            "ok": True,
                            "status": "partial",
                            "result": "partial",
                            "errors": errors,
                            "company_id": company_id,
                            "merchant_uid": merchant_uid,
                            "orders_sync": orders_result,
                            "goods_import_result": {
                                "status": "skipped",
                                "code": "offers_missing",
                                "detail": "offers_missing",
                                "request_id": rid,
                            },
                            "offers_feed_result": {
                                "status": "skipped",
                                "code": "offers_missing",
                                "detail": "offers_missing",
                                "phase": "goods_import",
                                "request_id": rid,
                            },
                        }
                    )
                    return JSONResponse(status_code=status.HTTP_200_OK, content=payload, headers={"X-Request-ID": rid})

                import_client = KaspiGoodsImportClient(token=token or "", base_url="https://kaspi.kz")

                payload_json = build_payload_json(payload)
                payload_hash = compute_payload_hash(payload_json)

                existing_run = await _find_recent_import_run_by_hash(
                    session,
                    company_id=company_id,
                    merchant_uid=merchant_uid,
                    payload_hash=payload_hash,
                )

                if existing_run:
                    import_run = existing_run
                else:
                    import_run = await _create_import_run(
                        session,
                        company_id=company_id,
                        merchant_uid=merchant_uid,
                        request_id=rid,
                    )
                try:
                    if existing_run:
                        response = existing_run.status_json or {}
                    elif remaining is not None:
                        response = await asyncio.wait_for(import_client.submit_import(payload_json), timeout=remaining)
                    else:
                        response = await import_client.submit_import(payload_json)
                except asyncio.TimeoutError:
                    errors.append(
                        _err(
                            code="kaspi_sync_timeout",
                            detail="Kaspi goods import timed out",
                            phase_for_error="goods_import",
                            status="timeout",
                        )
                    )
                    payload = _with_phase(
                        {
                            "ok": True,
                            "status": "partial",
                            "errors": errors,
                            "company_id": company_id,
                            "merchant_uid": merchant_uid,
                            "orders_sync": orders_result,
                            "goods_import_result": {
                                "status": "timeout",
                                "code": "kaspi_sync_timeout",
                                "detail": "Kaspi goods import timed out",
                                "request_id": rid,
                            },
                            "offers_feed_result": {
                                "status": "skipped",
                                "code": "kaspi_sync_timeout",
                                "detail": "Kaspi goods import timed out",
                                "phase": "goods_import",
                                "request_id": rid,
                            },
                        }
                    )
                    return JSONResponse(status_code=status.HTTP_200_OK, content=payload, headers={"X-Request-ID": rid})
                except KaspiImportNotAuthenticated as exc:
                    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED") from exc
                except KaspiImportUpstreamUnavailable:
                    raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_unavailable")
                except KaspiImportUpstreamError:
                    raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_error")

                import_code = _extract_import_code(response) or (import_run.kaspi_import_code if existing_run else None)
                if not import_code:
                    raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_import_code_missing")

                status_value = response.get("status") or "UPLOADED"
                now = datetime.utcnow()

                if not existing_run:
                    import_run.kaspi_import_code = str(import_code)
                    import_run.status = str(status_value)
                    import_run.payload_hash = payload_hash
                    import_run.request_payload = payload
                    import_run.status_json = response
                    import_run.last_checked_at = now
                    import_run.error_code = None
                    import_run.error_message = None
                    import_run.attempts = int(getattr(import_run, "attempts", 0) or 0) + 1
                    from app.core.config import settings
                    from app.services.kaspi_import_run_utils import compute_next_poll_at

                    import_run.next_poll_at = compute_next_poll_at(
                        now=now,
                        status=import_run.status,
                        attempts=import_run.attempts,
                        base_delay_seconds=settings.KASPI_IMPORT_POLL_BACKOFF_BASE_SECONDS,
                        max_delay_seconds=settings.KASPI_IMPORT_POLL_BACKOFF_MAX_SECONDS,
                        result_payload=import_run.result_json,
                    )
                    await session.commit()
                    await session.refresh(import_run)

                if refresh_once:
                    try:
                        if remaining is not None:
                            status_response = await asyncio.wait_for(
                                import_client.get_status(import_code=str(import_code)),
                                timeout=remaining,
                            )
                        else:
                            status_response = await import_client.get_status(import_code=str(import_code))
                        import_run.status = str(status_response.get("status") or import_run.status)
                        import_run.status_json = status_response
                        import_run.last_checked_at = datetime.utcnow()
                        from app.services.kaspi_import_run_utils import is_terminal_import_status

                        if is_terminal_import_status(import_run.status):
                            import_run.result_json = await import_client.get_result(import_code=str(import_code))
                        from app.core.config import settings
                        from app.services.kaspi_import_run_utils import compute_next_poll_at

                        import_run.attempts = int(getattr(import_run, "attempts", 0) or 0) + 1
                        import_run.next_poll_at = compute_next_poll_at(
                            now=import_run.last_checked_at,
                            status=import_run.status,
                            attempts=import_run.attempts,
                            base_delay_seconds=settings.KASPI_IMPORT_POLL_BACKOFF_BASE_SECONDS,
                            max_delay_seconds=settings.KASPI_IMPORT_POLL_BACKOFF_MAX_SECONDS,
                            result_payload=import_run.result_json,
                        )
                        await session.commit()
                        await session.refresh(import_run)
                    except asyncio.TimeoutError:
                        errors.append(
                            _err(
                                code="import_pending",
                                detail="kaspi_import_pending",
                                phase_for_error="goods_import",
                            )
                        )
                        payload = _with_phase(
                            {
                                "ok": True,
                                "status": "partial",
                                "result": "partial",
                                "errors": errors,
                                "company_id": company_id,
                                "merchant_uid": merchant_uid,
                                "orders_sync": orders_result,
                                "goods_import_result": {
                                    "status": "pending",
                                    "code": "import_pending",
                                    "detail": "kaspi_import_pending",
                                    "import_code": str(import_run.import_code),
                                    "kaspi_import_code": str(import_run.kaspi_import_code or ""),
                                    "import_status": str(import_run.status or ""),
                                    "result_summary": {
                                        "errors": 0,
                                        "warnings": 0,
                                        "skipped": 0,
                                        "total": 0,
                                    },
                                    "request_id": rid,
                                },
                                "offers_feed_result": {
                                    "status": "skipped",
                                    "code": "import_pending",
                                    "detail": "kaspi_import_pending",
                                    "phase": "goods_import",
                                    "request_id": rid,
                                },
                            }
                        )
                        return JSONResponse(
                            status_code=status.HTTP_200_OK, content=payload, headers={"X-Request-ID": rid}
                        )
                    except KaspiImportNotAuthenticated as exc:
                        raise HTTPException(
                            status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED"
                        ) from exc
                    except KaspiImportUpstreamUnavailable:
                        import_run.error_code = "upstream_unavailable"
                        import_run.error_message = "kaspi_upstream_unavailable"
                        import_run.last_checked_at = datetime.utcnow()
                        await session.commit()
                        raise HTTPException(
                            status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_unavailable"
                        )
                    except KaspiImportUpstreamError:
                        import_run.error_code = "upstream_error"
                        import_run.error_message = "kaspi_upstream_error"
                        import_run.last_checked_at = datetime.utcnow()
                        await session.commit()
                        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_error")

                phase = "offers_feed"
                offers = (
                    (
                        await session.execute(
                            sa.select(KaspiOffer)
                            .where(
                                KaspiOffer.company_id == company_id,
                                KaspiOffer.merchant_uid == merchant_uid,
                            )
                            .order_by(KaspiOffer.updated_at.desc())
                        )
                    )
                    .scalars()
                    .all()
                )
                if not offers:
                    from app.services.kaspi_import_run_utils import (
                        classify_import_result,
                        normalize_import_status,
                    )

                    goods_status_norm = normalize_import_status(import_run.status)
                    goods_class, result_summary = classify_import_result(import_run.status, import_run.result_json)
                    if goods_class == "success_applied":
                        goods_result_status = "success"
                        goods_ok = True
                    elif goods_class == "success_noop":
                        goods_result_status = "noop"
                        goods_ok = True
                    elif goods_class == "failed":
                        goods_result_status = "failed"
                        goods_ok = False
                    else:
                        goods_result_status = "pending"
                        goods_ok = False
                    errors.append(
                        _err(
                            code="offers_missing",
                            detail="offers_missing",
                            phase_for_error="offers_feed",
                        )
                    )
                    payload = _with_phase(
                        {
                            "ok": True,
                            "status": "partial",
                            "result": "partial",
                            "errors": errors,
                            "company_id": company_id,
                            "merchant_uid": merchant_uid,
                            "orders_sync": orders_result,
                            "goods_import_result": {
                                "ok": goods_ok,
                                "status": goods_result_status,
                                "import_id": str(import_run.id),
                                "import_code": str(import_run.import_code),
                                "kaspi_import_code": str(import_run.kaspi_import_code or ""),
                                "import_status": goods_status_norm or import_run.status,
                                "result_summary": result_summary,
                                "code": "import_failed"
                                if goods_class == "failed"
                                else "import_pending"
                                if goods_class == "pending"
                                else "import_noop"
                                if goods_class == "success_noop"
                                else None,
                                "detail": "kaspi_import_failed"
                                if goods_class == "failed"
                                else "kaspi_import_pending"
                                if goods_class == "pending"
                                else "kaspi_import_noop"
                                if goods_class == "success_noop"
                                else None,
                                "request_id": rid if goods_class != "success_applied" else None,
                            },
                            "offers_feed_result": {
                                "status": "skipped",
                                "code": "offers_missing",
                                "detail": "offers_missing",
                                "phase": "offers_feed",
                                "request_id": rid,
                            },
                        }
                    )
                    return JSONResponse(status_code=status.HTTP_200_OK, content=payload, headers={"X-Request-ID": rid})

                company_name = (company.name if company else None) or f"Company {company_id}"
                _build_kaspi_offers_xml(offers, company=company_name, merchant_id=merchant_uid)

                settings_obj = _load_company_settings(company)
                generated_at = datetime.utcnow()
                settings_obj["kaspi.feed_last_generated_at"] = generated_at.isoformat()
                settings_obj["kaspi.feed_last_generated_merchant_uid"] = merchant_uid
                if company is not None:
                    company.settings = json.dumps(settings_obj, ensure_ascii=False, separators=(",", ":"))
                    await session.commit()

                from app.services.kaspi_import_run_utils import (
                    classify_import_result,
                    normalize_import_status,
                )

                goods_status_norm = normalize_import_status(import_run.status)
                goods_class, result_summary = classify_import_result(import_run.status, import_run.result_json)

                status_value = "ok" if goods_class == "success_applied" and not orders_timed_out else "partial"
                if goods_class == "failed":
                    errors.append(
                        _err(
                            code="import_failed",
                            detail="kaspi_import_failed",
                            phase_for_error="goods_import",
                        )
                    )
                elif goods_class == "pending":
                    errors.append(
                        _err(
                            code="import_pending",
                            detail="kaspi_import_pending",
                            phase_for_error="goods_import",
                        )
                    )
                elif goods_class == "success_noop":
                    errors.append(
                        _err(
                            code="import_noop",
                            detail="kaspi_import_noop",
                            phase_for_error="goods_import",
                        )
                    )
                elapsed_ms = int((time.perf_counter() - started_at) * 1000)
                logger.info(
                    "kaspi_sync_now done",
                    extra={
                        "company_id": company_id,
                        "merchant_uid": merchant_uid,
                        "request_id": rid,
                        "status": status_value,
                        "elapsed_ms": elapsed_ms,
                    },
                )

                goods_import_result = {
                    "ok": goods_class in {"success_applied", "success_noop"},
                    "status": "success"
                    if goods_class == "success_applied"
                    else "noop"
                    if goods_class == "success_noop"
                    else "failed"
                    if goods_class == "failed"
                    else "pending",
                    "import_id": str(import_run.id),
                    "import_code": str(import_run.import_code),
                    "kaspi_import_code": str(import_run.kaspi_import_code or ""),
                    "import_status": goods_status_norm or import_run.status,
                    "result_summary": result_summary,
                }
                if goods_class == "failed":
                    goods_import_result.update(
                        {
                            "code": "import_failed",
                            "detail": "kaspi_import_failed",
                            "request_id": rid,
                        }
                    )
                elif goods_class == "pending":
                    goods_import_result.update(
                        {
                            "code": "import_pending",
                            "detail": "kaspi_import_pending",
                            "request_id": rid,
                        }
                    )
                elif goods_class == "success_noop":
                    goods_import_result.update(
                        {
                            "code": "import_noop",
                            "detail": "kaspi_import_noop",
                            "request_id": rid,
                        }
                    )
                offers_feed_result = {
                    "ok": True,
                    "status": "success",
                    "generated_at": generated_at.isoformat(),
                }
                offers_feed_upload_result: dict[str, Any] | None = None
                feed_upload_enabled = bool(getattr(settings, "KASPI_FEED_UPLOAD_ENABLED", False))
                if feed_upload_enabled:
                    try:
                        _xml_body, payload_hash, _last_modified = await _build_offers_xml_for_company(
                            session,
                            company_id=company.id,
                            merchant_uid=merchant_uid,
                        )
                        offers_feed_upload_result = {
                            "status": "success",
                            "mode": "pull",
                            "payload_hash": payload_hash,
                        }
                    except HTTPException as exc:
                        offers_feed_upload_result = {
                            "status": "failed",
                            "mode": "pull",
                            "code": "feed_pull_failed",
                            "detail": str(exc.detail),
                        }
                status_value = "ok" if goods_class == "success_applied" and not orders_timed_out else "partial"
                if feed_upload_enabled and offers_feed_upload_result is not None:
                    if offers_feed_upload_result.get("status") != "success":
                        status_value = "partial"
                        errors.append(
                            _err(
                                code=str(offers_feed_upload_result.get("code") or "feed_upload_pending"),
                                detail=str(offers_feed_upload_result.get("detail") or "feed_upload_pending"),
                                phase_for_error="offers_feed",
                            )
                        )

                return KaspiSyncNowOut(
                    ok=True,
                    status=status_value,
                    phase=phase,
                    errors=errors,
                    company_id=company_id,
                    merchant_uid=merchant_uid,
                    orders_sync=orders_result,
                    goods_import_id=str(import_run.id),
                    goods_import_code=str(import_run.kaspi_import_code or import_run.import_code),
                    goods_import_status=import_run.status,
                    goods_import_result=goods_import_result,
                    feed_last_generated_at=generated_at,
                    offers_feed_result=offers_feed_result,
                    offers_feed_upload_result=offers_feed_upload_result,
                )

            return await _run_sync_now()
        finally:
            if lock_acquired:
                await _release_sync_now_lock(session, company_id=company_id, merchant_uid=merchant_uid)

    try:
        response_obj = await asyncio.wait_for(_run_sync_now_bounded(), timeout=timeout_sec)
        if isinstance(response_obj, JSONResponse):
            http_status = response_obj.status_code
            outcome = "ok" if response_obj.status_code < 400 else "error"
        else:
            status_value = str(getattr(response_obj, "status", "") or "")
            outcome = status_value if status_value in {"ok", "partial", "timeout"} else "ok"
            http_status = status.HTTP_200_OK
        return response_obj
    except asyncio.TimeoutError as exc:
        exc_type = type(exc).__name__
        exc_repr = repr(exc) if _ci_diag_enabled() else None
        _safe_log_warning(
            "kaspi_sync_now_outer_timeout",
            request_id=rid,
            company_id=company_id,
            merchant_uid=merchant_uid,
            phase=phase,
            exc_type=exc_type,
            exc_repr=exc_repr,
            task_cancelled=_task_cancelled_flag(),
        )
        payload = {
            "ok": True,
            "status": "partial",
            "phase": phase,
            "errors": [
                _err(
                    code="kaspi_sync_timeout",
                    detail="Kaspi sync now timed out",
                    phase_for_error=phase,
                    status="timeout",
                )
            ],
            "company_id": company_id,
            "merchant_uid": merchant_uid,
            "orders_sync": {
                "status": "timeout",
                "code": "kaspi_sync_timeout",
                "detail": "Kaspi sync now timed out",
                "request_id": rid,
            },
            "goods_import_result": {
                "status": "skipped",
                "code": "kaspi_sync_timeout",
                "detail": "Kaspi sync now timed out",
                "request_id": rid,
            },
            "offers_feed_result": {
                "status": "skipped",
                "code": "kaspi_sync_timeout",
                "detail": "Kaspi sync now timed out",
                "phase": phase,
                "request_id": rid,
            },
        }
        logger.warning(
            "kaspi_sync_now orchestration timeout",
            extra={"company_id": company_id, "merchant_uid": merchant_uid, "request_id": rid, "phase": phase},
        )
        if hard:
            hard_payload = _with_phase(
                {
                    "detail": "kaspi_sync_timeout",
                    "code": "kaspi_sync_timeout",
                    "request_id": rid,
                    "ok": False,
                }
            )
            http_status = status.HTTP_504_GATEWAY_TIMEOUT
            outcome = "timeout"
            return JSONResponse(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                content=hard_payload,
                headers={"X-Request-ID": rid},
            )
        http_status = status.HTTP_200_OK
        outcome = "partial"
        return JSONResponse(status_code=status.HTTP_200_OK, content=payload, headers={"X-Request-ID": rid})
    except asyncio.CancelledError as exc:
        exc_type = type(exc).__name__
        exc_repr = repr(exc) if _ci_diag_enabled() else None
        _safe_log_warning(
            "kaspi_sync_now_cancelled",
            request_id=rid,
            company_id=company_id,
            merchant_uid=merchant_uid,
            phase=phase,
            exc_type=exc_type,
            exc_repr=exc_repr,
            task_cancelled=_task_cancelled_flag(),
        )
        raise
    except httpx.TimeoutException as exc:
        exc_type = type(exc).__name__
        exc_repr = repr(exc) if _ci_diag_enabled() else None
        _safe_log_warning(
            "kaspi_sync_now_httpx_timeout",
            request_id=rid,
            company_id=company_id,
            merchant_uid=merchant_uid,
            phase=phase,
            exc_type=exc_type,
            exc_repr=exc_repr,
            task_cancelled=_task_cancelled_flag(),
        )
        raise
    except Exception as exc:
        exc_type = type(exc).__name__
        exc_repr = repr(exc) if _ci_diag_enabled() else None
        if _HAS_EXCEPTION_GROUP and isinstance(exc, ExceptionGroup):  # type: ignore[name-defined]
            extra = _exception_group_info(exc)
            _safe_log_warning(
                "kaspi_sync_now_exception_group",
                request_id=rid,
                company_id=company_id,
                merchant_uid=merchant_uid,
                phase=phase,
                exc_type=exc_type,
                exc_repr=exc_repr,
                task_cancelled=_task_cancelled_flag(),
                **extra,
            )
            raise
        _safe_log_warning(
            "kaspi_sync_now_unhandled_error",
            request_id=rid,
            company_id=company_id,
            merchant_uid=merchant_uid,
            phase=phase,
            exc_type=exc_type,
            exc_repr=exc_repr,
            task_cancelled=_task_cancelled_flag(),
        )
        raise
    finally:
        finished_mono = time.perf_counter()
        elapsed_ms = int((finished_mono - started_mono) * 1000)
        _safe_log_info(
            "kaspi_sync_now_exit",
            request_id=rid,
            company_id=company_id,
            merchant_uid=merchant_uid,
            hard=hard,
            timeout_sec=timeout_sec,
            phase=phase,
            monotonic_start=started_mono,
            monotonic_end=finished_mono,
            elapsed_ms=elapsed_ms,
            http_status=http_status,
            outcome=outcome,
            exc_type=exc_type,
            exc_repr=exc_repr,
        )


async def kaspi_token_health(
    request: Request,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    store_name, token = await _resolve_kaspi_token(session, company_id)
    request_id = getattr(getattr(request, "state", None), "request_id", None)

    now = datetime.utcnow()
    ge_ms = int((now - timedelta(days=14)).timestamp() * 1000)
    le_ms = int(now.timestamp() * 1000)

    orders_url = "https://kaspi.kz/shop/api/v2/orders"
    orders_params = _build_kaspi_orders_params(ge_ms=ge_ms, le_ms=le_ms, state="NEW", page_number=0, page_size=1)

    orders_headers = {
        "X-Auth-Token": token,
        "Accept": "application/vnd.api+json",
    }

    goods_headers = {
        "X-Auth-Token": token,
        "Accept": "application/json",
    }

    try:
        async with _build_kaspi_httpx_client() as client:
            started = time.perf_counter()
            orders_resp = await client.get(orders_url, headers=orders_headers, params=orders_params)
            _log_kaspi_probe_response(
                request_id=request_id,
                company_id=company_id,
                store_name=store_name,
                method="GET",
                url=orders_url,
                status_code=orders_resp.status_code,
                response_text=orders_resp.text,
                elapsed_ms=int((time.perf_counter() - started) * 1000),
            )

            started = time.perf_counter()
            goods_url = "https://kaspi.kz/shop/api/products/import/schema"
            goods_resp = await client.get(goods_url, headers=goods_headers)
            _log_kaspi_probe_response(
                request_id=request_id,
                company_id=company_id,
                store_name=store_name,
                method="GET",
                url=goods_url,
                status_code=goods_resp.status_code,
                response_text=goods_resp.text,
                elapsed_ms=int((time.perf_counter() - started) * 1000),
            )
    except httpx.TimeoutException as exc:
        root_type, root_message = _extract_httpx_root_cause(exc)
        error_kind = _classify_httpx_error(exc, root_type, root_message)
        _log_kaspi_probe_error(
            request_id=request_id,
            company_id=company_id,
            store_name=store_name,
            method="GET",
            url=orders_url,
            exc=exc,
        )
        await KaspiStoreToken.update_selftest(
            session,
            store_name,
            status="upstream_unavailable",
            error_code=error_kind,
            error_message="upstream_unavailable",
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="upstream_unavailable")
    except httpx.RequestError as exc:
        root_type, root_message = _extract_httpx_root_cause(exc)
        error_kind = _classify_httpx_error(exc, root_type, root_message)
        _log_kaspi_probe_error(
            request_id=request_id,
            company_id=company_id,
            store_name=store_name,
            method="GET",
            url=orders_url,
            exc=exc,
        )
        await KaspiStoreToken.update_selftest(
            session,
            store_name,
            status="upstream_unavailable",
            error_code=error_kind,
            error_message="upstream_unavailable",
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="upstream_unavailable")

    if orders_resp.status_code in {401, 403} or goods_resp.status_code in {401, 403}:
        await KaspiStoreToken.update_selftest(
            session,
            store_name,
            status="invalid_token",
            error_code="NOT_AUTHENTICATED",
            error_message="NOT_AUTHENTICATED",
        )
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED")

    if not (200 <= orders_resp.status_code < 300) or not (200 <= goods_resp.status_code < 300):
        await KaspiStoreToken.update_selftest(
            session,
            store_name,
            status="upstream_unavailable",
            error_code="http_status",
            error_message="upstream_unavailable",
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="upstream_unavailable")

    await KaspiStoreToken.update_selftest(
        session,
        store_name,
        status="ok",
        error_code=None,
        error_message=None,
    )

    return KaspiTokenHealthOut(
        ok=True,
        orders_http=orders_resp.status_code,
        goods_http=goods_resp.status_code,
        cause=None,
    )


async def kaspi_token_selftest(
    request: Request,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    store_name, token = await _resolve_kaspi_token(session, company_id)

    now = datetime.utcnow()
    ge_ms = int((now - timedelta(days=14)).timestamp() * 1000)
    le_ms = int(now.timestamp() * 1000)

    orders_url = "https://kaspi.kz/shop/api/v2/orders"
    orders_params = _build_kaspi_orders_params(ge_ms=ge_ms, le_ms=le_ms, state="NEW", page_number=0, page_size=1)
    orders_headers = {
        "X-Auth-Token": token,
        "Accept": "application/vnd.api+json",
        "Content-Type": "application/vnd.api+json",
    }

    goods_headers = {
        "X-Auth-Token": token,
        "Accept": "application/json",
    }

    request_id = getattr(getattr(request, "state", None), "request_id", None)

    try:
        async with _build_kaspi_httpx_client() as client:
            started = time.perf_counter()
            orders_resp = await client.get(orders_url, headers=orders_headers, params=orders_params)
            _log_kaspi_probe_response(
                request_id=request_id,
                company_id=company_id,
                store_name=store_name,
                method="GET",
                url=orders_url,
                status_code=orders_resp.status_code,
                response_text=orders_resp.text,
                elapsed_ms=int((time.perf_counter() - started) * 1000),
            )

            goods_schema_url = "https://kaspi.kz/shop/api/products/import/schema"
            started = time.perf_counter()
            goods_schema_resp = await client.get(goods_schema_url, headers=goods_headers)
            _log_kaspi_probe_response(
                request_id=request_id,
                company_id=company_id,
                store_name=store_name,
                method="GET",
                url=goods_schema_url,
                status_code=goods_schema_resp.status_code,
                response_text=goods_schema_resp.text,
                elapsed_ms=int((time.perf_counter() - started) * 1000),
            )

            goods_categories_url = "https://kaspi.kz/shop/api/products/classification/categories"
            started = time.perf_counter()
            goods_categories_resp = await client.get(goods_categories_url, headers=goods_headers)
            _log_kaspi_probe_response(
                request_id=request_id,
                company_id=company_id,
                store_name=store_name,
                method="GET",
                url=goods_categories_url,
                status_code=goods_categories_resp.status_code,
                response_text=goods_categories_resp.text,
                elapsed_ms=int((time.perf_counter() - started) * 1000),
            )
    except httpx.TimeoutException as exc:
        root_type, root_message = _extract_httpx_root_cause(exc)
        error_kind = _classify_httpx_error(exc, root_type, root_message)
        _log_kaspi_probe_error(
            request_id=request_id,
            company_id=company_id,
            store_name=store_name,
            method="GET",
            url=orders_url,
            exc=exc,
        )
        await KaspiStoreToken.update_selftest(
            session,
            store_name,
            status="upstream_unavailable",
            error_code=error_kind,
            error_message="upstream_unavailable",
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            kind="kaspi_selftest",
            status="failed",
            request_id=request_id,
            error_code=error_kind,
            error_message="upstream_unavailable",
            meta_json={"orders_http": 0, "goods_schema_http": 0, "goods_categories_http": 0},
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="upstream_unavailable")
    except httpx.RequestError as exc:
        root_type, root_message = _extract_httpx_root_cause(exc)
        error_kind = _classify_httpx_error(exc, root_type, root_message)
        _log_kaspi_probe_error(
            request_id=request_id,
            company_id=company_id,
            store_name=store_name,
            method="GET",
            url=orders_url,
            exc=exc,
        )
        await KaspiStoreToken.update_selftest(
            session,
            store_name,
            status="upstream_unavailable",
            error_code=error_kind,
            error_message="upstream_unavailable",
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            kind="kaspi_selftest",
            status="failed",
            request_id=request_id,
            error_code=error_kind,
            error_message="upstream_unavailable",
            meta_json={"orders_http": 0, "goods_schema_http": 0, "goods_categories_http": 0},
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="upstream_unavailable")

    if orders_resp.status_code in {401, 403}:
        await KaspiStoreToken.update_selftest(
            session,
            store_name,
            status="invalid_token",
            error_code="NOT_AUTHENTICATED",
            error_message="NOT_AUTHENTICATED",
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            kind="kaspi_selftest",
            status="failed",
            request_id=request_id,
            error_code="NOT_AUTHENTICATED",
            error_message="NOT_AUTHENTICATED",
            meta_json={
                "orders_http": orders_resp.status_code,
                "goods_schema_http": goods_schema_resp.status_code,
                "goods_categories_http": goods_categories_resp.status_code,
            },
        )
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="NOT_AUTHENTICATED")

    orders_error = None
    status_value = "ok"
    if orders_resp.status_code >= 400:
        orders_error = "orders_request_failed"
        status_value = "failed"

    goods_access = None
    if orders_resp.status_code == 200 and (
        goods_schema_resp.status_code in {401, 403} or goods_categories_resp.status_code in {401, 403}
    ):
        goods_access = "missing_or_not_enabled"

    if not (200 <= orders_resp.status_code < 300):
        await KaspiStoreToken.update_selftest(
            session,
            store_name,
            status="upstream_unavailable",
            error_code="http_status",
            error_message="upstream_unavailable",
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            kind="kaspi_selftest",
            status="failed",
            request_id=request_id,
            error_code="http_status",
            error_message="upstream_unavailable",
            meta_json={
                "orders_http": orders_resp.status_code,
                "goods_schema_http": goods_schema_resp.status_code,
                "goods_categories_http": goods_categories_resp.status_code,
            },
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="upstream_unavailable")

    await KaspiStoreToken.update_selftest(
        session,
        store_name,
        status=status_value,
        error_code=orders_error,
        error_message=orders_error,
    )
    await _record_kaspi_event(
        session,
        company_id=company_id,
        kind="kaspi_selftest",
        status="success" if status_value == "ok" else "failed",
        request_id=request_id,
        error_code=orders_error,
        error_message=orders_error,
        meta_json={
            "orders_http": orders_resp.status_code,
            "goods_schema_http": goods_schema_resp.status_code,
            "goods_categories_http": goods_categories_resp.status_code,
        },
    )

    return KaspiTokenSelftestOut(
        orders_http=orders_resp.status_code,
        goods_schema_http=goods_schema_resp.status_code,
        goods_categories_http=goods_categories_resp.status_code,
        goods_access=goods_access,
        orders_error=orders_error,
    )


async def kaspi_catalog_import(
    file: UploadFile = File(...),
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    dry_run: bool = Query(False, alias="dry_run"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)

    if not merchant_uid or not merchant_uid.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_merchant_uid")

    merchant_uid = merchant_uid.strip()

    company_id = _resolve_company_id(current_user)

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="empty_file")

    filename = file.filename or "catalog.csv"
    rows = parse_catalog_file(content, filename)

    batch: CatalogImportBatch | None = None
    if not dry_run:
        content_hash = sha256(content).hexdigest()
        batch = CatalogImportBatch(
            company_id=company_id,
            source="kaspi",
            filename=filename,
            content_hash=content_hash,
            status="RUNNING",
            merchant_uid=merchant_uid,
            started_at=datetime.utcnow(),
        )
        session.add(batch)
        await session.commit()
        await session.refresh(batch)

    rows_total = len(rows)
    rows_ok = 0
    rows_skipped = 0
    error_counts: dict[str, int] = {}
    row_records: list[dict[str, Any]] = []
    offer_records: list[dict[str, Any]] = []

    for idx, row in enumerate(rows, start=1):
        raw = row.get("raw")
        sku = row.get("sku")
        master_sku = row.get("master_sku")
        title = row.get("title")
        price = row.get("price")
        old_price = row.get("old_price")
        stock_count = row.get("stock_count")
        pre_order = row.get("pre_order")
        stock_specified = row.get("stock_specified")
        updated_at = row.get("updated_at")

        error = None
        if not sku:
            error = "missing_sku"
        if error:
            rows_skipped += 1
            error_counts[error] = error_counts.get(error, 0) + 1
            if not dry_run and batch is not None:
                row_records.append(
                    {
                        "batch_id": batch.id,
                        "company_id": company_id,
                        "row_num": idx,
                        "raw": raw,
                        "sku": sku,
                        "master_sku": master_sku,
                        "title": title,
                        "price": price,
                        "old_price": old_price,
                        "stock_count": stock_count,
                        "pre_order": pre_order,
                        "stock_specified": stock_specified,
                        "updated_at": updated_at,
                        "error": error,
                    }
                )
            continue

        rows_ok += 1
        if not dry_run and batch is not None:
            row_records.append(
                {
                    "batch_id": batch.id,
                    "company_id": company_id,
                    "row_num": idx,
                    "raw": raw,
                    "sku": sku,
                    "master_sku": master_sku,
                    "title": title,
                    "price": price,
                    "old_price": old_price,
                    "stock_count": stock_count,
                    "pre_order": pre_order,
                    "stock_specified": stock_specified,
                    "updated_at": updated_at,
                    "error": None,
                }
            )
            offer_records.append(
                {
                    "company_id": company_id,
                    "merchant_uid": merchant_uid,
                    "sku": sku,
                    "master_sku": master_sku,
                    "title": title,
                    "price": price,
                    "old_price": old_price,
                    "stock_count": stock_count,
                    "pre_order": pre_order,
                    "stock_specified": stock_specified,
                    "raw": raw,
                    "updated_at": datetime.utcnow(),
                }
            )

    top_errors = [
        {"error": key, "count": count}
        for key, count in sorted(error_counts.items(), key=lambda item: (-item[1], item[0]))
    ][:5]

    if dry_run or batch is None:
        return KaspiCatalogImportOut(
            batch_id=None,
            status="DRY_RUN",
            rows_total=rows_total,
            rows_ok=rows_ok,
            rows_skipped=rows_skipped,
            top_errors=top_errors,
            dry_run=True,
        )

    try:
        if row_records:
            await session.execute(sa.insert(CatalogImportRow), row_records)

        if offer_records:
            deduped: dict[tuple[int, str, str], dict[str, Any]] = {}
            for record in offer_records:
                key = (record["company_id"], record["merchant_uid"], record["sku"])
                deduped[key] = record
            deduped_records = list(deduped.values())

            for chunk_start in range(0, len(deduped_records), 500):
                chunk = deduped_records[chunk_start : chunk_start + 500]
                stmt = sa.dialects.postgresql.insert(KaspiOffer).values(chunk)
                stmt = stmt.on_conflict_do_update(
                    index_elements=["company_id", "merchant_uid", "sku"],
                    set_={
                        "master_sku": sa.func.coalesce(
                            sa.func.nullif(stmt.excluded.master_sku, ""),
                            KaspiOffer.master_sku,
                        ),
                        "title": sa.func.coalesce(sa.func.nullif(stmt.excluded.title, ""), KaspiOffer.title),
                        "price": sa.func.coalesce(stmt.excluded.price, KaspiOffer.price),
                        "old_price": sa.func.coalesce(stmt.excluded.old_price, KaspiOffer.old_price),
                        "stock_count": sa.func.coalesce(stmt.excluded.stock_count, KaspiOffer.stock_count),
                        "pre_order": sa.func.coalesce(stmt.excluded.pre_order, KaspiOffer.pre_order),
                        "stock_specified": sa.func.coalesce(stmt.excluded.stock_specified, KaspiOffer.stock_specified),
                        "raw": sa.func.coalesce(stmt.excluded.raw, KaspiOffer.raw),
                        "updated_at": datetime.utcnow(),
                    },
                )
                await session.execute(stmt)

        batch.rows_total = rows_total
        batch.rows_ok = rows_ok
        batch.rows_failed = rows_skipped
        batch.status = "DONE"
        batch.finished_at = datetime.utcnow()
        if rows_skipped:
            batch.error_summary = "; ".join({e["error"] for e in top_errors})

        await session.commit()
    except Exception as exc:
        await session.rollback()
        batch = await session.get(CatalogImportBatch, batch.id)
        if batch:
            batch.status = "FAILED"
            batch.finished_at = datetime.utcnow()
            batch.error_summary = str(exc)[:500]
            await session.commit()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="catalog_import_failed")

    return KaspiCatalogImportOut(
        batch_id=str(batch.id),
        status=batch.status,
        rows_total=rows_total,
        rows_ok=rows_ok,
        rows_skipped=rows_skipped,
        top_errors=top_errors,
        dry_run=False,
    )


register_kaspi_goods_routes(
    router,
    kaspi_goods_schema=kaspi_goods_schema,
    kaspi_goods_categories=kaspi_goods_categories,
    kaspi_goods_attributes=kaspi_goods_attributes,
    kaspi_goods_attribute_values=kaspi_goods_attribute_values,
    kaspi_goods_import_upload=kaspi_goods_import_upload,
    kaspi_goods_import_status_by_code=kaspi_goods_import_status_by_code,
    kaspi_goods_import=kaspi_goods_import,
    kaspi_goods_import_status=kaspi_goods_import_status,
    kaspi_goods_import_result=kaspi_goods_import_result,
    kaspi_goods_import_create=kaspi_goods_import_create,
    kaspi_goods_import_list=kaspi_goods_import_list,
    kaspi_goods_import_get=kaspi_goods_import_get,
    kaspi_goods_import_refresh=kaspi_goods_import_refresh,
    kaspi_sync_now=kaspi_sync_now,
    kaspi_token_health=kaspi_token_health,
    kaspi_token_selftest=kaspi_token_selftest,
    kaspi_catalog_import=kaspi_catalog_import,
    kaspi_goods_upload_out_model=KaspiGoodsUploadOut,
    kaspi_goods_status_out_model=KaspiGoodsStatusOut,
    kaspi_goods_import_out_model=KaspiGoodsImportOut,
    kaspi_goods_result_out_model=KaspiGoodsResultOut,
    kaspi_goods_import_record_out_model=KaspiGoodsImportRecordOut,
    kaspi_sync_now_out_model=KaspiSyncNowOut,
    kaspi_token_health_out_model=KaspiTokenHealthOut,
    kaspi_token_selftest_out_model=KaspiTokenSelftestOut,
    kaspi_catalog_import_out_model=KaspiCatalogImportOut,
)


# ============================= MC SESSION + SYNC =============================
#
# TODO(MVP-DEFERRED): Kaspi Merchant Cabinet (MC) cookie-based automation is
# experimental and deferred. MVP uses official flows only: Orders API, XML
# feed generation/public link, and goods import/export. MC automation would
# require browser automation (e.g., Playwright) and is out of scope for now.


async def kaspi_mc_session_upsert(
    payload: KaspiMcSessionIn,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    merchant_uid = (payload.merchant_uid or "").strip()
    cookies = (payload.cookies or "").strip()
    if not merchant_uid or not cookies:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_merchant_uid_or_cookies")

    row = await KaspiMcSession.upsert_session(
        session,
        company_id=company_id,
        merchant_uid=merchant_uid,
        cookies=cookies,
        is_active=True,
    )

    return KaspiMcSessionOut(
        merchant_uid=row.merchant_uid,
        is_active=row.is_active,
        created_at=row.created_at,
        updated_at=row.updated_at,
        last_used_at=row.last_used_at,
        last_error=row.last_error,
        cookies_masked=_mask_secret(cookies),
    )


async def kaspi_mc_session_status(
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    q = sa.select(KaspiMcSession).where(KaspiMcSession.company_id == company_id)
    if merchant_uid:
        q = q.where(KaspiMcSession.merchant_uid == merchant_uid.strip())

    rows = (await session.execute(q)).scalars().all()
    if merchant_uid and not rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="mc_session_not_found")

    items: list[KaspiMcSessionOut] = []
    for row in rows:
        masked = None
        if row.is_active:
            cookies = await KaspiMcSession.get_cookies(
                session,
                company_id=company_id,
                merchant_uid=row.merchant_uid,
            )
            masked = _mask_secret(cookies)
        items.append(
            KaspiMcSessionOut(
                merchant_uid=row.merchant_uid,
                is_active=row.is_active,
                created_at=row.created_at,
                updated_at=row.updated_at,
                last_used_at=row.last_used_at,
                last_error=row.last_error,
                cookies_masked=masked,
            )
        )

    return KaspiMcSessionListOut(items=items)


async def kaspi_catalog_sync_mc(
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    if not merchant_uid or not merchant_uid.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_merchant_uid")
    merchant_uid = merchant_uid.strip()

    row = (
        (
            await session.execute(
                sa.select(KaspiMcSession).where(
                    KaspiMcSession.company_id == company_id,
                    KaspiMcSession.merchant_uid == merchant_uid,
                )
            )
        )
        .scalars()
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="mc_session_not_found")
    if not row.is_active:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="mc_session_inactive")

    cookies = await KaspiMcSession.get_cookies(
        session,
        company_id=company_id,
        merchant_uid=merchant_uid,
    )
    if not cookies:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="mc_session_not_configured")

    try:
        summary = await sync_kaspi_mc_offers(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            cookies=cookies,
        )
    except httpx.HTTPStatusError:
        await mark_mc_session_error(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            error="kaspi_mc_upstream_error",
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_mc_upstream_error")
    except httpx.RequestError:
        await mark_mc_session_error(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            error="kaspi_mc_upstream_unavailable",
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_mc_upstream_unavailable")

    return KaspiMcSyncOut(**summary)


async def kaspi_catalog_import_batches(
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    result = await session.execute(
        sa.select(CatalogImportBatch)
        .where(CatalogImportBatch.company_id == company_id)
        .order_by(CatalogImportBatch.created_at.desc())
        .limit(limit)
        .offset(offset)
    )
    batches = result.scalars().all()

    return [
        KaspiCatalogImportBatchOut(
            id=str(batch.id),
            merchant_uid=batch.merchant_uid,
            filename=batch.filename,
            status=batch.status,
            rows_total=batch.rows_total,
            rows_ok=batch.rows_ok,
            rows_failed=batch.rows_failed,
            started_at=batch.started_at,
            finished_at=batch.finished_at,
            created_at=batch.created_at,
            error_summary=batch.error_summary,
        )
        for batch in batches
    ]


async def kaspi_catalog_import_batch_detail(
    batch_id: str,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    result = await session.execute(
        sa.select(CatalogImportBatch).where(
            CatalogImportBatch.company_id == company_id,
            CatalogImportBatch.id == batch_id,
        )
    )
    batch = result.scalars().first()
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="batch_not_found")

    duration_seconds = None
    if batch.started_at and batch.finished_at:
        duration_seconds = int((batch.finished_at - batch.started_at).total_seconds())

    return KaspiCatalogImportBatchDetailOut(
        id=str(batch.id),
        merchant_uid=batch.merchant_uid,
        filename=batch.filename,
        status=batch.status,
        rows_total=batch.rows_total,
        rows_ok=batch.rows_ok,
        rows_failed=batch.rows_failed,
        started_at=batch.started_at,
        finished_at=batch.finished_at,
        created_at=batch.created_at,
        error_summary=batch.error_summary,
        duration_seconds=duration_seconds,
    )


async def kaspi_catalog_import_batch_errors(
    batch_id: str,
    limit: int = 200,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    batch = (
        (
            await session.execute(
                sa.select(CatalogImportBatch).where(
                    CatalogImportBatch.company_id == company_id,
                    CatalogImportBatch.id == batch_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="batch_not_found")

    result = await session.execute(
        sa.select(CatalogImportRow)
        .where(
            CatalogImportRow.company_id == company_id,
            CatalogImportRow.batch_id == batch_id,
            CatalogImportRow.error.is_not(None),
        )
        .order_by(CatalogImportRow.row_num.asc())
        .limit(limit)
        .offset(offset)
    )
    rows = result.scalars().all()

    return [
        KaspiCatalogImportErrorOut(
            row_num=row.row_num,
            error=row.error,
            sku=row.sku,
            master_sku=row.master_sku,
            title=row.title,
            raw=_truncate_raw(row.raw),
        )
        for row in rows
    ]


async def kaspi_offers_list(
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    q: str | None = None,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    conditions = [KaspiOffer.company_id == company_id]
    if merchant_uid:
        conditions.append(KaspiOffer.merchant_uid == merchant_uid)
    if q:
        like = f"%{q}%"
        conditions.append(sa.or_(KaspiOffer.sku.ilike(like), KaspiOffer.title.ilike(like)))

    total = (await session.execute(sa.select(sa.func.count()).select_from(KaspiOffer).where(*conditions))).scalar_one()

    result = await session.execute(
        sa.select(KaspiOffer).where(*conditions).order_by(KaspiOffer.updated_at.desc()).limit(limit).offset(offset)
    )
    offers = result.scalars().all()

    items = [
        KaspiOfferOut(
            id=offer.id,
            merchant_uid=offer.merchant_uid,
            sku=offer.sku,
            master_sku=offer.master_sku,
            title=offer.title,
            price=float(offer.price) if offer.price is not None else None,
            old_price=float(offer.old_price) if offer.old_price is not None else None,
            stock_count=offer.stock_count,
            pre_order=offer.pre_order,
            stock_specified=offer.stock_specified,
            updated_at=offer.updated_at,
        )
        for offer in offers
    ]

    return KaspiOfferListOut(items=items, total=int(total or 0), limit=limit, offset=offset)


register_kaspi_mc_routes(
    router,
    kaspi_mc_session_upsert=kaspi_mc_session_upsert,
    kaspi_mc_session_status=kaspi_mc_session_status,
    kaspi_catalog_sync_mc=kaspi_catalog_sync_mc,
    kaspi_catalog_import_batches=kaspi_catalog_import_batches,
    kaspi_catalog_import_batch_detail=kaspi_catalog_import_batch_detail,
    kaspi_catalog_import_batch_errors=kaspi_catalog_import_batch_errors,
    kaspi_offers_list=kaspi_offers_list,
    kaspi_mc_session_out_model=KaspiMcSessionOut,
    kaspi_mc_session_list_out_model=KaspiMcSessionListOut,
    kaspi_mc_sync_out_model=KaspiMcSyncOut,
    kaspi_catalog_import_batch_out_model=KaspiCatalogImportBatchOut,
    kaspi_catalog_import_batch_detail_out_model=KaspiCatalogImportBatchDetailOut,
    kaspi_catalog_import_error_out_model=KaspiCatalogImportErrorOut,
    kaspi_offer_list_out_model=KaspiOfferListOut,
)


class KaspiOfferSeedIn(BaseModel):
    merchant_uid: str = Field(..., min_length=1)
    sku: str | None = None
    title: str | None = None
    price: int | None = Field(None, ge=0)


class KaspiOfferSeedOut(BaseModel):
    created: bool
    merchant_uid: str
    sku: str
    offer_id: int | None = None


_FEED_UPLOAD_SUCCESS = {"done", "success", "completed", "published"}
_FEED_UPLOAD_FAILED = {"failed", "error"}


def _classify_feed_upload_status(status: str | None) -> str:
    value = (status or "").strip().lower()
    if value in _FEED_UPLOAD_SUCCESS:
        return "success"
    if value in _FEED_UPLOAD_FAILED:
        return "failed"
    return "pending"


def _build_feed_upload_env(token: str) -> dict[str, str]:
    base_url = (settings.KASPI_FEED_BASE_URL or "https://kaspi.kz").rstrip("/")
    upload_path = settings.KASPI_FEED_UPLOAD_PATH or "/shop/api/feeds/import"
    upload_url = (
        settings.KASPI_FEED_UPLOAD_URL
        or f"{base_url}{upload_path if upload_path.startswith('/') else '/' + upload_path}"
    )
    status_url = settings.KASPI_FEED_STATUS_URL or f"{base_url}/shop/api/feeds/import/status"
    result_url = settings.KASPI_FEED_RESULT_URL or f"{base_url}/shop/api/feeds/import/result"
    return {
        "KASPI_FEED_UPLOAD_URL": upload_url,
        "KASPI_FEED_STATUS_URL": status_url,
        "KASPI_FEED_RESULT_URL": result_url,
        "KASPI_FEED_TOKEN": token,
        "KASPI_TOKEN": token,
    }


def _format_http_datetime(dt: datetime) -> str:
    return formatdate(timeval=dt.timestamp(), usegmt=True)


def _parse_http_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        parsed = parsedate_to_datetime(value)
        if parsed.tzinfo:
            return parsed.astimezone(UTC).replace(tzinfo=None)
        return parsed
    except Exception:
        return None


def _public_feed_etag(payload_hash: str) -> str:
    return f'"{payload_hash}"'


async def _load_public_offers_feed(
    session: AsyncSession,
    *,
    token_value: str,
    merchant_uid: str | None = None,
) -> tuple[KaspiFeedPublicToken, str, str, datetime]:
    token_hash = sha256(token_value.encode("utf-8")).hexdigest()
    token_row = (
        (
            await session.execute(
                sa.select(KaspiFeedPublicToken).where(
                    KaspiFeedPublicToken.token_hash == token_hash,
                    KaspiFeedPublicToken.revoked_at.is_(None),
                )
            )
        )
        .scalars()
        .first()
    )
    if not token_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="not_found")

    effective_merchant_uid = (merchant_uid or "").strip()
    if effective_merchant_uid:
        if not token_row.merchant_uid or token_row.merchant_uid != effective_merchant_uid:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="not_found")
    else:
        effective_merchant_uid = (token_row.merchant_uid or "").strip()
        if not effective_merchant_uid:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="not_found")

    xml_body, payload_hash, last_modified = await _build_offers_xml_for_company(
        session,
        company_id=token_row.company_id,
        merchant_uid=effective_merchant_uid,
    )
    return token_row, xml_body, payload_hash, last_modified


async def _build_offers_xml_for_company(
    session: AsyncSession,
    *,
    company_id: int,
    merchant_uid: str,
) -> tuple[str, str, datetime]:
    offers = (
        (
            await session.execute(
                sa.select(KaspiOffer)
                .where(
                    KaspiOffer.company_id == company_id,
                    KaspiOffer.merchant_uid == merchant_uid,
                )
                .order_by(KaspiOffer.updated_at.desc())
            )
        )
        .scalars()
        .all()
    )
    if not offers:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="not_found")

    company = await session.get(Company, company_id)
    company_name = (company.name if company else None) or f"Company {company_id}"
    xml_body = _build_kaspi_offers_xml(offers, company=company_name, merchant_id=merchant_uid)
    payload_hash = compute_feed_payload_hash(xml_body)
    last_modified = max((offer.updated_at for offer in offers if offer.updated_at), default=datetime.utcnow())
    return xml_body, payload_hash, last_modified


def _log_public_feed_access(
    *,
    request: Request,
    token_row: KaspiFeedPublicToken,
    status_code: int,
    size_bytes: int,
) -> None:
    request_id = getattr(getattr(request, "state", None), "request_id", None) or request.headers.get("X-Request-ID")
    _safe_log_info(
        "kaspi_public_feed_access",
        request_id=request_id,
        company_id=token_row.company_id,
        token_id=token_row.id,
        status_code=status_code,
        bytes=size_bytes,
        user_agent=request.headers.get("User-Agent"),
    )


async def kaspi_offers_seed(
    payload: KaspiOfferSeedIn,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    if not _is_dev_environment():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="not_found")

    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)
    merchant_uid = payload.merchant_uid.strip()

    existing = await session.execute(
        sa.select(KaspiOffer)
        .where(KaspiOffer.company_id == company_id, KaspiOffer.merchant_uid == merchant_uid)
        .limit(1)
    )
    offer = existing.scalars().first()
    if offer:
        return KaspiOfferSeedOut(created=False, merchant_uid=merchant_uid, sku=offer.sku, offer_id=offer.id)

    sku = (payload.sku or f"SMOKE-{uuid4().hex[:8]}").strip()
    title = (payload.title or "Smoke Item").strip()
    price = int(payload.price) if payload.price is not None else 1000

    offer = KaspiOffer(
        company_id=company_id,
        merchant_uid=merchant_uid,
        sku=sku,
        title=title,
        price=price,
        stock_count=1,
        stock_specified=True,
    )
    session.add(offer)
    await session.commit()
    await session.refresh(offer)

    return KaspiOfferSeedOut(created=True, merchant_uid=merchant_uid, sku=sku, offer_id=offer.id)


TEMPLATE_RESPONSES = {
    200: {
        "description": "Template file (CSV or XLSX)",
        "content": {
            "text/csv": {
                "schema": {"type": "string", "format": "binary"},
            },
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                "schema": {"type": "string", "format": "binary"},
            },
        },
    }
}

LEGACY_CSV_RESPONSES = {
    200: {
        "description": "CSV template file",
        "content": {
            "text/csv": {
                "schema": {"type": "string", "format": "binary"},
            }
        },
    }
}


async def kaspi_catalog_import_template_csv(
    current_user: User = Depends(get_current_user),
):
    await _require_store_admin_company_scoped(current_user)
    content = _kaspi_catalog_template_csv()
    headers = {"Content-Disposition": "attachment; filename=kaspi_catalog_template.csv"}
    return Response(content=content, media_type="text/csv; charset=utf-8", headers=headers)


KASPI_CATALOG_TEMPLATE_HEADERS: list[str] = [
    "sku",
    "master_sku",
    "title",
    "price",
    "old_price",
    "stock_count",
    "pre_order",
    "stock_specified",
    "updated_at",
]


def _kaspi_catalog_template_csv() -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(KASPI_CATALOG_TEMPLATE_HEADERS)
    writer.writerow(
        [
            "SKU-001",
            "MASTER-001",
            "Sample title",
            "1000",
            "1200",
            "5",
            "false",
            "true",
            "2026-01-17T12:00:00",
        ]
    )
    return output.getvalue()


def _kaspi_catalog_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "template"
    ws.append(KASPI_CATALOG_TEMPLATE_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def kaspi_catalog_import_template(
    format: Literal["csv", "xlsx"] = Query("xlsx"),
    current_user: User = Depends(get_current_user),
):
    await _require_store_admin_company_scoped(current_user)
    if format == "csv":
        content = _kaspi_catalog_template_csv()
        headers = {"Content-Disposition": "attachment; filename=kaspi_catalog_template.csv"}
        return Response(content=content, media_type="text/csv; charset=utf-8", headers=headers)

    content = _kaspi_catalog_template_xlsx()
    headers = {"Content-Disposition": "attachment; filename=kaspi_catalog_template.xlsx"}
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


async def kaspi_products_list(
    limit: int = 50,
    offset: int = 0,
    q: str | None = None,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    """
    Возвращает список продуктов каталога Kaspi для текущей компании.

    Args:
        limit: Максимум записей (default 50, max 200)
        offset: Смещение для пагинации (default 0)
        q: Опциональный поиск по name/sku (ILIKE)

    Returns:
        Список продуктов с безопасными полями (без raw)
    """
    company_id = _resolve_company_id(current_user)

    # Validate limit
    limit = max(1, min(limit, 200))
    offset = max(0, offset)

    try:
        # Build query
        query = sa.select(KaspiCatalogProduct).where(KaspiCatalogProduct.company_id == company_id)

        # Optional search
        if q:
            search_pattern = f"%{q}%"
            query = query.where(
                sa.or_(
                    KaspiCatalogProduct.name.ilike(search_pattern),
                    KaspiCatalogProduct.sku.ilike(search_pattern),
                )
            )

        # Count total
        count_query = sa.select(sa.func.count()).select_from(query.subquery())
        total_result = await session.execute(count_query)
        total = total_result.scalar() or 0

        # Apply pagination
        query = query.limit(limit).offset(offset).order_by(KaspiCatalogProduct.id)

        # Execute
        result = await session.execute(query)
        products = result.scalars().all()

        # Map to response model (safe fields only)
        items = [
            KaspiProductOut(
                offer_id=p.offer_id,
                name=p.name,
                sku=p.sku,
                price=str(p.price) if p.price is not None else None,
                qty=p.qty,
                is_active=p.is_active,
            )
            for p in products
        ]

        return KaspiProductListOut(
            items=items,
            total=total,
            limit=limit,
            offset=offset,
        )

    except Exception as e:
        logger.error("Kaspi products list failed: company_id=%s error=%s", company_id, e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve products",
        )


register_kaspi_tooling_routes(
    router,
    kaspi_offers_seed=kaspi_offers_seed,
    kaspi_catalog_import_template_csv=kaspi_catalog_import_template_csv,
    kaspi_catalog_import_template=kaspi_catalog_import_template,
    kaspi_products_list=kaspi_products_list,
    kaspi_offer_seed_out_model=KaspiOfferSeedOut,
    kaspi_product_list_out_model=KaspiProductListOut,
    legacy_csv_responses=LEGACY_CSV_RESPONSES,
    template_responses=TEMPLATE_RESPONSES,
)


# ============================= FEED EXPORTS ==============================


class KaspiFeedExportOut(BaseModel):
    """Response model for feed export metadata with retry diagnostics."""

    id: int
    kind: str
    format: str
    status: str  # generated, uploading, uploaded, failed
    checksum: str
    stats_json: dict | None = None
    last_error: str | None = None
    attempts: int = 0
    last_attempt_at: str | None = None
    uploaded_at: str | None = None
    duration_ms: int | None = None
    created_at: str | None = None
    updated_at: str | None = None


class KaspiFeedPublicTokenIn(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    merchant_uid: str = Field(..., min_length=1, alias="merchantUid")
    comment: str | None = None


class KaspiFeedPublicTokenOut(BaseModel):
    id: int
    merchant_uid: str | None = None
    token: str | None = None
    created_at: datetime
    revoked_at: datetime | None = None
    last_used_at: datetime | None = None
    comment: str | None = None


class KaspiFeedPublicTokenListOut(BaseModel):
    items: list[KaspiFeedPublicTokenOut]


async def _rotate_public_feed_token(
    session: AsyncSession,
    *,
    company_id: int,
    merchant_uid: str,
    comment: str | None,
) -> tuple[KaspiFeedPublicToken, str]:
    now = datetime.utcnow()
    existing = (
        (
            await session.execute(
                sa.select(KaspiFeedPublicToken).where(
                    KaspiFeedPublicToken.company_id == company_id,
                    KaspiFeedPublicToken.merchant_uid == merchant_uid,
                    KaspiFeedPublicToken.revoked_at.is_(None),
                )
            )
        )
        .scalars()
        .all()
    )
    for row in existing:
        row.revoked_at = now

    token_value = None
    token_hash = None
    for _ in range(3):
        candidate = secrets.token_urlsafe(32)
        candidate_hash = sha256(candidate.encode("utf-8")).hexdigest()
        exists = (
            await session.execute(
                sa.select(sa.func.count())
                .select_from(KaspiFeedPublicToken)
                .where(KaspiFeedPublicToken.token_hash == candidate_hash)
            )
        ).scalar_one()
        if not exists:
            token_value = candidate
            token_hash = candidate_hash
            break

    if not token_value or not token_hash:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="token_generation_failed")

    token_row = KaspiFeedPublicToken(
        company_id=company_id,
        merchant_uid=merchant_uid,
        token_hash=token_hash,
        comment=comment,
    )
    session.add(token_row)
    await session.commit()
    await session.refresh(token_row)
    return token_row, token_value


class KaspiFeedGenerateOut(BaseModel):
    """Response model for feed generation."""

    ok: bool
    export_id: int
    company_id: int
    total: int
    active: int
    checksum: str
    is_new: bool


class KaspiFeedUploadOut(BaseModel):
    """Response model for feed upload with retry diagnostics."""

    ok: bool
    export_id: int
    status: str
    upload_id: str | None = None
    error: str | None = None
    is_retryable: bool | None = None
    already_uploaded: bool = False
    upload_in_progress: bool = False


class KaspiFeedListOut(BaseModel):
    """Response model for feed exports list."""

    items: list[KaspiFeedExportOut]
    total: int
    limit: int
    offset: int


# ============================= FEED EXPORTS (MVP) =============================


async def kaspi_feed_export_create(
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    store_id: str | None = Query(None, alias="storeId"),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    if not merchant_uid or not merchant_uid.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_merchant_uid")

    company_id = _resolve_company_id(current_user)
    merchant_uid = merchant_uid.strip()

    started_at = datetime.utcnow()
    started_perf = time.perf_counter()

    export = KaspiFeedExport(
        company_id=company_id,
        kind="offers",
        format="xml",
        status="RUNNING",
        checksum="",
        payload_text="",
        stats_json=None,
        last_error=None,
        attempts=0,
        last_attempt_at=started_at,
    )
    session.add(export)
    try:
        await session.commit()
        await session.refresh(export)

        result = await session.execute(
            sa.select(KaspiOffer)
            .where(KaspiOffer.company_id == company_id, KaspiOffer.merchant_uid == merchant_uid)
            .order_by(KaspiOffer.updated_at.desc())
        )
        offers = result.scalars().all()
        company = await session.get(Company, company_id)
        company_name = (company.name if company else None) or f"Company {company_id}"
        xml_body = _build_kaspi_offers_xml(offers, company=company_name, merchant_id=merchant_uid)
        checksum = sha256(xml_body.encode("utf-8")).hexdigest()
        duration_ms = int((time.perf_counter() - started_perf) * 1000)

        export.checksum = checksum
        export.payload_text = xml_body
        export.stats_json = {"total": len(offers), "merchant_uid": merchant_uid, "store_id": store_id}
        export.duration_ms = duration_ms
        export.status = "DONE"
        await session.commit()
        await session.refresh(export)
        return _feed_export_to_out(export)
    except Exception as exc:
        await session.rollback()
        export = await session.get(KaspiFeedExport, export.id)
        if export:
            export.status = "FAILED"
            export.last_error = str(exc)[:500]
            export.duration_ms = int((time.perf_counter() - started_perf) * 1000)
            await session.commit()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="feed_export_failed")


async def kaspi_feed_exports_list(
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    limit = max(1, min(limit, 200))
    offset = max(0, offset)

    result = await session.execute(
        sa.select(KaspiFeedExport)
        .where(KaspiFeedExport.company_id == company_id)
        .order_by(KaspiFeedExport.created_at.desc())
        .limit(limit)
        .offset(offset)
    )
    exports = result.scalars().all()
    return [_feed_export_to_out(export) for export in exports]


async def _create_offers_feed_upload(
    *,
    session: AsyncSession,
    company: Company,
    merchant_uid: str,
    store_name: str,
    token: str,
    request_id: str | None,
    comment: str | None,
    refresh: bool,
) -> tuple[KaspiFeedUpload, bool]:
    if refresh:
        rebuild_offers = await _build_offers_from_products(
            session,
            company_id=company.id,
            merchant_uid=merchant_uid,
            include_inactive=False,
        )
        if rebuild_offers:
            await _upsert_kaspi_offers(session, offers=rebuild_offers)
            await session.commit()

    offers = (
        (
            await session.execute(
                sa.select(KaspiOffer)
                .where(
                    KaspiOffer.company_id == company.id,
                    KaspiOffer.merchant_uid == merchant_uid,
                )
                .order_by(KaspiOffer.updated_at.desc())
            )
        )
        .scalars()
        .all()
    )
    if not offers:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="offers_not_found")

    company_name = (company.name or "").strip() or f"Company {company.id}"
    xml_body = _build_kaspi_offers_xml(offers, company=company_name, merchant_id=merchant_uid)
    payload_hash = compute_feed_payload_hash(xml_body)

    existing = await find_recent_successful_upload_by_hash(
        session,
        company_id=company.id,
        merchant_uid=merchant_uid,
        payload_hash=payload_hash,
    )
    if existing:
        return existing, True

    export = await get_or_create_feed_export(
        session,
        company_id=company.id,
        kind="offers",
        xml_body=xml_body,
    )

    job = await create_feed_upload_job(
        session,
        company_id=company.id,
        merchant_uid=merchant_uid,
        export_id=export.id,
        source="offers",
        request_id=request_id,
        comment=comment,
        payload_hash=payload_hash,
    )

    now_attempt = datetime.utcnow()
    job.attempts = int(job.attempts or 0) + 1
    job.last_attempt_at = now_attempt
    job.updated_at = now_attempt
    await session.commit()
    await session.refresh(job)

    extra_env = _build_feed_upload_env(token)
    upload_url = extra_env.get("KASPI_FEED_UPLOAD_URL")
    if should_block_feed_upload_url(upload_url):
        await update_feed_upload_job(
            session,
            job=job,
            status="failed",
            error_code="feed_upload_endpoint_invalid",
            error_message="feed_upload_endpoint_invalid",
            last_attempt_at=now_attempt,
            next_attempt_at=None,
        )
        return job, False
    tmp_dir = settings.tmp_dir()
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"kaspi_feed_{company.id}_{uuid4().hex}.xml"

    max_attempts = int(getattr(settings, "KASPI_FEED_UPLOAD_MAX_ATTEMPTS", 5) or 5)
    base_delay = int(getattr(settings, "KASPI_FEED_UPLOAD_BACKOFF_BASE_SECONDS", 30) or 30)
    max_delay = int(getattr(settings, "KASPI_FEED_UPLOAD_BACKOFF_MAX_SECONDS", 900) or 900)

    try:
        tmp_path.write_text(xml_body, encoding="utf-8")
        response = KaspiAdapter().feed_upload(
            store_name,
            str(tmp_path),
            comment=comment,
            extra_env=extra_env,
        )
    except KaspiAdapterError as exc:
        error_message = str(exc)[:500]
        unsupported = is_unsupported_content_type_error(error_message=error_message)
        status_value = "failed" if unsupported or job.attempts >= max_attempts else "pending"
        next_attempt = None
        if status_value == "pending":
            next_attempt = compute_next_attempt_at(
                now=now_attempt,
                attempts=job.attempts,
                base_delay_seconds=base_delay,
                max_delay_seconds=max_delay,
            )
        await update_feed_upload_job(
            session,
            job=job,
            status=status_value,
            error_code="unsupported_content_type" if unsupported else "upstream_unavailable",
            error_message=error_message,
            last_attempt_at=now_attempt,
            next_attempt_at=next_attempt,
        )
        return job, False
    except Exception as exc:
        error_message = str(exc)[:500]
        unsupported = is_unsupported_content_type_error(error_message=error_message)
        status_value = "failed" if unsupported or job.attempts >= max_attempts else "pending"
        next_attempt = None
        if status_value == "pending":
            next_attempt = compute_next_attempt_at(
                now=now_attempt,
                attempts=job.attempts,
                base_delay_seconds=base_delay,
                max_delay_seconds=max_delay,
            )
        await update_feed_upload_job(
            session,
            job=job,
            status=status_value,
            error_code="unsupported_content_type" if unsupported else "upstream_unavailable",
            error_message=error_message,
            last_attempt_at=now_attempt,
            next_attempt_at=next_attempt,
        )
        return job, False
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            logger.warning("Kaspi feed upload temp cleanup failed: path=%s", tmp_path)

    normalized = normalize_kaspi_payload(_normalize_kaspi_response(response))
    import_code = _extract_import_code(normalized)
    if not import_code:
        await update_feed_upload_job(
            session,
            job=job,
            status="failed",
            error_code="kaspi_import_code_missing",
            error_message="kaspi_import_code_missing",
            response_json=normalized,
        )
        return job, False

    status_value = str(normalized.get("status") or "uploaded")
    status_class = _classify_feed_upload_status(status_value)
    error_code, error_message = _extract_error_info(normalized)
    if is_unsupported_content_type_error(error_message=error_message, response_payload=normalized):
        error_code = "unsupported_content_type"
        status_value = "failed"
    await update_feed_upload_job(
        session,
        job=job,
        status=status_value,
        import_code=str(import_code),
        error_code=error_code,
        error_message=error_message,
        response_json=normalized,
        next_attempt_at=None
        if status_class == "success" or error_code == "unsupported_content_type"
        else compute_next_attempt_at(
            now=now_attempt,
            attempts=job.attempts,
            base_delay_seconds=base_delay,
            max_delay_seconds=max_delay,
        ),
    )
    return job, False


async def kaspi_offers_feed_upload(
    request: Request,
    body: KaspiOffersFeedUploadIn,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_FEED_UPLOADS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    company = await session.get(Company, company_id)
    merchant_uid = _resolve_kaspi_offers_merchant_uid(
        company=company,
        raw_merchant_uid=body.merchant_uid,
    )

    if body.refresh:
        rebuild_offers = await _build_offers_from_products(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            include_inactive=False,
        )
        if rebuild_offers:
            await _upsert_kaspi_offers(session, offers=rebuild_offers)
            await session.commit()

    xml_body, payload_hash, _last_modified = await _build_offers_xml_for_company(
        session,
        company_id=company_id,
        merchant_uid=merchant_uid,
    )
    token_row, token_value = await _rotate_public_feed_token(
        session,
        company_id=company_id,
        merchant_uid=merchant_uid,
        comment=body.comment,
    )
    public_base = str(settings.PUBLIC_URL or request.base_url).rstrip("/")
    url = f"{public_base}/public/kaspi/price-list/{token_value}.xml"
    generated_at = datetime.utcnow()

    return KaspiOffersFeedUploadOut(
        url=url,
        token_id=token_row.id,
        payload_hash=payload_hash,
        generated_at=generated_at,
        merchant_uid=merchant_uid,
    )


async def kaspi_feed_upload_create(
    request: Request,
    body: KaspiFeedUploadIn,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_FEED_UPLOADS)),
    session: AsyncSession = Depends(get_async_db),
):
    merchant_uid = (body.merchant_uid or "").strip()
    if not merchant_uid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_merchant_uid")

    source = (body.source or "").strip()
    allowed_sources = {"public_token", "export_id", "local_file_path"}
    if source not in allowed_sources:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_source")

    if source == "export_id" and not body.export_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="export_id_required")
    if source == "local_file_path" and not body.local_file_path:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="local_file_path_required")

    company_id = _resolve_company_id(current_user)
    store_name, token = await _resolve_kaspi_token(session, company_id)
    request_id = getattr(getattr(request, "state", None), "request_id", None) or request.headers.get("X-Request-ID")

    existing = await get_feed_upload_by_request_id(session, company_id=company_id, request_id=request_id)
    if existing:
        return _feed_upload_to_out(existing)

    xml_body: str
    if source == "export_id":
        export = await session.get(KaspiFeedExport, body.export_id)
        if not export or export.company_id != company_id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="export_not_found")
        if not export.payload_text:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="export_payload_missing")
        xml_body = export.payload_text
    elif source == "local_file_path":
        file_path = Path(body.local_file_path)
        if not file_path.is_file():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="local_file_not_found")
        xml_body = file_path.read_text(encoding="utf-8", errors="replace")
    else:
        offers = (
            (
                await session.execute(
                    sa.select(KaspiOffer)
                    .where(
                        KaspiOffer.company_id == company_id,
                        KaspiOffer.merchant_uid == merchant_uid,
                    )
                    .order_by(KaspiOffer.updated_at.desc())
                )
            )
            .scalars()
            .all()
        )
        if not offers:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="offers_not_found")
        company = await session.get(Company, company_id)
        company_name = (company.name if company else None) or f"Company {company_id}"
        xml_body = _build_kaspi_offers_xml(offers, company=company_name, merchant_id=merchant_uid)

    payload_hash = compute_feed_payload_hash(xml_body)
    job = await create_feed_upload_job(
        session,
        company_id=company_id,
        merchant_uid=merchant_uid,
        export_id=body.export_id if source == "export_id" else None,
        source=source,
        request_id=request_id,
        comment=body.comment,
        payload_hash=payload_hash,
    )
    now_attempt = datetime.utcnow()
    job.attempts = int(job.attempts or 0) + 1
    job.last_attempt_at = now_attempt
    job.updated_at = now_attempt
    await session.commit()
    await session.refresh(job)

    extra_env = _build_feed_upload_env(token)
    upload_url = extra_env.get("KASPI_FEED_UPLOAD_URL")
    if should_block_feed_upload_url(upload_url):
        await update_feed_upload_job(
            session,
            job=job,
            status="failed",
            error_code="feed_upload_endpoint_invalid",
            error_message="feed_upload_endpoint_invalid",
            last_attempt_at=now_attempt,
            next_attempt_at=None,
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code="feed_upload_endpoint_invalid",
            error_message="feed_upload_endpoint_invalid",
            meta_json={"upload_id": str(job.id), "import_code": None},
        )
        payload = jsonable_encoder(_feed_upload_to_out(job))
        return JSONResponse(status_code=status.HTTP_201_CREATED, content=payload)

    tmp_dir = settings.tmp_dir()
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"kaspi_feed_{company_id}_{uuid4().hex}.xml"

    try:
        tmp_path.write_text(xml_body, encoding="utf-8")
        response = KaspiAdapter().feed_upload(
            store_name,
            str(tmp_path),
            comment=body.comment,
            extra_env=extra_env,
        )
    except KaspiAdapterError as exc:
        error_message = str(exc)[:500]
        unsupported = is_unsupported_content_type_error(error_message=error_message)
        status_value = "failed" if unsupported else "pending"
        next_attempt = None
        if status_value == "pending":
            next_attempt = compute_next_attempt_at(
                now=now_attempt,
                attempts=job.attempts,
                base_delay_seconds=int(getattr(settings, "KASPI_FEED_UPLOAD_BACKOFF_BASE_SECONDS", 30) or 30),
                max_delay_seconds=int(getattr(settings, "KASPI_FEED_UPLOAD_BACKOFF_MAX_SECONDS", 900) or 900),
            )
        error_code = "unsupported_content_type" if unsupported else "upstream_unavailable"
        await update_feed_upload_job(
            session,
            job=job,
            status=status_value,
            error_code=error_code,
            error_message=error_message,
            last_attempt_at=now_attempt,
            next_attempt_at=next_attempt,
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code=error_code,
            error_message=error_message,
            meta_json={"upload_id": str(job.id), "import_code": None},
        )
        payload = jsonable_encoder(_feed_upload_to_out(job))
        return JSONResponse(status_code=status.HTTP_201_CREATED, content=payload)
    except Exception as exc:
        logger.error("Kaspi feed upload failed: company_id=%s error=%s", company_id, exc)
        error_message = str(exc)[:500]
        unsupported = is_unsupported_content_type_error(error_message=error_message)
        status_value = "failed" if unsupported else "pending"
        next_attempt = None
        if status_value == "pending":
            next_attempt = compute_next_attempt_at(
                now=now_attempt,
                attempts=job.attempts,
                base_delay_seconds=int(getattr(settings, "KASPI_FEED_UPLOAD_BACKOFF_BASE_SECONDS", 30) or 30),
                max_delay_seconds=int(getattr(settings, "KASPI_FEED_UPLOAD_BACKOFF_MAX_SECONDS", 900) or 900),
            )
        error_code = "unsupported_content_type" if unsupported else "upstream_unavailable"
        await update_feed_upload_job(
            session,
            job=job,
            status=status_value,
            error_code=error_code,
            error_message=error_message,
            last_attempt_at=now_attempt,
            next_attempt_at=next_attempt,
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code=error_code,
            error_message=error_message,
            meta_json={"upload_id": str(job.id), "import_code": None},
        )
        payload = jsonable_encoder(_feed_upload_to_out(job))
        return JSONResponse(status_code=status.HTTP_201_CREATED, content=payload)
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            logger.warning("Kaspi feed upload temp cleanup failed: path=%s", tmp_path)

    normalized = normalize_kaspi_payload(_normalize_kaspi_response(response))
    import_code = _extract_import_code(normalized)
    if not import_code:
        await update_feed_upload_job(
            session,
            job=job,
            status="failed",
            error_code="kaspi_import_code_missing",
            error_message="kaspi_import_code_missing",
            response_json=normalized,
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code="kaspi_import_code_missing",
            error_message="kaspi_import_code_missing",
            meta_json={"upload_id": str(job.id), "import_code": None},
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_import_code_missing")

    status_value = normalized.get("status") or "uploaded"
    error_code, error_message = _extract_error_info(normalized)
    if is_unsupported_content_type_error(error_message=error_message, response_payload=normalized):
        error_code = "unsupported_content_type"
        status_value = "failed"
    status_class = _classify_feed_upload_status(status_value)
    await update_feed_upload_job(
        session,
        job=job,
        status=str(status_value),
        import_code=str(import_code),
        error_code=error_code,
        error_message=error_message,
        response_json=normalized,
        next_attempt_at=None
        if status_class == "success" or error_code == "unsupported_content_type"
        else compute_next_attempt_at(
            now=now_attempt,
            attempts=job.attempts,
            base_delay_seconds=int(getattr(settings, "KASPI_FEED_UPLOAD_BACKOFF_BASE_SECONDS", 30) or 30),
            max_delay_seconds=int(getattr(settings, "KASPI_FEED_UPLOAD_BACKOFF_MAX_SECONDS", 900) or 900),
        ),
    )
    await _record_kaspi_event(
        session,
        company_id=company_id,
        merchant_uid=merchant_uid,
        kind="kaspi_feed",
        status="success",
        request_id=request_id,
        error_code=error_code,
        error_message=error_message,
        meta_json={"upload_id": str(job.id), "import_code": str(import_code)},
    )

    return _feed_upload_to_out(job)


async def kaspi_feed_uploads_list(
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_FEED_UPLOADS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)

    limit = max(1, min(limit, 200))
    offset = max(0, offset)

    result = await session.execute(
        sa.select(KaspiFeedUpload)
        .where(KaspiFeedUpload.company_id == company_id)
        .order_by(KaspiFeedUpload.created_at.desc())
        .limit(limit)
        .offset(offset)
    )
    uploads = result.scalars().all()
    return [_feed_upload_to_out(record) for record in uploads]


async def kaspi_feed_upload_get(
    upload_id: UUID,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_FEED_UPLOADS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)

    record = (
        (
            await session.execute(
                sa.select(KaspiFeedUpload).where(
                    KaspiFeedUpload.company_id == company_id,
                    KaspiFeedUpload.id == upload_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="upload_not_found")
    return _feed_upload_to_out(record)


async def kaspi_feed_upload_refresh(
    request: Request,
    upload_id: UUID,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_FEED_UPLOADS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)

    request_id = getattr(getattr(request, "state", None), "request_id", None) or request.headers.get("X-Request-ID")

    record = (
        (
            await session.execute(
                sa.select(KaspiFeedUpload).where(
                    KaspiFeedUpload.company_id == company_id,
                    KaspiFeedUpload.id == upload_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="upload_not_found")

    if not record.import_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="import_code_missing")

    store_name, token = await _resolve_kaspi_token(session, company_id)
    base_url = os.getenv("KASPI_FEED_BASE_URL", "https://kaspi.kz")
    upload_url = os.getenv("KASPI_FEED_UPLOAD_URL", f"{base_url.rstrip('/')}/shop/api/feeds/import")
    status_url = os.getenv("KASPI_FEED_STATUS_URL", f"{base_url.rstrip('/')}/shop/api/feeds/import/status")
    result_url = os.getenv("KASPI_FEED_RESULT_URL", f"{base_url.rstrip('/')}/shop/api/feeds/import/result")
    extra_env = {
        "KASPI_FEED_UPLOAD_URL": upload_url,
        "KASPI_FEED_STATUS_URL": status_url,
        "KASPI_FEED_RESULT_URL": result_url,
        "KASPI_FEED_TOKEN": token,
        "KASPI_TOKEN": token,
    }
    try:
        response = KaspiAdapter().feed_import_status(
            store_name,
            import_id=record.import_code,
            extra_env=extra_env,
        )
    except KaspiAdapterError as exc:
        await update_feed_upload_job(
            session,
            job=record,
            status="failed",
            error_code="kaspi_upstream_unavailable",
            error_message=str(exc)[:500],
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=record.merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code="kaspi_upstream_unavailable",
            error_message=str(exc)[:500],
            meta_json={"upload_id": str(record.id), "import_code": record.import_code},
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_unavailable") from exc
    except Exception as exc:
        logger.error("Kaspi feed status failed: company_id=%s error=%s", company_id, exc)
        await update_feed_upload_job(
            session,
            job=record,
            status="failed",
            error_code="kaspi_feed_status_failed",
            error_message=str(exc)[:500],
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=record.merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code="kaspi_feed_status_failed",
            error_message=str(exc)[:500],
            meta_json={"upload_id": str(record.id), "import_code": record.import_code},
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_feed_status_failed") from exc

    normalized = normalize_kaspi_payload(_normalize_kaspi_response(response))
    status_value = normalized.get("status") or record.status
    error_code, error_message = _extract_error_info(normalized)

    await update_feed_upload_job(
        session,
        job=record,
        status=str(status_value),
        error_code=error_code,
        error_message=error_message,
        response_json=normalized,
    )
    await _record_kaspi_event(
        session,
        company_id=company_id,
        merchant_uid=record.merchant_uid,
        kind="kaspi_feed",
        status="success",
        request_id=request_id,
        error_code=error_code,
        error_message=error_message,
        meta_json={"upload_id": str(record.id), "import_code": record.import_code},
    )
    return _feed_upload_to_out(record)


async def kaspi_feed_upload_refresh_compat(
    request: Request,
    upload_id: UUID,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_FEED_UPLOADS)),
    session: AsyncSession = Depends(get_async_db),
):
    return await kaspi_feed_upload_refresh(
        request=request,
        upload_id=upload_id,
        current_user=current_user,
        session=session,
    )


async def kaspi_feed_upload_publish(
    request: Request,
    upload_id: UUID,
    current_user: User = Depends(require_store_admin_then_feature(FEATURE_KASPI_FEED_UPLOADS)),
    session: AsyncSession = Depends(get_async_db),
):
    company_id = _resolve_company_id(current_user)
    request_id = getattr(getattr(request, "state", None), "request_id", None) or request.headers.get("X-Request-ID")

    record = (
        (
            await session.execute(
                sa.select(KaspiFeedUpload).where(
                    KaspiFeedUpload.company_id == company_id,
                    KaspiFeedUpload.id == upload_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="upload_not_found")

    if not record.import_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="import_code_missing")

    store_name, token = await _resolve_kaspi_token(session, company_id)
    base_url = os.getenv("KASPI_FEED_BASE_URL", "https://kaspi.kz")
    upload_url = os.getenv("KASPI_FEED_UPLOAD_URL", f"{base_url.rstrip('/')}/shop/api/feeds/import")
    status_url = os.getenv("KASPI_FEED_STATUS_URL", f"{base_url.rstrip('/')}/shop/api/feeds/import/status")
    result_url = os.getenv("KASPI_FEED_RESULT_URL", f"{base_url.rstrip('/')}/shop/api/feeds/import/result")
    extra_env = {
        "KASPI_FEED_UPLOAD_URL": upload_url,
        "KASPI_FEED_STATUS_URL": status_url,
        "KASPI_FEED_RESULT_URL": result_url,
        "KASPI_FEED_TOKEN": token,
        "KASPI_TOKEN": token,
    }

    try:
        response = KaspiAdapter().feed_import_status(
            store_name,
            import_id=record.import_code,
            extra_env=extra_env,
        )
    except KaspiAdapterError as exc:
        await update_feed_upload_job(
            session,
            job=record,
            status="failed",
            error_code="kaspi_upstream_unavailable",
            error_message=str(exc)[:500],
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=record.merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code="kaspi_upstream_unavailable",
            error_message=str(exc)[:500],
            meta_json={"upload_id": str(record.id), "import_code": record.import_code},
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_upstream_unavailable") from exc
    except Exception as exc:
        logger.error("Kaspi feed publish failed: company_id=%s error=%s", company_id, exc)
        await update_feed_upload_job(
            session,
            job=record,
            status="failed",
            error_code="kaspi_feed_publish_failed",
            error_message=str(exc)[:500],
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=record.merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code="kaspi_feed_publish_failed",
            error_message=str(exc)[:500],
            meta_json={"upload_id": str(record.id), "import_code": record.import_code},
        )
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="kaspi_feed_publish_failed") from exc

    normalized = normalize_kaspi_payload(_normalize_kaspi_response(response))
    status_value = str(normalized.get("status") or record.status)
    error_code, error_message = _extract_error_info(normalized)

    if status_value.lower() not in {"done", "success", "completed", "published"}:
        await update_feed_upload_job(
            session,
            job=record,
            status=status_value,
            error_code=error_code,
            error_message=error_message,
            response_json=normalized,
        )
        await _record_kaspi_event(
            session,
            company_id=company_id,
            merchant_uid=record.merchant_uid,
            kind="kaspi_feed",
            status="error",
            request_id=request_id,
            error_code="feed_not_ready_for_publish",
            error_message="feed_not_ready_for_publish",
            meta_json={"upload_id": str(record.id), "import_code": record.import_code},
        )
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="feed_not_ready_for_publish")

    await update_feed_upload_job(
        session,
        job=record,
        status="published",
        error_code=error_code,
        error_message=error_message,
    )
    await _record_kaspi_event(
        session,
        company_id=company_id,
        merchant_uid=record.merchant_uid,
        kind="kaspi_feed",
        status="success",
        request_id=request_id,
        error_code=error_code,
        error_message=error_message,
        meta_json={"upload_id": str(record.id), "import_code": record.import_code},
    )
    return _feed_upload_to_out(record)


async def kaspi_feed_export_detail(
    export_id: int,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    export = (
        (
            await session.execute(
                sa.select(KaspiFeedExport).where(
                    KaspiFeedExport.company_id == company_id,
                    KaspiFeedExport.id == export_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not export:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="export_not_found")
    return _feed_export_to_out(export)


async def kaspi_feed_export_download(
    export_id: int,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    export = (
        (
            await session.execute(
                sa.select(KaspiFeedExport).where(
                    KaspiFeedExport.company_id == company_id,
                    KaspiFeedExport.id == export_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not export:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="export_not_found")

    if export.status != "DONE" or not export.payload_text:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="export_not_ready")

    merchant_uid = None
    if isinstance(export.stats_json, dict):
        merchant_uid = export.stats_json.get("merchant_uid")
    safe_merchant = merchant_uid or "unknown"
    filename = f"kaspi_offers_{safe_merchant}_{export.id}.xml"

    return Response(
        content=export.payload_text,
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


register_kaspi_feed_routes_phase_one(
    router,
    kaspi_feed_export_create=kaspi_feed_export_create,
    kaspi_feed_exports_list=kaspi_feed_exports_list,
    kaspi_offers_feed_upload=kaspi_offers_feed_upload,
    kaspi_feed_upload_create=kaspi_feed_upload_create,
    kaspi_feed_uploads_list=kaspi_feed_uploads_list,
    kaspi_feed_upload_get=kaspi_feed_upload_get,
    kaspi_feed_upload_refresh=kaspi_feed_upload_refresh,
    kaspi_feed_upload_refresh_compat=kaspi_feed_upload_refresh_compat,
    kaspi_feed_upload_publish=kaspi_feed_upload_publish,
    kaspi_feed_export_detail=kaspi_feed_export_detail,
    kaspi_feed_export_download=kaspi_feed_export_download,
    kaspi_feed_export_out_model=KaspiFeedExportOut,
    kaspi_offers_feed_upload_out_model=KaspiOffersFeedUploadOut,
    kaspi_feed_upload_record_out_model=KaspiFeedUploadRecordOut,
)


# ============================= PUBLIC FEED TOKENS =============================


async def kaspi_feed_public_token_create(
    payload: KaspiFeedPublicTokenIn,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)
    env_is_dev = settings.is_development or settings.is_testing

    merchant_uid = (payload.merchant_uid or "").strip()
    if not merchant_uid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_merchant_uid")

    token_value = None
    token_hash = None
    for _ in range(3):
        candidate = secrets.token_urlsafe(32)
        candidate_hash = sha256(candidate.encode("utf-8")).hexdigest()
        exists = (
            await session.execute(
                sa.select(sa.func.count())
                .select_from(KaspiFeedPublicToken)
                .where(KaspiFeedPublicToken.token_hash == candidate_hash)
            )
        ).scalar_one()
        if not exists:
            token_value = candidate
            token_hash = candidate_hash
            break

    if not token_value or not token_hash:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="token_generation_failed")

    token_row = KaspiFeedPublicToken(
        company_id=company_id,
        merchant_uid=merchant_uid,
        token_hash=token_hash,
        comment=payload.comment,
    )
    session.add(token_row)
    await session.commit()
    await session.refresh(token_row)

    return KaspiFeedPublicTokenOut(
        id=token_row.id,
        merchant_uid=token_row.merchant_uid,
        token=token_value if env_is_dev else None,
        created_at=token_row.created_at,
        revoked_at=token_row.revoked_at,
        last_used_at=token_row.last_used_at,
        comment=token_row.comment,
    )


async def kaspi_feed_public_tokens_list(
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    result = await session.execute(
        sa.select(KaspiFeedPublicToken)
        .where(KaspiFeedPublicToken.company_id == company_id)
        .order_by(KaspiFeedPublicToken.created_at.desc())
    )
    tokens = result.scalars().all()
    return KaspiFeedPublicTokenListOut(
        items=[
            KaspiFeedPublicTokenOut(
                id=row.id,
                merchant_uid=row.merchant_uid,
                token=None,
                created_at=row.created_at,
                revoked_at=row.revoked_at,
                last_used_at=row.last_used_at,
                comment=row.comment,
            )
            for row in tokens
        ]
    )


async def kaspi_feed_public_token_revoke(
    token_id: int,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_async_db),
):
    await _require_store_admin_company_scoped(current_user)
    company_id = _resolve_company_id(current_user)

    token_row = (
        (
            await session.execute(
                sa.select(KaspiFeedPublicToken).where(
                    KaspiFeedPublicToken.company_id == company_id,
                    KaspiFeedPublicToken.id == token_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not token_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="token_not_found")

    token_row.revoked_at = datetime.utcnow()
    await session.commit()
    await session.refresh(token_row)

    return KaspiFeedPublicTokenOut(
        id=token_row.id,
        merchant_uid=token_row.merchant_uid,
        token=None,
        created_at=token_row.created_at,
        revoked_at=token_row.revoked_at,
        last_used_at=token_row.last_used_at,
        comment=token_row.comment,
    )


async def kaspi_public_price_list(
    request: Request,
    token: str,
    session: AsyncSession = Depends(get_async_db),
):
    await enforce_rate_limit(
        tag="kaspi_public_feed",
        ident=request.client.host if request.client else "0.0.0.0",
        max_requests=int(getattr(settings, "RATE_LIMIT_PER_MINUTE", 100) or 100),
        window_seconds=int(getattr(settings, "RATE_LIMIT_WINDOW_SECONDS", 60) or 60),
        detail="rate_limited",
    )
    token_row, xml_body, payload_hash, last_modified = await _load_public_offers_feed(
        session,
        token_value=token,
    )
    etag = _public_feed_etag(payload_hash)
    if_none_match = (request.headers.get("if-none-match") or "").strip()
    if_modified_since = _parse_http_datetime(request.headers.get("if-modified-since"))
    if if_none_match and if_none_match == etag:
        _log_public_feed_access(request=request, token_row=token_row, status_code=304, size_bytes=0)
        return Response(
            status_code=status.HTTP_304_NOT_MODIFIED,
            headers={"ETag": etag, "Last-Modified": _format_http_datetime(last_modified)},
        )
    if if_modified_since and if_modified_since >= last_modified:
        _log_public_feed_access(request=request, token_row=token_row, status_code=304, size_bytes=0)
        return Response(
            status_code=status.HTTP_304_NOT_MODIFIED,
            headers={"ETag": etag, "Last-Modified": _format_http_datetime(last_modified)},
        )

    token_row.last_used_at = datetime.utcnow()
    await session.commit()
    size_bytes = len(xml_body.encode("utf-8"))
    _log_public_feed_access(request=request, token_row=token_row, status_code=200, size_bytes=size_bytes)
    return Response(
        content=xml_body,
        media_type="application/xml; charset=utf-8",
        headers={
            "Cache-Control": "public, max-age=300",
            "ETag": etag,
            "Last-Modified": _format_http_datetime(last_modified),
        },
    )


async def kaspi_public_offers_feed(
    request: Request,
    token: str | None = None,
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    session: AsyncSession = Depends(get_async_db),
):
    if not token:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="not_found")

    await enforce_rate_limit(
        tag="kaspi_public_feed",
        ident=request.client.host if request.client else "0.0.0.0",
        max_requests=int(getattr(settings, "RATE_LIMIT_PER_MINUTE", 100) or 100),
        window_seconds=int(getattr(settings, "RATE_LIMIT_WINDOW_SECONDS", 60) or 60),
        detail="rate_limited",
    )
    token_row, xml_body, payload_hash, last_modified = await _load_public_offers_feed(
        session,
        token_value=token,
        merchant_uid=merchant_uid,
    )
    etag = _public_feed_etag(payload_hash)
    if_none_match = (request.headers.get("if-none-match") or "").strip()
    if_modified_since = _parse_http_datetime(request.headers.get("if-modified-since"))
    if if_none_match and if_none_match == etag:
        _log_public_feed_access(request=request, token_row=token_row, status_code=304, size_bytes=0)
        return Response(
            status_code=status.HTTP_304_NOT_MODIFIED,
            headers={"ETag": etag, "Last-Modified": _format_http_datetime(last_modified)},
        )
    if if_modified_since and if_modified_since >= last_modified:
        _log_public_feed_access(request=request, token_row=token_row, status_code=304, size_bytes=0)
        return Response(
            status_code=status.HTTP_304_NOT_MODIFIED,
            headers={"ETag": etag, "Last-Modified": _format_http_datetime(last_modified)},
        )

    token_row.last_used_at = datetime.utcnow()
    await session.commit()
    size_bytes = len(xml_body.encode("utf-8"))
    _log_public_feed_access(request=request, token_row=token_row, status_code=200, size_bytes=size_bytes)
    return Response(
        content=xml_body,
        media_type="application/xml; charset=utf-8",
        headers={
            "Cache-Control": "public, max-age=300",
            "ETag": etag,
            "Last-Modified": _format_http_datetime(last_modified),
        },
    )


register_kaspi_public_routes(
    router,
    public_router,
    kaspi_feed_public_token_create=kaspi_feed_public_token_create,
    kaspi_feed_public_tokens_list=kaspi_feed_public_tokens_list,
    kaspi_feed_public_token_revoke=kaspi_feed_public_token_revoke,
    kaspi_public_price_list=kaspi_public_price_list,
    kaspi_public_offers_feed=kaspi_public_offers_feed,
    kaspi_feed_public_token_out_model=KaspiFeedPublicTokenOut,
    kaspi_feed_public_token_list_out_model=KaspiFeedPublicTokenListOut,
)


async def kaspi_feed_generate_products(
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    """
    Генерирует фид продуктов для текущей компании.
    Идемпотентен: повторный вызов вернёт существующий фид если контент не изменился.
    """
    from app.services.kaspi_feed_export_service import generate_products_feed

    company_id = _resolve_company_id(current_user)

    try:
        result = await generate_products_feed(session, company_id)
        return KaspiFeedGenerateOut(**result)
    except Exception as e:
        logger.error("Kaspi feed generation failed: company_id=%s error=%s", company_id, e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate feed",
        )


async def kaspi_feed_upload(
    export_id: int,
    request: Request,
    merchant_uid: str | None = Query(None, alias="merchantUid"),
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    """
    Загружает фид на Kaspi. Текущий пользователь должен иметь доступ к компании фида.
    """
    company_id = _resolve_company_id(current_user)

    export = (
        (
            await session.execute(
                sa.select(KaspiFeedExport).where(
                    KaspiFeedExport.id == export_id,
                    KaspiFeedExport.company_id == company_id,
                )
            )
        )
        .scalars()
        .first()
    )
    if not export:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="export_not_found")

    export_merchant_uid = None
    if isinstance(export.stats_json, dict):
        export_merchant_uid = export.stats_json.get("merchant_uid")
    merchant_uid = (merchant_uid or export_merchant_uid or "").strip()
    if not merchant_uid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="missing_merchant_uid")

    terminal_statuses = {
        "done",
        "failed",
        "uploaded",
        "published",
        "success",
        "completed",
        "rejected",
        "invalid",
        "error",
    }

    existing = (
        (
            await session.execute(
                sa.select(KaspiFeedUpload)
                .where(
                    KaspiFeedUpload.company_id == company_id,
                    KaspiFeedUpload.merchant_uid == merchant_uid,
                    KaspiFeedUpload.export_id == export_id,
                    sa.or_(
                        KaspiFeedUpload.status.is_(None),
                        sa.func.lower(KaspiFeedUpload.status).not_in(terminal_statuses),
                    ),
                )
                .order_by(KaspiFeedUpload.created_at.desc())
                .limit(1)
            )
        )
        .scalars()
        .first()
    )

    if existing:
        rid = getattr(getattr(request, "state", None), "request_id", None) or request.headers.get("X-Request-ID")
        if not rid:
            rid = str(uuid4())
        payload = {
            "detail": "upload_not_claimable",
            "code": "HTTP_409",
            "request_id": rid,
            "existing_upload_id": str(existing.id),
            "status": existing.status,
        }
        return JSONResponse(status_code=status.HTTP_409_CONFLICT, content=payload, headers={"X-Request-ID": rid})

    body = KaspiFeedUploadIn(
        merchant_uid=merchant_uid,
        source="export_id",
        export_id=str(export_id),
    )
    return await kaspi_feed_upload_create(
        request=request,
        body=body,
        current_user=current_user,
        session=session,
    )


async def kaspi_feeds_list(
    kind: str | None = None,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    """
    Возвращает список фидов для текущей компании с опциональной фильтрацией по kind.
    """
    from app.models.kaspi_feed_export import KaspiFeedExport

    company_id = _resolve_company_id(current_user)

    # Validate limit
    limit = max(1, min(limit, 200))
    offset = max(0, offset)

    try:
        # Build query
        query = sa.select(KaspiFeedExport).where(KaspiFeedExport.company_id == company_id)

        # Optional filter by kind
        if kind:
            query = query.where(KaspiFeedExport.kind == kind)

        # Count total
        count_query = sa.select(sa.func.count()).select_from(query.subquery())
        total_result = await session.execute(count_query)
        total = total_result.scalar() or 0

        # Apply pagination and ordering
        query = query.order_by(KaspiFeedExport.created_at.desc()).limit(limit).offset(offset)

        # Execute
        result = await session.execute(query)
        exports = result.scalars().all()

        # Map to response models
        items = [
            KaspiFeedExportOut(
                id=e.id,
                kind=e.kind,
                format=e.format,
                status=e.status,
                checksum=e.checksum,
                stats_json=e.stats_json,
                last_error=e.last_error,
                attempts=e.attempts or 0,
                last_attempt_at=e.last_attempt_at.isoformat() if e.last_attempt_at else None,
                uploaded_at=e.uploaded_at.isoformat() if e.uploaded_at else None,
                duration_ms=e.duration_ms,
                created_at=e.created_at.isoformat() if e.created_at else None,
                updated_at=e.updated_at.isoformat() if e.updated_at else None,
            )
            for e in exports
        ]

        return KaspiFeedListOut(
            items=items,
            total=total,
            limit=limit,
            offset=offset,
        )

    except Exception as e:
        logger.error("Kaspi feeds list failed: company_id=%s error=%s", company_id, e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve feeds",
        )


async def kaspi_feed_get(
    export_id: int,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    """
    Возвращает метаданные фида (без payload).
    """
    from app.models.kaspi_feed_export import KaspiFeedExport

    company_id = _resolve_company_id(current_user)

    try:
        stmt = sa.select(KaspiFeedExport).where(
            sa.and_(
                KaspiFeedExport.id == export_id,
                KaspiFeedExport.company_id == company_id,
            )
        )
        result = await session.execute(stmt)
        export = result.scalars().first()

        if not export:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Export not found",
            )

        return KaspiFeedExportOut(
            id=export.id,
            kind=export.kind,
            format=export.format,
            status=export.status,
            checksum=export.checksum,
            stats_json=export.stats_json,
            last_error=export.last_error,
            attempts=export.attempts or 0,
            last_attempt_at=export.last_attempt_at.isoformat() if export.last_attempt_at else None,
            uploaded_at=export.uploaded_at.isoformat() if export.uploaded_at else None,
            duration_ms=export.duration_ms,
            created_at=export.created_at.isoformat() if export.created_at else None,
            updated_at=export.updated_at.isoformat() if export.updated_at else None,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Kaspi feed get failed: export_id=%s company_id=%s error=%s", export_id, company_id, e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve feed",
        )


async def kaspi_feed_get_payload(
    export_id: int,
    current_user: User = Depends(_auth_user),
    session: AsyncSession = Depends(get_async_db),
):
    """
    Возвращает XML payload фида с типом application/xml.
    """
    from app.models.kaspi_feed_export import KaspiFeedExport

    company_id = _resolve_company_id(current_user)

    try:
        stmt = sa.select(KaspiFeedExport).where(
            sa.and_(
                KaspiFeedExport.id == export_id,
                KaspiFeedExport.company_id == company_id,
            )
        )
        result = await session.execute(stmt)
        export = result.scalars().first()

        if not export:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Export not found",
            )

        return Response(
            content=export.payload_text,
            media_type="application/xml",
            headers={"Content-Disposition": f'attachment; filename="kaspi_feed_{export_id}.xml"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Kaspi feed payload failed: export_id=%s company_id=%s error=%s", export_id, company_id, e)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve payload",
        )


register_kaspi_feed_routes_phase_two(
    router,
    kaspi_feed_generate_products=kaspi_feed_generate_products,
    kaspi_feed_upload=kaspi_feed_upload,
    kaspi_feeds_list=kaspi_feeds_list,
    kaspi_feed_get=kaspi_feed_get,
    kaspi_feed_get_payload=kaspi_feed_get_payload,
    kaspi_feed_generate_out_model=KaspiFeedGenerateOut,
    kaspi_feed_upload_record_out_model=KaspiFeedUploadRecordOut,
    kaspi_feed_list_out_model=KaspiFeedListOut,
    kaspi_feed_export_out_model=KaspiFeedExportOut,
)


register_kaspi_status_routes(
    router,
    auth_dependency=_auth_user,
    get_async_db_dependency=get_async_db,
    resolve_company_id_fn=_resolve_company_id,
    status_last_error_max_len=STATUS_LAST_ERROR_MAX_LEN,
    logger=logger,
    company_model=Company,
    kaspi_feed_export_model=KaspiFeedExport,
    kaspi_catalog_product_model=KaspiCatalogProduct,
    kaspi_order_sync_state_model=KaspiOrderSyncState,
    kaspi_store_token_model=KaspiStoreToken,
)
